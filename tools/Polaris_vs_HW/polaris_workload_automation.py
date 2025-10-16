#!/usr/bin/env python3
"""
Polaris Workload Automation Script

This script automates the process of:
1. Pulling the latest Polaris repository
2. Running all workloads with various configurations
3. Collecting and aggregating metrics
4. Generating comprehensive Excel reports

"""

import argparse
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import subprocess
import shutil
import urllib.request
import json
import csv
from datetime import datetime
import tempfile

# Import common utilities
from common_utils import (
    pull_git_repo, create_excel_workbook, setup_excel_headers,
    apply_cell_formatting, adjust_column_widths, save_excel_workbook,
    ensure_directory, archive_directory, create_orderly_output_structure, EXCEL_COLORS
)

# Third-party imports
try:
    import git
    import yaml
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, Reference
    import matplotlib.pyplot as plt
    import seaborn as sns
except ImportError as e:
    print(f"Missing required dependencies: {e}")
    print("Please install required packages:")
    print("pip install GitPython PyYAML pandas openpyxl matplotlib seaborn")
    sys.exit(1)

# Constants
POLARIS_REPO_URL = "https://github.com/tenstorrent/polaris"
DEFAULT_CONFIG_FILES = {
    'workloads': ['config/all_workloads.yaml', 'config/ip_workloads.yaml'],
    'mapping': 'config/wl2archmapping.yaml'
}

# Logging will be configured after directory structure is set up
logger = logging.getLogger(__name__)


@dataclass
class ExecutionConfig:
    """Configuration for a single workload execution"""
    workload_name: str
    workload_instance: str
    architecture: str
    batch_size: int = 1
    frequency_mhz: int = 1000
    study_name: str = ""
    output_dir: str = ""
    arch_config: str = "config/all_archs.yaml"
    workload_config: str = "config/all_workloads.yaml"

    def get_command_args(self) -> List[str]:
        """Generate command line arguments for polaris execution"""
        args = [
            "--archspec", str(self.arch_config),
            "--wlspec", str(self.workload_config),
            "--wlmapspec", f"config/wl2archmapping.yaml",
            "--study", str(self.study_name),
            "--odir", str(self.output_dir),
            "--filterwl", str(self.workload_name),
            "--filterwli", str(self.workload_instance),
            "--filterarch", str(self.architecture)
        ]

        # Only add batchsize if it's different from default (1)
        # Polaris doesn't allow batchsize ranges where start == end
        if self.batch_size != 1:
            # For llama2, use a proper batchsize range to avoid precision issues
            if "llama2" in self.study_name:
                args.extend(["--batchsize", str(self.batch_size), str(self.batch_size * 2), "2"])
            else:
                args.extend(["--batchsize", str(self.batch_size), str(self.batch_size), "1"])

        # Only add frequency if it's different from default (1000 MHz)
        # Polaris doesn't allow frequency ranges where start == end
        if self.frequency_mhz != 1000:
            # For llama2, use a proper frequency range to avoid precision issues
            if "llama2" in self.study_name:
                args.extend(["--frequency", str(self.frequency_mhz), str(self.frequency_mhz + 100), "200"])
            else:
                args.extend(["--frequency", str(self.frequency_mhz), str(self.frequency_mhz), "100"])

        return args


@dataclass
class ExecutionResult:
    """Result of a single workload execution"""
    config: ExecutionConfig
    success: bool
    execution_time: float
    error_message: Optional[str] = None
    summary_csv_path: Optional[str] = None
    detailed_json_path: Optional[str] = None
    metrics: Dict[str, Any] = field(default_factory=dict)
    stdout: str = ""
    stderr: str = ""


class PolarisAutomation:
    """Main class for Polaris workload automation"""

    def __init__(self, repo_path: str = ".", clean_repo: bool = False, clean_results: bool = True):
        """
        Initialize the automation system

        Args:
            repo_path: Path to clone/pull the Polaris repository
            clean_repo: Whether to clean the repository before operations
        """
        self.repo_path = Path(repo_path).absolute()
        self.clean_repo = clean_repo
        self.clean_results = clean_results
        self.repo: Optional[git.Repo] = None

        # Don't create the directory here - let setup_repository handle it

        logger.info(f"Initialized Polaris automation with repo path: {self.repo_path}")

    def setup_repository(self, clean_repo: bool = False) -> bool:
        """
        Setup the Polaris repository - clone or pull latest changes.

        Args:
            clean_repo: If True, remove existing repo and re-clone

        Returns:
            bool: True if successful
        """
        # If repo_path is a subdirectory inside the polaris repo (e.g., tools/Polaris_vs_HW),
        # resolve to the actual repo root containing .git
        try:
            path = self.repo_path
            for _ in range(5):  # limit ascent
                if (path / ".git").exists():
                    if path != self.repo_path:
                        logger.info(f"Resolved Polaris repo root at: {path}")
                        self.repo_path = path
                    break
                if path.parent == path:
                    break
                path = path.parent
        except Exception:
            pass

        success = pull_git_repo("tenstorrent/polaris", self.repo_path, clean_repo)

        if success:
            # Additional verification for Polaris-specific files
            if not (self.repo_path / "polaris.py").exists():
                logger.error("Invalid repository - polaris.py not found")
                return False

        return success

    def get_repository_info(self) -> Dict[str, str]:
        """Get information about the current repository state"""
        if not self.repo:
            return {}

        try:
            head = self.repo.head.commit
            return {
                'commit_hash': head.hexsha,
                'commit_message': head.message.strip(),
                'author': head.author.name,
                'date': head.committed_datetime.isoformat(),
                'branch': self.repo.active_branch.name if not self.repo.head.is_detached else 'detached',
                'is_dirty': self.repo.is_dirty()
            }
        except Exception as e:
            logger.warning(f"Could not get repository info: {e}")
            return {}

    def load_configurations(self, arch_configs: List[str], workload_configs: Optional[List[str]]) -> Tuple[Dict[str, Any], Dict[str, Any], List[str]]:
        """
        Load architecture and workload configurations from YAML files.
        
        Args:
            arch_configs: List of architecture config file paths
            workload_configs: List of workload config file paths
            
        Returns:
            Tuple of (arch_config, workload_config, target_architectures)
        """
        logger.info(f"Loading architecture configs: {arch_configs}")
        logger.info(f"Loading workload configs: {workload_configs}")
        
        arch_config = {}
        workload_config = {}
        target_architectures = []
        
        # Load architecture configurations
        for config_file in arch_configs:
            try:
                config_path = self.repo_path / config_file
                if config_path.exists():
                    with open(config_path, 'r') as f:
                        data = yaml.safe_load(f)
                        arch_config.update(data or {})
                    logger.info(f"Loaded architecture config: {config_file}")
                else:
                    logger.warning(f"Architecture config not found: {config_path}")
            except Exception as e:
                logger.warning(f"Could not load architecture config {config_file}: {e}")
        
        # Load workload configurations
        if workload_configs:
            for config_file in workload_configs:
                try:
                    # Handle both relative paths (within repo) and absolute paths (local configs)
                    if Path(config_file).is_absolute():
                        config_path = Path(config_file)
                    else:
                        config_path = self.repo_path / config_file

                    if config_path.exists():
                        with open(config_path, 'r') as f:
                            data = yaml.safe_load(f)
                            workload_config[config_file] = data or {}
                        logger.info(f"Loaded workload config: {config_file}")
                    else:
                        logger.warning(f"Workload config not found: {config_path}")
                except Exception as e:
                    logger.warning(f"Could not load workload config {config_file}: {e}")
        else:
            # Default workload configs - check for local wh_supported.yaml alongside scripts
            script_dir = Path(__file__).parent
            local_config = script_dir / 'wh_supported.yaml'
            if local_config.exists():
                try:
                    with open(local_config, 'r') as f:
                        data = yaml.safe_load(f)
                        workload_config[str(local_config)] = data or {}
                    logger.info(f"Loaded default local workload config: {local_config}")
                except Exception as e:
                    logger.warning(f"Could not load default local workload config {local_config}: {e}")
            else:
                # Fallback to polaris repo configs
                default_configs = ['config/all_workloads.yaml', 'config/ip_workloads.yaml']
                for config_file in default_configs:
                    try:
                        config_path = self.repo_path / config_file
                        if config_path.exists():
                            with open(config_path, 'r') as f:
                                data = yaml.safe_load(f)
                                workload_config[config_file] = data or {}
                            logger.info(f"Loaded fallback workload config: {config_file}")
                        else:
                            logger.warning(f"Fallback workload config not found: {config_path}")
                    except Exception as e:
                        logger.warning(f"Could not load fallback workload config {config_file}: {e}")
        
        # Download external configs if any config has yaml_cfg_path
        has_external = False
        all_workload_groups = []
        for data in workload_config.values():
            if 'workloads' in data:
                all_workload_groups.extend(data['workloads'])
                for wl_group in data['workloads']:
                    if 'instances' in wl_group:
                        for instance in wl_group['instances'].values():
                            if 'yaml_cfg_path' in instance:
                                has_external = True
                                break
                        if has_external:
                            break
            if has_external:
                break
        
        if has_external:
            logger.info("External configs detected, downloading...")
            self._download_external_configs(all_workload_groups)
        else:
            logger.info("No external configs needed")
        
        # Extract target architectures from arch config
        for pkg in arch_config.get('packages', []):
            for instance in pkg.get('instances', []):
                instance_name = instance.get('name', '')
                target_architectures.append(instance_name)

        if not target_architectures:
            target_architectures = ['n150', 'n300']
            logger.info(f"No target architectures found in config, using defaults: {target_architectures}")

        logger.info(f"Loaded {len(arch_config)} architecture packages, {len(workload_config)} workload configs")
        logger.info(f"Target architectures: {target_architectures}")

        return arch_config, workload_config, target_architectures

    def discover_workloads(self, target_architectures: Optional[List[str]] = None, arch_config: Dict[str, Any] = None, workload_configs: List[str] = None, run_type: str = "default") -> List[ExecutionConfig]:
        """
        Discover workload/architecture combinations, optionally filtered by target architectures

        Args:
            target_architectures: Optional list of architecture names to filter by.
            arch_config: Path to architecture configuration file.
            workload_configs: Optional list of workload configuration file paths.
            run_type: Type of run for unique output directories.

        Returns:
            List of ExecutionConfig objects for the combinations
        """
        architectures_config = arch_config
        workloads_config = self.load_workload_configs(workload_configs)

        # Download external config files if needed
        self._download_external_configs(workloads_config)

        execution_configs = []
        all_architectures = self._extract_architectures(architectures_config)

        # Filter architectures if specified
        if target_architectures:
            architectures = [arch for arch in all_architectures if arch in target_architectures]
            if not architectures:
                logger.warning(f"No matching architectures found for filter: {target_architectures}")
                logger.info(f"Available architectures: {all_architectures}")
                return []
        else:
            architectures = all_architectures

        # Determine which workload config file each workload comes from
        workload_to_config_map = {}
        # Always build the map from default configs
        default_configs = ['config/all_workloads.yaml', 'config/ip_workloads.yaml']
        for config_path in default_configs:
            try:
                config_file = self.repo_path / config_path
                if config_file.exists():
                    with open(config_file, 'r') as f:
                        config_data = yaml.safe_load(f)
                    if 'workloads' in config_data:
                        for workload_group in config_data['workloads']:
                            workload_name = workload_group.get('name', '')
                            if workload_name:  # Avoid empty names
                                workload_to_config_map[workload_name] = config_path
                    logger.debug(f"Built map from {config_path}: added {len([k for k in workload_to_config_map if workload_to_config_map[k] == config_path])} workloads")
            except Exception as e:
                logger.warning(f"Could not build map from {config_path}: {e}")

        if workload_configs:
            # Override/add from specified configs
            for config_file in workload_configs:
                try:
                    with open(Path(self.repo_path) / config_file, 'r') as f:
                        config_data = yaml.safe_load(f)
                    if 'workloads' in config_data:
                        for workload_group in config_data['workloads']:
                            workload_name = workload_group.get('name', '')
                            workload_to_config_map[workload_name] = config_file
                except Exception as e:
                    logger.warning(f"Could not read workload config {config_file}: {e}")

        logger.debug(f"Final workload_to_config_map: {list(workload_to_config_map.keys())}")

        # Process workloads
        for workload_group in workloads_config.get('workloads', []):
            api = workload_group.get('api', 'TTSIM')
            workload_name = workload_group.get('name', '')
            basedir = workload_group.get('basedir', 'workloads')

            # Get the config file for this workload
            workload_config_file = workload_to_config_map.get(workload_name, "config/all_workloads.yaml")

            # Get workload instances
            for instance_name, instance_config in workload_group.get('instances', {}).items():
                for arch in architectures:
                    # Special handling for llama2 workloads to avoid TTSIM precision issues
                    batch_size = 1
                    frequency_mhz = 1000

                    if workload_name == "llama2":
                        # Try different parameters for llama2 to avoid TTSIM precision issues
                        batch_size = 2  # Try batch size 2
                        frequency_mhz = 1000  # Keep default frequency

                    config = ExecutionConfig(
                        workload_name=workload_name,
                        workload_instance=instance_name,
                        architecture=arch,
                        batch_size=batch_size,
                        frequency_mhz=frequency_mhz,
                        study_name=f"{workload_name}_{instance_name}_{arch}_{run_type}",
                        output_dir=f"results/{workload_name}_{instance_name}_{arch}_{run_type}",
                        arch_config='config/tt_wh.yaml',
                        workload_config=workload_config_file
                    )
                    execution_configs.append(config)

        logger.info(f"Discovered {len(execution_configs)} workload/architecture combinations")
        return execution_configs

    def load_workload_configs(self, workload_config_paths: List[str]) -> Dict[str, Any]:
        workloads_config = {'workloads': []}
        for workload_path in workload_config_paths:
            workload_file = Path(self.repo_path) / workload_path
            if workload_file.exists():
                with open(workload_file, 'r') as f:
                    config = yaml.safe_load(f)
                    if 'workloads' in config:
                        workloads_config['workloads'].extend(config['workloads'])
        return workloads_config

    def load_mapping_config(self) -> Dict[str, Any]:
        config_dir = self.repo_path / "config"
        with open(config_dir / "wl2archmapping.yaml", 'r') as f:
            return yaml.safe_load(f)

    def _download_external_configs(self, workload_groups: List[Dict]) -> None:
        """
        Download external configuration files referenced in workload instances.
        
        Args:
            workload_groups: List of workload group dictionaries
        """
        external_dir = self.repo_path / "config" / "external_configs"
        external_dir.mkdir(parents=True, exist_ok=True)
        
        updated_paths = []
        
        for group in workload_groups:
            if 'instances' in group:
                for instance_name, instance_config in group['instances'].items():
                    yaml_cfg_path = instance_config.get('yaml_cfg_path')
                    if yaml_cfg_path:
                        # Check if it's a remote URL
                        if yaml_cfg_path.startswith("http"):
                            # Download remote config
                            try:
                                filename = yaml_cfg_path.split('/')[-1]
                                local_path = external_dir / filename

                                if not local_path.exists():
                                    logger.info(f"Downloading instance config: {yaml_cfg_path}")
                                    urllib.request.urlretrieve(yaml_cfg_path, local_path)

                                    # Update the config to use local path
                                    instance_config['yaml_cfg_path'] = f"config/external_configs/{filename}"
                                    updated_paths.append(filename)
                                    logger.info(f"Saved instance config locally: config/external_configs/{filename}")
                                else:
                                    instance_config['yaml_cfg_path'] = f"config/external_configs/{filename}"
                                    logger.debug(f"External config already exists: config/external_configs/{filename}")
                            except Exception as e:
                                logger.warning(f"Failed to download {yaml_cfg_path}: {e}")
                        # Check if it's a local path that doesn't exist but has a known remote
                        elif not (self.repo_path / yaml_cfg_path).exists():
                            # Hardcoded mapping for known external YOLO configs
                            filename = Path(yaml_cfg_path).name
                            remote_url = None

                            if "yolov8" in filename:
                                yolo_v8_base_url = "https://raw.githubusercontent.com/autogyro/yolo-V8/refs/heads/main/ultralytics/models/v8/"
                                remote_url = yolo_v8_base_url + filename
                            elif "yolov7" in filename:
                                yolo_v7_base_url = "https://raw.githubusercontent.com/WongKinYiu/yolov7/refs/heads/main/cfg/deploy/"
                                remote_url = yolo_v7_base_url + filename

                            if remote_url:
                                local_path = external_dir / filename
                                try:
                                    if not local_path.exists():
                                        logger.info(f"Downloading instance config: {remote_url}")
                                        urllib.request.urlretrieve(remote_url, local_path)
                                        instance_config['yaml_cfg_path'] = f"config/external_configs/{filename}"
                                        updated_paths.append(filename)
                                        logger.info(f"Saved instance config locally: config/external_configs/{filename}")
                                    else:
                                        instance_config['yaml_cfg_path'] = f"config/external_configs/{filename}"
                                        logger.debug(f"External config already exists: config/external_configs/{filename}")
                                except Exception as e:
                                    logger.warning(f"Failed to download {remote_url}: {e}")
                            else:
                                logger.warning(f"Local external config not found and no known remote: {yaml_cfg_path}")
        
        if updated_paths:
            logger.info(f"Updated {len(updated_paths)} external config paths")
        else:
            logger.info("No external configs to download")

    def _extract_architectures(self, arch_config: Dict) -> List[str]:
        """Extract architecture names from architecture configuration"""
        architectures = []

        # Extract from packages
        for package in arch_config.get('packages', []):
            package_name = package.get('name', '')
            for instance in package.get('instances', []):
                instance_name = instance.get('name', '')
                architectures.append(instance_name)

        return architectures

    def _check_polaris_success(self, config: ExecutionConfig, result: ExecutionResult) -> bool:
        """
        Check if Polaris actually produced meaningful results.
        Polaris can return exit code 0 but still fail to run experiments.
        We check for the presence of non-empty result files.
        """
        try:
            study_name = config.study_name

            # Try multiple possible directory structures (Polaris can create different structures)
            possible_dirs = [
                Path("results") / study_name / study_name,  # automation script structure
                Path("results") / study_name,               # direct run structure
                Path(config.output_dir) / study_name,       # custom output dir with study name
                Path(config.output_dir),                    # direct output dir
            ]

            summary_json = None
            stats_files = []

            for base_dir in possible_dirs:
                # Look for SUMMARY directory
                summary_file = base_dir / "SUMMARY" / "study-summary.json"
                if summary_file.exists():
                    summary_json = summary_file
                    break

                # Look for study-summary.json directly
                direct_summary = base_dir / "study-summary.json"
                if direct_summary.exists():
                    summary_json = direct_summary
                    break

                # Look for STATS files
                stats_dir = base_dir / "STATS"
                if stats_dir.exists():
                    json_files = list(stats_dir.glob("*.json"))
                    if json_files:
                        stats_files.extend(json_files)

            # Check summary file
            if summary_json:
                try:
                    with open(summary_json, 'r') as f:
                        data = json.load(f)
                        if data.get('summary') and len(data['summary']) > 0:
                            result.summary_csv_path = str(summary_json)
                            logger.debug(f"Found valid summary at: {summary_json}")
                            return True
                except (json.JSONDecodeError, KeyError) as e:
                    logger.debug(f"Error reading summary file {summary_json}: {e}")

            # Check stats files as fallback
            for json_file in stats_files:
                try:
                    with open(json_file, 'r') as f:
                        data = json.load(f)
                        if data and len(str(data)) > 100:  # More substantial check
                            result.detailed_json_path = str(json_file)
                            logger.debug(f"Found valid stats file at: {json_file}")
                            return True
                except (json.JSONDecodeError, IOError) as e:
                    logger.debug(f"Error reading stats file {json_file}: {e}")
                    continue

            # Debug: log what directories we checked
            logger.debug(f"No valid results found for {config.workload_name}.{config.workload_instance}. Checked dirs: {[str(d) for d in possible_dirs]}")
            return False

        except Exception as e:
            logger.debug(f"Error checking Polaris success: {e}")
            return False

    def execute_workload(self, config: ExecutionConfig) -> ExecutionResult:
        """
        Execute a single workload configuration

        Args:
            config: Execution configuration

        Returns:
            ExecutionResult with metrics and status
        """
        result = ExecutionResult(config=config, success=False, execution_time=0.0)

        try:
            logger.info(f"Executing workload: {config.workload_name}.{config.workload_instance} on {config.architecture}")

            # Change to repository directory
            original_cwd = os.getcwd()
            os.chdir(self.repo_path)

            # Prepare command
            cmd = [sys.executable, "polaris.py"] + config.get_command_args()
            logger.info(f"Running Polaris command: {' '.join(cmd)}")
            logger.info(f"Current working directory: {os.getcwd()}")

            # Execute
            start_time = time.time()
            process = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=3600  # 1 hour timeout
            )
            end_time = time.time()

            result.execution_time = end_time - start_time
            result.stdout = process.stdout
            result.stderr = process.stderr

            if process.returncode == 0:
                # Check if Polaris actually produced meaningful results
                # Polaris can return 0 but still fail to run experiments
                result.success = self._check_polaris_success(config, result)

                if result.success:
                    logger.info(f"Successfully executed {config.workload_name}.{config.workload_instance}")
                else:
                    logger.warning(f"Polaris returned success but no meaningful results for {config.workload_name}.{config.workload_instance}")
                    result.error_message = "Polaris completed but produced no results"

                # Load metrics if available (even if marked as failed, might have partial results)
                result.metrics = self._extract_metrics(result)

            else:
                result.error_message = process.stderr
                logger.error(f"Failed to execute {config.workload_name}.{config.workload_instance}: {process.stderr}")

        except subprocess.TimeoutExpired:
            result.error_message = "Execution timed out"
            logger.error(f"Execution timed out for {config.workload_name}.{config.workload_instance}")
        except Exception as e:
            result.error_message = str(e)
            logger.error(f"Unexpected error executing {config.workload_name}.{config.workload_instance}: {e}")
        finally:
            os.chdir(original_cwd)

        return result

    def execute_workloads(self, configs: List[ExecutionConfig]) -> List[ExecutionResult]:
        """
        Execute a list of workload configurations.

        Args:
            configs: List of ExecutionConfig objects

        Returns:
            List of ExecutionResult objects
        """
        results = []
        for config in configs:
            result = self.execute_workload(config)
            results.append(result)
        return results

    def _extract_metrics(self, result: ExecutionResult) -> Dict[str, Any]:
        """Extract metrics from result files (JSON format)"""
        metrics = {}

        try:
            # Load summary JSON
            if result.summary_csv_path and Path(result.summary_csv_path).exists():
                with open(result.summary_csv_path, 'r') as f:
                    summary_data = json.load(f)

                # Extract metrics from summary data
                summary_list = summary_data.get('summary', [])
                if summary_list:
                    # Take the first (and typically only) summary entry
                    summary = summary_list[0]
                    metrics.update({
                        'frequency_mhz': summary.get('freq_MHz', summary.get('frequency_MHz', summary.get('freq', 1000))),
                        'total_cycles': summary.get('tot_cycles', summary.get('cycles', summary.get('total_cycles', 0))),
                        'total_msecs': summary.get('tot_msecs', summary.get('msecs', summary.get('total_msecs', summary.get('time_ms', 0)))),
                        'ideal_throughput': summary.get('ideal_throughput', 0),
                        'memory_size_gb': summary.get('mem_size_GB', summary.get('memory_size', summary.get('mem_size', 0))),
                        'device_memory_gb': summary.get('device_mem_size_GB', summary.get('device_memory', 0)),
                        'fits_device': summary.get('fits_device', False),
                        'performance_projection': summary.get('perf_projection', summary.get('performance', 0)),
                        'resource_compute': summary.get('rsrc_comp', summary.get('compute_resource', 0)),
                        'resource_memory': summary.get('rsrc_mem', summary.get('memory_resource', 0)),
                        'batch_size': summary.get('batch', summary.get('batch_size', 1))
                    })

            # Load detailed JSON if available (STATS files)
            if result.detailed_json_path and Path(result.detailed_json_path).exists():
                with open(result.detailed_json_path, 'r') as f:
                    detailed_data = json.load(f)
                    metrics['operator_count'] = len(detailed_data.get('operatorstats', []))
                    # Override batch_size if available in detailed data
                    if 'batch' in detailed_data:
                        metrics['batch_size'] = detailed_data['batch']

        except Exception as e:
            logger.warning(f"Could not extract metrics: {e}")

        return metrics

    def run_all_workloads(self, max_parallel: int = 1, dry_run: bool = False, target_architectures: Optional[List[str]] = None, arch_config: str = "config/all_archs.yaml", workload_configs: List[str] = None) -> List[ExecutionResult]:
        """
        Run all discovered workloads

        Args:
            max_parallel: Maximum number of parallel executions
            dry_run: If True, only show what would be executed
            target_architectures: Optional list of architectures to filter by
            arch_config: Path to architecture configuration file

        Returns:
            List of execution results
        """
        configs = self.discover_workloads(target_architectures, arch_config, workload_configs)
        results = []

        logger.info(f"Starting execution of {len(configs)} workload configurations")

        if dry_run:
            logger.info("DRY RUN - Would execute the following configurations:")
            for config in configs[:5]:  # Show first 5
                logger.info(f"  - {config.workload_name}.{config.workload_instance} on {config.architecture}")
            if len(configs) > 5:
                logger.info(f"  ... and {len(configs) - 5} more")
            return []

        for i, config in enumerate(configs):
            logger.info(f"Progress: {i+1}/{len(configs)} - Executing {config.workload_name}.{config.workload_instance}")

            result = self.execute_workload(config)
            results.append(result)

            # Log progress
            success_count = sum(1 for r in results if r.success)
            logger.info(f"Completed: {success_count}/{len(results)} successful")

        logger.info(f"Execution completed. {sum(1 for r in results if r.success)}/{len(results)} successful")
        return results

    def generate_excel_report(self, results: List[ExecutionResult], output_path: Path) -> bool:
        """
        Generate comprehensive Excel report from execution results

        Args:
            results: List of execution results
            output_path: Path to save the Excel file

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"Generating Excel report: {output_path}")

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create consolidated worksheet
            self._create_consolidated_sheet(wb, results)

            # Save workbook
            wb.save(str(output_path))
            logger.info(f"Excel report saved successfully: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return False

    def save_detailed_outputs(self, results: List[ExecutionResult], output_dir: Path) -> None:
        """
        Save detailed run logs for each execution, including command, stdout, and stderr.
        
        Args:
            results: List of execution results
            output_dir: Directory to save the log files
        """
        logger.info(f"Saving detailed run outputs to {output_dir}")
        
        for result in results:
            log_content = f"""POLARIS RUN LOG
====================

Workload: {result.config.workload_name}
Instance: {result.config.workload_instance}
Architecture: {result.config.architecture}
Success: {result.success}
Execution Time: {result.execution_time:.2f} seconds
Error Message: {result.error_message or 'None'}

COMMAND EXECUTED:
{ ' '.join([sys.executable, 'polaris.py'] + result.config.get_command_args()) }

STDOUT:
{result.stdout}

STDERR:
{result.stderr}
"""
            
            filename = f"{result.config.workload_name}_{result.config.workload_instance}_{result.config.architecture}_run_log.txt"
            log_path = output_dir / filename
            
            with open(log_path, 'w') as f:
                f.write(log_content)
            
            logger.debug(f"Saved run log: {log_path}")

        logger.info("Detailed run outputs saved successfully")

    def archive_raw_results(self, repo_path: Path, output_dir: Path) -> None:
        """
        Archive the raw Polaris results directory to the output directory.

        Args:
            repo_path: Path to the Polaris repository
            output_dir: Directory to save the archived results
        """
        results_src = repo_path / "results"
        archive_directory(results_src, output_dir, "raw_polaris_results")

    def _create_consolidated_sheet(self, wb: Workbook, results: List[ExecutionResult]) -> None:
        """Create a single consolidated worksheet with all metrics and results"""
        ws = wb.create_sheet("Polaris Performance Report")

        # Title and metadata
        ws.cell(row=1, column=1, value="Polaris Workload Performance Report")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=2, column=1).font = Font(italic=True)

        # Summary statistics
        total_workloads = len(results)
        successful_results = [r for r in results if r.success]
        failed_results = [r for r in results if not r.success]

        ws.cell(row=4, column=1, value="EXECUTION SUMMARY")
        ws.cell(row=4, column=1).font = Font(bold=True, size=14)
        ws.cell(row=4, column=1).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        ws.cell(row=5, column=1, value="Total Workloads:")
        ws.cell(row=5, column=2, value=total_workloads)
        ws.cell(row=6, column=1, value="Successful:")
        ws.cell(row=6, column=2, value=f"{len(successful_results)} ({len(successful_results)/total_workloads*100:.1f}%)")
        ws.cell(row=7, column=1, value="Failed:")
        ws.cell(row=7, column=2, value=f"{len(failed_results)} ({len(failed_results)/total_workloads*100:.1f}%)")

        # Performance metrics section
        current_row = 10
        ws.cell(row=current_row, column=1, value="PERFORMANCE METRICS")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        current_row += 3
        headers = [
            "Workload", "Instance", "Architecture", "Status", "Exec Time (ms)",
            "Total Cycles", "Total Time (ms)", "Memory (GB)", "Fits Device",
            "Perf Projection", "Compute %", "Memory %", "Frequency (MHz)", "Batch Size",
            "Ideal Throughput", "Device Mem (GB)", "Operator Count"
        ]

        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        current_row += 1

        # Data rows for all results
        for result in results:
            ws.cell(row=current_row, column=1, value=result.config.workload_name)
            ws.cell(row=current_row, column=2, value=result.config.workload_instance)
            ws.cell(row=current_row, column=3, value=result.config.architecture)
            ws.cell(row=current_row, column=4, value="Success" if result.success else "Failed")
            ws.cell(row=current_row, column=5, value=round(result.execution_time * 1000, 2))

            if result.success and result.metrics:
                ws.cell(row=current_row, column=6, value=result.metrics.get('total_cycles', 0))
                ws.cell(row=current_row, column=7, value=result.metrics.get('total_msecs', 0))
                ws.cell(row=current_row, column=8, value=result.metrics.get('memory_size_gb', 0))
                ws.cell(row=current_row, column=9, value="Yes" if result.metrics.get('fits_device', False) else "No")
                ws.cell(row=current_row, column=10, value=result.metrics.get('performance_projection', 0))
                ws.cell(row=current_row, column=11, value=result.metrics.get('resource_compute', 0))
                ws.cell(row=current_row, column=12, value=result.metrics.get('resource_memory', 0))
                ws.cell(row=current_row, column=13, value=result.metrics.get('frequency_mhz', 0))
                ws.cell(row=current_row, column=14, value=result.metrics.get('batch_size', 1))
                ws.cell(row=current_row, column=15, value=result.metrics.get('ideal_throughput', 0))
                ws.cell(row=current_row, column=16, value=result.metrics.get('device_memory_gb', 0))
                ws.cell(row=current_row, column=17, value=result.metrics.get('operator_count', 0))
            else:
                # For failed runs, put error message in a separate column if needed
                if result.error_message and current_row <= 1048576 - 1:  # Excel row limit
                    ws.cell(row=current_row, column=18, value=result.error_message[:100])

            current_row += 1

        # Failed runs section (if any)
        if failed_results:
            current_row += 5  # More spacing before new section
            ws.cell(row=current_row, column=1, value="FAILED RUNS DETAILS")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

            current_row += 3
            error_headers = ["Workload", "Instance", "Architecture", "Error Message", "Exec Time (ms)"]
            for col, header in enumerate(error_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            current_row += 1
            for result in failed_results:
                ws.cell(row=current_row, column=1, value=result.config.workload_name)
                ws.cell(row=current_row, column=2, value=result.config.workload_instance)
                ws.cell(row=current_row, column=3, value=result.config.architecture)
                ws.cell(row=current_row, column=4, value=result.error_message or "Unknown error")
                ws.cell(row=current_row, column=5, value=round(result.execution_time * 1000, 2))
                current_row += 1

        # Performance comparison section
        if successful_results:
            current_row += 5  # More spacing before new section
            ws.cell(row=current_row, column=1, value="PERFORMANCE COMPARISON")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")

            current_row += 3
            comp_headers = ["Workload.Instance", "Architecture", "Cycles", "Time (ms)", "Throughput", "Memory (GB)"]
            for col, header in enumerate(comp_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")

            current_row += 1
            for result in successful_results:
                if result.metrics:
                    workload_id = f"{result.config.workload_name}.{result.config.workload_instance}"
                    ws.cell(row=current_row, column=1, value=workload_id)
                    ws.cell(row=current_row, column=2, value=result.config.architecture)
                    ws.cell(row=current_row, column=3, value=result.metrics.get('total_cycles', 0))
                    ws.cell(row=current_row, column=4, value=result.metrics.get('total_msecs', 0))
                    ws.cell(row=current_row, column=5, value=result.metrics.get('ideal_throughput', 0))
                    ws.cell(row=current_row, column=6, value=result.metrics.get('memory_size_gb', 0))
                    current_row += 1

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col in range(1, 19):  # Cover all possible columns
            max_length = 10  # Minimum width
            for row in range(1, current_row):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    max_length = max(max_length, len(cell_value))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)  # Cap at 30 for readability

    def _create_summary_sheet(self, wb: Workbook, results: List[ExecutionResult]) -> None:
        """Create summary worksheet with high-level metrics"""
        # Filter to only successful results for the main summary
        successful_results = [r for r in results if r.success]
        ws = wb.create_sheet("Results Summary")

        # Headers
        headers = [
            "Workload", "Instance", "Architecture", "Status", "Execution Time (ms)",
            "Total Cycles", "Total Time (ms)", "Memory Usage (GB)", "Fits Device",
            "Performance Projection", "Compute Resource %", "Memory Resource %"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data rows
        for row, result in enumerate(successful_results, 2):
            ws.cell(row=row, column=1, value=result.config.workload_name)
            ws.cell(row=row, column=2, value=result.config.workload_instance)
            ws.cell(row=row, column=3, value=result.config.architecture)
            ws.cell(row=row, column=4, value="Success")  # All are successful since we filtered
            ws.cell(row=row, column=5, value=round(result.execution_time * 1000, 2))

            if result.metrics:
                ws.cell(row=row, column=6, value=result.metrics.get('total_cycles', 0))
                ws.cell(row=row, column=7, value=result.metrics.get('total_msecs', 0))
                ws.cell(row=row, column=8, value=result.metrics.get('memory_size_gb', 0))
                ws.cell(row=row, column=9, value="Yes" if result.metrics.get('fits_device', False) else "No")
                ws.cell(row=row, column=10, value=result.metrics.get('performance_projection', 0))
                ws.cell(row=row, column=11, value=result.metrics.get('resource_compute', 0))
                ws.cell(row=row, column=12, value=result.metrics.get('resource_memory', 0))
            else:
                ws.cell(row=row, column=4, value="Failed")
                if result.error_message:
                    ws.cell(row=row, column=13, value=result.error_message[:100])  # Truncate long errors

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_detailed_sheet(self, wb: Workbook, results: List[ExecutionResult]) -> None:
        """Create detailed metrics worksheet"""
        ws = wb.create_sheet("Performance Metrics")

        # Headers
        headers = [
            "Workload", "Instance", "Architecture", "Frequency (MHz)", "Batch Size",
            "Ideal Throughput", "Device Memory (GB)", "Operator Count"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data rows
        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=result.config.workload_name)
            ws.cell(row=row, column=2, value=result.config.workload_instance)
            ws.cell(row=row, column=3, value=result.config.architecture)

            if result.success and result.metrics:
                ws.cell(row=row, column=4, value=result.metrics.get('frequency_mhz', 0))
                ws.cell(row=row, column=5, value=result.metrics.get('batch_size', 1))
                ws.cell(row=row, column=6, value=result.metrics.get('ideal_throughput', 0))
                ws.cell(row=row, column=7, value=result.metrics.get('device_memory_gb', 0))
                ws.cell(row=row, column=8, value=result.metrics.get('operator_count', 0))

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_performance_sheet(self, wb: Workbook, results: List[ExecutionResult]) -> None:
        """Create performance analysis worksheet"""
        ws = wb.create_sheet("Architecture Analysis")

        # Filter successful results
        successful_results = [r for r in results if r.success]

        if not successful_results:
            ws.cell(row=1, column=1, value="No successful results to analyze")
            return

        # Create performance comparison table
        ws.cell(row=1, column=1, value="Performance Comparison")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Headers
        headers = ["Workload.Instance", "Architecture", "Cycles", "Time (ms)", "Throughput", "Memory (GB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        for row, result in enumerate(successful_results, 3):
            workload_id = f"{result.config.workload_name}.{result.config.workload_instance}"
            ws.cell(row=row, column=1, value=workload_id)
            ws.cell(row=row, column=2, value=result.config.architecture)
            ws.cell(row=row, column=3, value=result.metrics.get('total_cycles', 0))
            ws.cell(row=row, column=4, value=result.metrics.get('total_msecs', 0))
            ws.cell(row=row, column=5, value=result.metrics.get('ideal_throughput', 0))
            ws.cell(row=row, column=6, value=result.metrics.get('memory_size_gb', 0))

    def _create_error_sheet(self, wb: Workbook, results: List[ExecutionResult]) -> None:
        """Create error analysis worksheet"""
        ws = wb.create_sheet("Failed Runs")

        # Filter failed results
        failed_results = [r for r in results if not r.success]

        if not failed_results:
            ws.cell(row=1, column=1, value="No errors to report")
            return

        # Headers
        headers = ["Workload", "Instance", "Architecture", "Error Message", "Execution Time (ms)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        for row, result in enumerate(failed_results, 2):
            ws.cell(row=row, column=1, value=result.config.workload_name)
            ws.cell(row=row, column=2, value=result.config.workload_instance)
            ws.cell(row=row, column=3, value=result.config.architecture)
            ws.cell(row=row, column=4, value=result.error_message or "Unknown error")
            ws.cell(row=row, column=5, value=round(result.execution_time * 1000, 2))

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col in range(1, 6):  # Include error message column
            ws.column_dimensions[get_column_letter(col)].width = 20


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description="Polaris Workload Automation Script")
    parser.add_argument("--repo-path", default=".", help="Path to Polaris repository (default: current repo)")
    parser.add_argument("--output-dir", default=None, help="Output directory for results (default: unified structure)")
    parser.add_argument("--unified-output-dir", default=None,
                        help="Unified output directory for all profiling results (default: auto-generated)")
    parser.add_argument("--target-archs", default=None,
                        help="Comma-separated list of target architectures to run (e.g., 'n150,n300')")
    parser.add_argument("--arch-config", default="config/tt_wh.yaml",
                        help="Path to architecture configuration YAML file (default: config/tt_wh.yaml)")
    parser.add_argument("--workload-configs", default=None,
                        help="Comma-separated list of workload configuration YAML files (default: config/wh_supported.yaml)")
    parser.add_argument("--local-workload-config", type=Path, default=None,
                        help="Path to local workload configuration YAML file (overrides workload-configs)")
    parser.add_argument("--run-type", default="default",
                        help="Type of run for unique output directories (default: default)")
    parser.add_argument("--clean-repo", action="store_true", help="Clean and re-clone the repository")
    parser.add_argument("--dry-run", action="store_true", help="Dry run - discover workloads without executing")
    
    args = parser.parse_args()

    # Setup unified output directory structure
    if args.unified_output_dir:
        unified_base = Path(args.unified_output_dir)
    else:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unified_base = Path(f"HW_Polaris_comparison_reports_{timestamp}")

    # Create orderly directory structure
    structure_paths = create_orderly_output_structure(unified_base)

    # Set up output directory within unified structure
    if args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        output_dir = structure_paths['polaris_simulation_results']

    # Setup logging to unified logs directory
    logs_dir = structure_paths['logs']
    log_file = logs_dir / "polaris_automation.log"

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    logger = logging.getLogger(__name__)

    try:
        # Initialize automation
        automation = PolarisAutomation(args.repo_path)

        # Setup repository
        if not automation.setup_repository(clean_repo=args.clean_repo):
            logger.error("Failed to setup repository")
            return 1

        # Parse target architectures
        if args.target_archs:
            target_architectures = [arch.strip() for arch in args.target_archs.split(',') if arch.strip()]
            logger.info(f"Targeting architectures: {target_architectures}")
        else:
            # Default to n150 and n300 from tt_wh.yaml
            target_architectures = ['n150', 'n300']
            logger.info(f"Default targeting architectures: {target_architectures}")

        # Prepare configs as lists
        arch_configs = [args.arch_config]

        # Handle local workload config (takes priority)
        if args.local_workload_config:
            local_config_path = Path(args.local_workload_config)
            if local_config_path.exists():
                # For local config, we need to pass it differently since it's outside the repo
                logger.info(f"Using local workload config: {local_config_path}")
                # We'll handle this in the load_configurations method
                workload_configs = [str(local_config_path)]  # Pass as list for compatibility
            else:
                logger.error(f"Local workload config not found: {local_config_path}")
                return 1
        elif args.workload_configs:
            workload_configs = [c.strip() for c in args.workload_configs.split(',') if c.strip()]
            logger.info(f"Using workload configs: {workload_configs}")
        else:
            workload_configs = None
            logger.info("Using default workload configs")

        # Load configurations
        arch_config, workload_config, discovered_targets = automation.load_configurations(arch_configs, workload_configs)

        if args.dry_run:
            # Discover workloads
            configs = automation.discover_workloads(target_architectures, arch_config, list(workload_config.keys()), args.run_type)
            logger.info(f"Dry run: Discovered {len(configs)} workload/architecture combinations")
            for config in configs:
                logger.info(f"  - {config.workload_name}.{config.workload_instance} on {config.architecture}")
            return 0

        # Discover workloads
        configs = automation.discover_workloads(target_architectures, arch_config, list(workload_config.keys()), args.run_type)
        logger.info(f"Discovered {len(configs)} workload/architecture combinations")

        if not configs:
            logger.warning("No workload configurations discovered")
            return 1

        # Execute workloads
        results = automation.execute_workloads(configs)
        logger.info(f"Execution completed. {sum(1 for r in results if r.success)}/{len(results)} successful")

        # Generate Excel report in unified structure
        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_filename = f"polaris_performance_report_{timestamp}.xlsx"
            report_path = structure_paths['reports'] / report_filename
            automation.generate_excel_report(results, report_path)
            logger.info(f"Report generated: {report_path}")
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")

        # Save detailed run outputs to unified structure
        automation.save_detailed_outputs(results, structure_paths['raw_data'])

        # Archive raw Polaris results to unified structure
        automation.archive_raw_results(automation.repo_path, structure_paths['raw_data'])

        # Also copy to the expected polaris_simulation_results directory for analyzer compatibility
        automation.archive_raw_results(automation.repo_path, output_dir)

        # Copy automation log is already handled by unified logging setup
        logger.info("Polaris automation completed successfully")
        logger.info(f"📁 All results organized in: {unified_base}")
        return 0

    except Exception as e:
        logger.error(f"Polaris automation failed: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
