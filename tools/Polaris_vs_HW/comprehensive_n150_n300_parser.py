#!/usr/bin/env python3
"""
Comprehensive n150/n300 Hardware Metrics Parser

Extracts ALL available performance metrics and model support data
from TT-Metal repository for n150 and n300 targets.

"""

import os
import sys
import re
import subprocess
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime

import yaml
from loguru import logger
import requests

# Import common utilities
from common_utils import (
    pull_git_repo, create_excel_workbook, setup_excel_headers,
    apply_cell_formatting, adjust_column_widths, save_excel_workbook,
    ensure_directory, create_orderly_output_structure
)


# Repository cloning function removed - using common_utils.pull_git_repo


@dataclass
class ModelSupport:
    """Represents model hardware support information"""
    model_name: str
    model_variant: str
    hardware: str
    source: str = "tt_transformers"


@dataclass
class PerformanceMetric:
    """Represents a performance metric from various sources"""
    model_name: str
    model_variant: str
    hardware: str
    metric_type: str
    value: float
    unit: str
    target: str = ""
    batch_size: Optional[int] = None
    sequence_length: Optional[int] = None
    source: str = ""
    notes: Optional[str] = None


class ComprehensiveHardwareParser:
    """Comprehensive parser for all n150/n300 hardware metrics"""

    def __init__(self, repo_dir: Path):
        self.model_support: List[ModelSupport] = []
        self.performance_metrics: List[PerformanceMetric] = []
        self.repo_dir = repo_dir
        self.session = requests.Session()  # Keep for backup web fetching
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (compatible; PolarisParser/1.0)'
        })

    def fetch_content(self, url: str) -> str:
        """Fetch content from local repository or URL with fallback
        
        Args:
            url: URL of the content (used to determine local path and as fallback)
            
        Returns:
            str: Content of the file or empty string on failure
        """
        try:
            # Convert GitHub URL to local path
            if "raw.githubusercontent.com" in url:
                # Extract path after repository name
                path_parts = url.split("/main/")
                if len(path_parts) == 2:
                    local_path = self.repo_dir / path_parts[1]
                    if local_path.exists():
                        logger.info(f"Reading from local file: {local_path}")
                        return local_path.read_text()
            
            # Fallback to web request
            logger.info(f"Falling back to web request: {url}")
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            return response.text
            
        except Exception as e:
            logger.error(f"Failed to fetch content from {url}: {e}")
            return ""

    def parse_tt_transformers_support(self) -> List[ModelSupport]:
        """Parse model support from tt_transformers README"""
        logger.info("Fetching tt_transformers model support...")

        url = "https://raw.githubusercontent.com/tenstorrent/tt-metal/main/models/tt_transformers/README.md"
        content = self.fetch_content(url)

        if not content:
            return []

        support_list = []

        # Find the main table (between the headers and details tag)
        lines = content.split('\n')
        in_table = False
        headers = []

        for line in lines:
            line = line.strip()

            if line.startswith('|') and 'Hardware' in line and 'Model' in line:
                # This is the header row
                headers = [col.strip() for col in line.split('|')[1:-1]]
                in_table = True
                continue

            if in_table and line.startswith('|') and not line.startswith('|---'):
                # This is a data row
                cols = [col.strip() for col in line.split('|')[1:-1]]

                if len(cols) >= 3 and len(headers) >= 3:
                    try:
                        model_info = cols[0] if len(cols) > 0 else ""
                        hardware_info = cols[1] if len(cols) > 1 else ""

                        # Parse model name from markdown link
                        if '[' in model_info and '](' in model_info:
                            model_name = model_info.split('](')[0][1:]
                        else:
                            model_name = model_info

                        # Parse hardware support
                        if '/' in hardware_info:
                            # Multiple hardware types
                            hardware_types = [h.strip() for h in hardware_info.split('/') if h.strip()]
                        else:
                            hardware_types = [hardware_info]

                        # Extract model name and variant
                        model_name_clean = model_name.split()[0] if model_name else ""
                        model_variant = ' '.join(model_name.split()[1:]) if len(model_name.split()) > 1 else ""

                        # Create support entries for n150/n300 only
                        for hw in hardware_types:
                            hw_clean = hw.lower().strip()
                            if hw_clean in ['n150', 'n300', 'wormhole']:
                                # Map to standardized names - only WH n150/n300
                                if hw_clean == 'wormhole':
                                    # Add both n150 and n300 for generic "wormhole" entries
                                    support_list.append(ModelSupport(
                                        model_name=model_name_clean,
                                        model_variant=model_variant,
                                        hardware='n150'
                                    ))
                                    support_list.append(ModelSupport(
                                        model_name=model_name_clean,
                                        model_variant=model_variant,
                                        hardware='n300'
                                    ))
                                else:
                                    support_list.append(ModelSupport(
                                        model_name=model_name_clean,
                                        model_variant=model_variant,
                                        hardware=hw_clean
                                    ))


                    except Exception as e:
                        logger.debug(f"Failed to parse row: {cols}, error: {e}")
                        continue

            if line.startswith('<details>'):
                # End of main table
                break

        logger.info(f"Found {len(support_list)} model support entries")
        return support_list

    def parse_model_updates_performance(self) -> List[PerformanceMetric]:
        """Parse performance metrics from MODEL_UPDATES.md"""
        logger.info("Fetching MODEL_UPDATES.md performance data...")

        url = "https://raw.githubusercontent.com/tenstorrent/tt-metal/main/models/docs/MODEL_UPDATES.md"
        content = self.fetch_content(url)

        if not content:
            return []

        metrics = []

        # Look for specific performance patterns
        lines = content.split('\n')

        for i, line in enumerate(lines):
            line = line.strip()

            # Pattern 1: "Achieved X t/s/u ... on Wormhole Galaxy"
            if 'achieved' in line.lower() and 't/s/u' in line.lower():
                metric = self._parse_tsu_metric(line, lines, i)
                if metric:
                    metrics.append(metric)

            # Pattern 2: Performance improvements mentioned
            elif 'increased from' in line.lower() and 't/s/u' in line.lower():
                metric = self._parse_improvement_metric(line)
                if metric:
                    metrics.append(metric)

        # Add known metrics from manual parsing
        manual_metrics = [
            PerformanceMetric(
                model_name="Llama",
                model_variant="3.1 70B",
                hardware="n150",
                metric_type="tokens/sec/user",
                value=65.0,
                unit="t/s/u",
                target="",
                batch_size=32,
                sequence_length=128,
                source="MODEL_UPDATES.md",
                notes="Decode mode with vLLM fork"
            ),
            PerformanceMetric(
                model_name="Various",
                model_variant="1B/3B/8B/11B",
                hardware="n150",
                metric_type="tokens/sec/user",
                value=28.0,
                unit="t/s/u",
                target="",
                source="MODEL_UPDATES.md",
                notes="8B model improvement from ~23 to ~28 t/s/u using BFP4 weights"
            ),
            PerformanceMetric(
                model_name="Various",
                model_variant="Demo",
                hardware="n150",
                metric_type="tokens/sec/user",
                value=23.0,
                unit="t/s/u",
                target="",
                source="MODEL_UPDATES.md",
                notes="TT-NN tracing demo"
            )
        ]

        metrics.extend(manual_metrics)
        logger.info(f"Found {len(metrics)} performance metrics from MODEL_UPDATES.md")
        return metrics

    def _parse_tsu_metric(self, line: str, lines: List[str], line_index: int) -> Optional[PerformanceMetric]:
        """Parse a t/s/u performance metric"""
        try:
            # Extract the value
            tsu_match = re.search(r'(\d+(?:\.\d+)?)\s*t/s/u', line.lower())
            if not tsu_match:
                return None

            value = float(tsu_match.group(1))

            # Determine hardware from context
            hardware = "n150"  # Default to n150 for Galaxy
            if "galaxy" in line.lower():
                hardware = "n150"
            elif "quietbox" in line.lower():
                hardware = "n300"

            # Extract batch size and sequence length from nearby lines
            batch_size = None
            seq_length = None

            # Look in current line and next few lines for context
            context_lines = lines[max(0, line_index-2):min(len(lines), line_index+5)]
            context = ' '.join(context_lines).lower()

            batch_match = re.search(r'batch\s*(?:size)?\s*(?:of\s*)?(\d+)', context)
            if batch_match:
                batch_size = int(batch_match.group(1))

            seq_match = re.search(r'(\d+)\s*(?:input\s*)?(?:sequence\s*)?length', context)
            if seq_match:
                seq_length = int(seq_match.group(1))

            # Extract model info from section header
            model_name = "Unknown"
            model_variant = ""

            # Look backwards for section header
            for j in range(line_index, max(0, line_index-20), -1):
                prev_line = lines[j].strip()
                if prev_line.startswith('### [') and '](' in prev_line:
                    model_text = prev_line.split('](')[0][4:]
                    model_name = model_text.split()[0]
                    model_variant = ' '.join(model_text.split()[1:])
                    break

            return PerformanceMetric(
                model_name=model_name,
                model_variant=model_variant,
                hardware=hardware,
                metric_type="tokens/sec/user",
                value=value,
                unit="t/s/u",
                target="",
                batch_size=batch_size,
                sequence_length=seq_length,
                source="MODEL_UPDATES.md",
                notes=f"From line: {line[:100]}..."
            )

        except Exception as e:
            logger.debug(f"Failed to parse t/s/u metric from: {line}, error: {e}")
            return None

    def _parse_improvement_metric(self, line: str) -> Optional[PerformanceMetric]:
        """Parse performance improvement metrics"""
        try:
            # Pattern: "increased from ~23 t/s/u to ~28 t/s/u"
            match = re.search(r'increased\s+from\s+.*?(\d+(?:\.\d+)?)\s*t/s/u\s+to\s+.*?(\d+(?:\.\d+)?)\s*t/s/u', line.lower())
            if match:
                old_value = float(match.group(1))
                new_value = float(match.group(2))

                return PerformanceMetric(
                    model_name="Various",
                    model_variant="1B/3B/8B/11B",
                    hardware="n150",
                    metric_type="tokens/sec/user",
                    value=new_value,
                    unit="t/s/u",
                    target="",
                    source="MODEL_UPDATES.md",
                    notes=f"Improvement from {old_value} to {new_value} t/s/u using BFP4 weights"
                )

        except Exception as e:
            logger.debug(f"Failed to parse improvement metric: {e}")

        return None

    def parse_comprehensive_model_matrix(self) -> Tuple[List[PerformanceMetric], List[PerformanceMetric]]:
        """Parse comprehensive performance data from models/README.md Model Matrix"""
        logger.info("Parsing comprehensive Model Matrix from models/README.md...")

        llm_metrics = []
        cnn_metrics = []

        try:
            # Fetch the models README
            response = requests.get("https://raw.githubusercontent.com/tenstorrent/tt-metal/main/models/README.md")
            content = response.text

            # Parse LLM section
            llm_section = self._extract_section(content, "## LLMs", "## Speech-to-Text")
            llm_metrics = self._parse_llm_table(llm_section)

            # Parse Speech-to-Text section (Whisper)
            speech_section = self._extract_section(content, "## Speech-to-Text", "## Diffusion Models")
            whisper_metrics = self._parse_whisper_table(speech_section)
            llm_metrics.extend(whisper_metrics)

            # Parse CNN section
            cnn_section = self._extract_section(content, "## CNNs and Vision Transformers", "## NLPs")
            cnn_metrics = self._parse_cnn_table(cnn_section)

        except Exception as e:
            logger.error(f"Failed to parse Model Matrix: {e}")

        logger.info(f"Found {len(llm_metrics)} LLM metrics and {len(cnn_metrics)} CNN metrics from Model Matrix")
        return llm_metrics, cnn_metrics

    def _extract_section(self, content: str, start_marker: str, end_marker: str) -> str:
        """Extract a section between two markers"""
        start_idx = content.find(start_marker)
        if start_idx == -1:
            return ""

        end_idx = content.find(end_marker, start_idx)
        if end_idx == -1:
            end_idx = len(content)

        return content[start_idx:end_idx]

    def _parse_llm_table(self, section: str) -> List[PerformanceMetric]:
        """Parse LLM performance table from models README"""
        metrics = []
        lines = section.split('\n')

        for line in lines:
            if '|' in line and '[' in line and ']' in line:
                # Parse table row
                parts = [p.strip() for p in line.split('|')[1:-1]]  # Remove empty parts at start/end
                if len(parts) >= 7:
                    try:
                        model_info = parts[0]
                        batch_size = int(parts[1]) if parts[1].strip() else 0
                        hardware_info = parts[2]
                        # Strip asterisks (*) from performance values and convert to float
                        ttft_str = parts[3].strip().rstrip('*')
                        tsu_str = parts[4].strip().rstrip('*')
                        target_tsu_str = parts[5].strip().rstrip('*')
                        ts_str = parts[6].strip().rstrip('*')

                        ttft = float(ttft_str) if ttft_str and ttft_str != '' else 0
                        tsu = float(tsu_str) if tsu_str and tsu_str != '' else 0
                        target_tsu = float(target_tsu_str) if target_tsu_str and target_tsu_str != '' and target_tsu_str != '|' else 0
                        ts = float(ts_str) if ts_str and ts_str != '' else 0

                        # Extract model name and variant
                        model_match = re.search(r'\[([^\]]+)\]', model_info)
                        if model_match:
                            model_full = model_match.group(1)
                            # Parse model name and configuration
                            model_name, model_variant = self._parse_model_name(model_full)

                            # Extract hardware
                            hardware = self._extract_hardware_from_link(hardware_info)

                            if hardware in ['n150', 'n300']:
                                # Format notes consistently for LLMs
                                ttft_str = f"{ttft}ms" if ttft > 0 else "N/A"
                                tsu_str = f"{tsu}" if tsu > 0 else "N/A"
                                target_tsu_str = f"{target_tsu}" if target_tsu > 0 else "N/A"

                                notes_llm = f"TTFT: {ttft_str}, T/S/U: {tsu_str}, Target T/S/U: {target_tsu_str}"

                                # Set target value for the column
                                target_value = target_tsu_str if target_tsu > 0 else ""

                                metrics.append(PerformanceMetric(
                                    model_name=model_name,
                                    model_variant=model_variant,
                                    hardware=hardware,
                                    metric_type="tokens/sec",
                                    value=ts,
                                    unit="t/s",
                                    target=target_value,
                                    batch_size=batch_size,
                                    source="models/README.md Model Matrix",
                                    notes=notes_llm
                                ))

                                # Also add T/S/U metric with consistent notes
                                if tsu > 0:
                                    metrics.append(PerformanceMetric(
                                        model_name=model_name,
                                        model_variant=model_variant,
                                        hardware=hardware,
                                        metric_type="tokens/sec/user",
                                        value=tsu,
                                        unit="t/s/u",
                                        target=target_value,
                                        batch_size=batch_size,
                                        source="models/README.md Model Matrix",
                                        notes=notes_llm
                                    ))

                    except (ValueError, IndexError) as e:
                        logger.debug(f"Failed to parse LLM row: {line} - {e}")

        return metrics

    def _parse_whisper_table(self, section: str) -> List[PerformanceMetric]:
        """Parse Whisper performance table"""
        metrics = []
        lines = section.split('\n')

        for line in lines:
            if 'Whisper' in line and '|' in line:
                parts = [p.strip() for p in line.split('|')[1:-1]]
                if len(parts) >= 7:
                    try:
                        model_info = parts[0]
                        batch_size = int(parts[1]) if parts[1].strip() else 0
                        hardware_info = parts[2]
                        ttft_str = parts[3].strip().rstrip('*')
                        tsu_str = parts[4].strip().rstrip('*')
                        ttft = float(ttft_str) if ttft_str else 0
                        tsu = float(tsu_str) if tsu_str else 0
                        target_tsu_str = parts[5].strip().rstrip('*')
                        ts_str = parts[6].strip().rstrip('*')
                        target_tsu = float(target_tsu_str) if target_tsu_str else 0
                        ts = float(ts_str) if ts_str else 0

                        hardware = self._extract_hardware_from_link(hardware_info)

                        if hardware in ['n150', 'n300']:
                            # Format notes consistently for Whisper (speech-to-text)
                            ttft_str = f"{ttft}ms" if ttft > 0 else "N/A"
                            tsu_str = f"{tsu}" if tsu > 0 else "N/A"
                            target_tsu_str = f"{target_tsu}" if target_tsu and str(target_tsu).strip() else "N/A"

                            notes_whisper = f"TTFT: {ttft_str}, T/S/U: {tsu_str}, Target T/S/U: {target_tsu_str}"

                            # Set target value for the column
                            target_value = target_tsu_str if target_tsu and str(target_tsu).strip() else ""

                            metrics.append(PerformanceMetric(
                                model_name="Whisper",
                                model_variant="distil-large-v3",
                                hardware=hardware,
                                metric_type="tokens/sec",
                                value=ts,
                                unit="t/s",
                                target=target_value,
                                batch_size=batch_size,
                                source="models/README.md Model Matrix",
                                notes=notes_whisper
                            ))

                    except (ValueError, IndexError) as e:
                        logger.debug(f"Failed to parse Whisper row: {line} - {e}")

        return metrics

    def _parse_cnn_table(self, section: str) -> List[PerformanceMetric]:
        """Parse CNN performance table from models README"""
        metrics = []
        lines = section.split('\n')

        current_category = None
        for line in lines:
            if '###' in line:
                if 'Classification' in line:
                    current_category = 'classification'
                elif 'Object Detection' in line:
                    current_category = 'detection'
                elif 'Segmentation' in line:
                    current_category = 'segmentation'

            if '|' in line and '[' in line and ']' in line and current_category:
                parts = [p.strip() for p in line.split('|')[1:-1]]
                if len(parts) >= 5:
                    try:
                        model_info = parts[0]
                        batch_size = int(parts[1]) if parts[1].strip() else 0
                        hardware_info = parts[2]
                        performance = parts[3]
                        target = parts[4] if len(parts) > 4 else ""

                        # Extract model name and variant
                        model_match = re.search(r'\[([^\]]+)\]', model_info)
                        if model_match:
                            model_full = model_match.group(1)
                            model_name, model_variant = self._parse_cnn_model_name(model_full)

                            hardware = self._extract_hardware_from_link(hardware_info)

                            # Parse performance value
                            perf_value = self._parse_performance_value(performance)

                            if perf_value > 0 and hardware in ['n150', 'n300']:
                                # Use img/sec for CNN and SD models as requested
                                if current_category in ['detection', 'segmentation']:
                                    unit = "fps"
                                    metric_type = "frames/sec"
                                else:
                                    unit = "img/sec"
                                    metric_type = "images/sec"

                                # Format target value
                                target_str = f"{target}" if target and target.strip() else "N/A"

                                notes_cnn = f"Category: {current_category}, Target: {target_str}"

                                # Set target value for the column
                                target_value = target_str if target and target.strip() else ""

                                metrics.append(PerformanceMetric(
                                    model_name=model_name,
                                    model_variant=model_variant,
                                    hardware=hardware,
                                    metric_type=metric_type,
                                    value=perf_value,
                                    unit=unit,
                                    target=target_value,
                                    batch_size=batch_size,
                                    source="models/README.md Model Matrix",
                                    notes=notes_cnn
                                ))

                    except (ValueError, IndexError) as e:
                        logger.debug(f"Failed to parse CNN row: {line} - {e}")

        return metrics

    def _parse_model_name(self, model_full: str) -> Tuple[str, str]:
        """Parse model name and variant from full model string"""
        # Handle various model formats
        if 'Llama' in model_full:
            base_match = re.search(r'(Llama \d+\.\d+) (\d+B)', model_full)
            if base_match:
                return base_match.group(1), base_match.group(2)
            # Handle Llama with TP/DP
            tp_match = re.search(r'(Llama \d+\.\d+ \d+B) \((TP|DP)=\d+\)', model_full)
            if tp_match:
                return tp_match.group(1), f"{tp_match.group(2)}={re.search(r'(\d+)', model_full).group(1)}"
        elif 'Qwen' in model_full:
            qwen_match = re.search(r'(Qwen [\d.]+ [\d\w]+)', model_full)
            if qwen_match:
                return qwen_match.group(1), ""
        elif 'Falcon' in model_full:
            falcon_match = re.search(r'(Falcon \d+B)', model_full)
            if falcon_match:
                return falcon_match.group(1), ""
        elif 'Mistral' in model_full:
            return "Mistral", "7B"
        elif 'Mixtral' in model_full:
            return "Mixtral", "8x7B"
        elif 'DeepSeek' in model_full:
            return "DeepSeek R1 Distill", "Llama 70B"
        elif 'Mamba' in model_full:
            return "Mamba", "2.8B"

        return model_full, ""

    def _parse_cnn_model_name(self, model_full: str) -> Tuple[str, str]:
        """Parse CNN model name and variant"""
        if 'ResNet' in model_full:
            resnet_match = re.search(r'(ResNet-\d+)', model_full)
            if resnet_match:
                return "ResNet", resnet_match.group(1).replace('-', ' ')
        elif 'ViT' in model_full:
            return "ViT", "base"
        elif 'MobileNet' in model_full:
            return "MobileNet", "v2"
        elif 'YOLO' in model_full:
            yolo_match = re.search(r'(YOLO\w+\d*\w*)', model_full)
            if yolo_match:
                return yolo_match.group(1), ""
        elif 'UNet' in model_full:
            return "UNet", "VGG19"

        return model_full, ""

    def _extract_hardware_from_link(self, hardware_info: str) -> str:
        """Extract hardware type from hardware link"""
        if 'n150' in hardware_info:
            return 'n150'
        elif 'n300' in hardware_info:
            return 'n300'
        elif 'Galaxy' in hardware_info:
            return 'galaxy'
        elif 'QuietBox' in hardware_info:
            return 'QuietBox'
        elif 'p150' in hardware_info:
            return 'p150'
        elif 'p100' in hardware_info:
            return 'p100'
        return hardware_info

    def _parse_performance_value(self, perf_str: str) -> float:
        """Parse performance value from string with commas and asterisks"""
        try:
            # Remove commas, asterisks, and extra spaces
            clean_str = perf_str.replace(',', '').replace('*', '').strip()
            return float(clean_str)
        except (ValueError, AttributeError):
            return 0.0

    def parse_cnn_performance(self) -> List[PerformanceMetric]:
        """Parse CNN performance metrics from Model Matrix (main source now)"""
        logger.info("Collecting CNN performance data from Model Matrix...")

        # Get data from Model Matrix
        _, cnn_metrics = self.parse_comprehensive_model_matrix()

        logger.info(f"Found {len(cnn_metrics)} CNN performance metrics from Model Matrix")
        return cnn_metrics

    def collect_all_data(self) -> Dict[str, Any]:
        """Collect all hardware metrics and model support data"""
        logger.info("Starting comprehensive n150/n300 data collection...")

        # Collect model support
        self.model_support = self.parse_tt_transformers_support()

        # Collect performance metrics from multiple sources
        matrix_llm_metrics, matrix_cnn_metrics = self.parse_comprehensive_model_matrix()
        updates_llm_metrics = self.parse_model_updates_performance()

        # Combine all metrics (prioritize Model Matrix data)
        self.performance_metrics = matrix_llm_metrics + matrix_cnn_metrics + updates_llm_metrics

        # Create comprehensive report - WH n150/n300 only
        data = {
            "model_support": self.model_support,
            "performance_metrics": self.performance_metrics,
            "summary": {
                "total_models_supported": len(set(f"{m.model_name}_{m.model_variant}" for m in self.model_support)),
                "n150_models": len(set(f"{m.model_name}_{m.model_variant}" for m in self.model_support if m.hardware == "n150")),
                "n300_models": len(set(f"{m.model_name}_{m.model_variant}" for m in self.model_support if m.hardware == "n300")),
                "total_performance_metrics": len(self.performance_metrics),
                "wh_n150_metrics": len([m for m in self.performance_metrics if m.hardware == "n150"]),
                "wh_n300_metrics": len([m for m in self.performance_metrics if m.hardware == "n300"]),
                "matrix_llm_metrics": len(matrix_llm_metrics),
                "matrix_cnn_metrics": len(matrix_cnn_metrics),
                "updates_llm_metrics": len(updates_llm_metrics),
                "generated_at": datetime.now().isoformat()
            }
        }

        logger.info(f"Collection complete: {len(self.model_support)} model support entries, {len(self.performance_metrics)} performance metrics ({len(matrix_llm_metrics) + len(updates_llm_metrics)} LLM, {len(matrix_cnn_metrics)} CNN)")
        return data

    def generate_comprehensive_report(self, output_path: Path, data: Dict[str, Any]) -> bool:
        """Generate comprehensive Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = create_excel_workbook()
            wb.remove(wb.active)

            # Sheet 1: Model Support Matrix
            ws1 = wb.create_sheet("Model Support Matrix")

            # Title
            ws1.cell(row=1, column=1, value="Comprehensive n150/n300 Model Support")
            ws1.cell(row=1, column=1).font = Font(bold=True, size=16)

            ws1.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws1.cell(row=2, column=1).font = Font(italic=True)

            # Summary
            summary = data["summary"]
            ws1.cell(row=4, column=1, value="SUMMARY")
            ws1.cell(row=4, column=1).font = Font(bold=True, size=14)

            ws1.cell(row=5, column=1, value="Total Models with Support:")
            ws1.cell(row=5, column=2, value=summary["total_models_supported"])
            ws1.cell(row=6, column=1, value="Models on n150:")
            ws1.cell(row=6, column=2, value=summary["n150_models"])
            ws1.cell(row=7, column=1, value="Models on n300:")
            ws1.cell(row=7, column=2, value=summary["n300_models"])
            ws1.cell(row=8, column=1, value="Total Performance Metrics:")
            ws1.cell(row=8, column=2, value=summary["total_performance_metrics"])
            ws1.cell(row=9, column=1, value="Model Matrix LLM Metrics:")
            ws1.cell(row=9, column=2, value=summary["matrix_llm_metrics"])
            ws1.cell(row=10, column=1, value="Model Matrix CNN Metrics:")
            ws1.cell(row=10, column=2, value=summary["matrix_cnn_metrics"])
            ws1.cell(row=11, column=1, value="MODEL_UPDATES LLM Metrics:")
            ws1.cell(row=11, column=2, value=summary["updates_llm_metrics"])

            # Model support table
            current_row = 13
            ws1.cell(row=current_row, column=1, value="MODEL SUPPORT MATRIX")
            ws1.cell(row=current_row, column=1).font = Font(bold=True, size=14)

            current_row += 2
            headers = ["Model", "Variant", "n150 Support", "n300 Support", "Source"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            current_row += 1

            # Group by model
            model_groups = {}
            for support in data["model_support"]:
                key = f"{support.model_name}_{support.model_variant}"
                if key not in model_groups:
                    model_groups[key] = {"model": support.model_name, "variant": support.model_variant, "n150": False, "n300": False}
                model_groups[key][support.hardware] = True

            for model_data in model_groups.values():
                ws1.cell(row=current_row, column=1, value=model_data["model"])
                ws1.cell(row=current_row, column=2, value=model_data["variant"])
                ws1.cell(row=current_row, column=3, value="✅" if model_data["n150"] else "❌")
                ws1.cell(row=current_row, column=4, value="✅" if model_data["n300"] else "❌")
                ws1.cell(row=current_row, column=5, value="tt_transformers")
                current_row += 1

            # Sheet 2: Performance Metrics
            ws2 = wb.create_sheet("Performance Metrics")

            ws2.cell(row=1, column=1, value="Performance Metrics for WH n150/n300")
            ws2.cell(row=1, column=1).font = Font(bold=True, size=16)

            current_row = 4
            ws2.cell(row=current_row, column=1, value="PERFORMANCE METRICS")
            ws2.cell(row=current_row, column=1).font = Font(bold=True, size=14)

            current_row += 2
            perf_headers = ["Model", "Variant", "Hardware", "Metric Type", "Value", "Unit", "Target", "Batch Size", "Seq Length", "Source", "Notes"]
            for col, header in enumerate(perf_headers, 1):
                cell = ws2.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            current_row += 1

            for metric in data["performance_metrics"]:
                ws2.cell(row=current_row, column=1, value=metric.model_name)
                ws2.cell(row=current_row, column=2, value=metric.model_variant)
                ws2.cell(row=current_row, column=3, value=metric.hardware)
                ws2.cell(row=current_row, column=4, value=metric.metric_type)
                ws2.cell(row=current_row, column=5, value=metric.value)
                ws2.cell(row=current_row, column=6, value=metric.unit)
                ws2.cell(row=current_row, column=7, value=metric.target or '')
                ws2.cell(row=current_row, column=8, value=metric.batch_size or '')
                ws2.cell(row=current_row, column=9, value=metric.sequence_length or '')
                ws2.cell(row=current_row, column=10, value=metric.source)
                ws2.cell(row=current_row, column=11, value=metric.notes or '')
                current_row += 1

            # Auto-adjust column widths for both sheets
            for ws in [ws1, ws2]:
                for col in range(1, 15):
                    max_length = 10
                    for row in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)

            save_excel_workbook(wb, Path(output_path))
            logger.info(f"✅ Comprehensive report saved: {output_path}")
            return True

        except Exception as e:
            logger.error(f"❌ Failed to generate comprehensive report: {e}")
            return False


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description="Collect Comprehensive n150/n300 Hardware Metrics")
    parser.add_argument("--output-dir", default=None,
                       help="Output directory for Excel reports (default: unified structure)")
    parser.add_argument("--unified-output-dir", default=None,
                       help="Unified output directory for all profiling results (default: auto-generated)")
    parser.add_argument("--tt-metal-repo", default="./tt-metal",
                       help="Path to tt-metal repository. Will be cloned if it doesn't exist.")
    parser.add_argument("--tt-metal-url",
                       default="https://github.com/tenstorrent/tt-metal.git",
                       help="URL of the tt-metal repository")
    parser.add_argument("--skip-repo-update", action="store_true",
                       help="Skip repository clone/pull operations")

    args = parser.parse_args()

    # Setup unified output directory structure
    if args.unified_output_dir:
        unified_base = Path(args.unified_output_dir)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unified_base = Path(f"HW_Polaris_comparison_reports_{timestamp}")

    # Create orderly directory structure
    structure_paths = create_orderly_output_structure(unified_base)

    # Set up output directory within unified structure
    if args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        # Save Excel reports in the Reports folder, not tt_metal_hw_results
        output_dir = structure_paths['reports']

    # Setup logging to unified logs directory
    logs_dir = structure_paths['logs']

    # Configure logging
    import logging
    log_file = logs_dir / "tt_metal_parser.log"
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    logger = logging.getLogger(__name__)

    repo_dir = Path(args.tt_metal_repo)

    # Handle repository operations
    if not args.skip_repo_update:
        # Extract repo name from URL (e.g., "https://github.com/tenstorrent/tt-metal.git" -> "tenstorrent/tt-metal")
        repo_name = args.tt_metal_url.replace("https://github.com/", "").replace(".git", "")
        if not pull_git_repo(repo_name, repo_dir):
            print("❌ Failed to setup tt-metal repository")
            return 1
    elif not repo_dir.exists():
        print(f"❌ tt-metal repository not found at {repo_dir} and --skip-repo-update was specified")
        return 1

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"comprehensive_n150_n300_report_{timestamp}.xlsx"

    # Collect and report
    collector = ComprehensiveHardwareParser(repo_dir)
    data = collector.collect_all_data()

    print("📊 Collection Summary:")
    print(f"   • Model Support Entries: {len(data['model_support'])}")
    print(f"   • Unique Models: {data['summary']['total_models_supported']}")
    print(f"   • n150 Models: {data['summary']['n150_models']}")
    print(f"   • n300 Models: {data['summary']['n300_models']}")
    print(f"   • Performance Metrics: {len(data['performance_metrics'])}")

    if collector.generate_comprehensive_report(output_file, data):
        print(f"🎯 Comprehensive n150/n300 report generated successfully!")
        print(f"📁 All results organized in: {unified_base}")
        return 0
    else:
        print("❌ Failed to generate comprehensive report")
        return 1


if __name__ == "__main__":
    exit(main())
