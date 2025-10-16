#!/usr/bin/env python3
"""
Common utilities shared across Polaris benchmarking scripts.

This module contains shared functions for:
- Git repository operations
- Excel file generation and formatting
- Model name normalization
- File system operations
"""

import subprocess
import shutil
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
import logging

logger = logging.getLogger(__name__)


# Git Operations
def pull_git_repo(repo_name: str, repo_path: Path, clean_repo: bool = False) -> bool:
    """
    Pull or clone a Git repository from GitHub.

    Args:
        repo_name: Repository name (e.g., 'tenstorrent/polaris')
        repo_path: Local path where repo should be located
        clean_repo: If True, remove existing repo and re-clone

    Returns:
        bool: True if successful
    """
    logger.info(f"Setting up {repo_name} repository at {repo_path}")

    # If the path is already inside a git repo (has .git in any parent), just pull at the repo root
    try:
        probe = repo_path
        for _ in range(6):
            if (probe / ".git").exists():
                repo_path = probe
                break
            if probe.parent == probe:
                break
            probe = probe.parent
    except Exception:
        pass

    if clean_repo:
        logger.info("Cleaning existing repository...")
        if repo_path.exists():
            shutil.rmtree(repo_path)
        logger.info(f"Cloning {repo_name} repository...")
        try:
            subprocess.run(["git", "clone", f"https://github.com/{repo_name}.git", str(repo_path)],
                         check=True, capture_output=True)
            logger.info("Successfully cloned repository.")
        except subprocess.CalledProcessError as e:
            logger.error(f"Failed to clone repository: {e}")
            return False
    else:
        # Check if this is a valid git repository
        is_valid_git_repo = repo_path.exists() and (repo_path / ".git").exists()

        if is_valid_git_repo:
            logger.info("Repository exists, pulling latest changes...")
            try:
                result = subprocess.run(["git", "pull"], cwd=str(repo_path),
                                      capture_output=True, text=True)
                if result.returncode == 0:
                    logger.info("Successfully pulled latest changes.")
                else:
                    logger.warning(f"Git pull returned non-zero exit code: {result.stderr}")
            except Exception as e:
                logger.error(f"Failed to pull repository: {e}")
                return False
        else:
            logger.info("Repository does not exist or is invalid, cloning...")
            # Remove the directory if it exists but is not a valid git repo
            if repo_path.exists():
                logger.info("Removing invalid repository directory...")
                shutil.rmtree(repo_path)
            try:
                subprocess.run(["git", "clone", f"https://github.com/{repo_name}.git", str(repo_path)],
                             check=True, capture_output=True)
                logger.info("Successfully cloned repository.")
            except subprocess.CalledProcessError as e:
                logger.error(f"Failed to clone repository: {e}")
                return False

    # Verify repository structure
    if not repo_path.exists():
        logger.error("Repository directory was not created")
        return False

    logger.info(f"Repository setup complete: {repo_path}")
    return True


# Excel Utilities
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    def create_excel_workbook() -> Workbook:
        """Create a new Excel workbook."""
        return Workbook()

    def setup_excel_headers(ws: Worksheet, headers: List[str], row: int = 1) -> None:
        """Setup column headers in an Excel worksheet."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    def apply_cell_formatting(ws: Worksheet, row: int, col: int,
                            font: Optional[Font] = None,
                            fill: Optional[PatternFill] = None,
                            alignment: Optional[Alignment] = None) -> None:
        """Apply formatting to a specific cell."""
        cell = ws.cell(row=row, column=col)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment

    def adjust_column_widths(ws: Worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def save_excel_workbook(wb: Workbook, output_path: Path) -> bool:
        """Save Excel workbook to file."""
        try:
            wb.save(str(output_path))
            logger.info(f"Excel file saved: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            return False

    # Color constants for consistent styling
    EXCEL_COLORS = {
        'header': PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
        'hw_data': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
        'polaris_data': PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"),
        'difference': PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
    }

except ImportError:
    logger.warning("openpyxl not available, Excel utilities disabled")
    Workbook = None


# Model Normalization
def normalize_model_name(model_key: str) -> str:
    """
    Normalize model names for better matching between TT-Metal and Polaris.

    Args:
        model_key: Raw model key string

    Returns:
        Normalized model key string
    """
    # Handle ResNet variations
    if 'ResNet' in model_key or 'RESNET' in model_key.upper():
        # TT-Metal: "ResNet_n150" -> "resnet50_n150"
        # Polaris: "RESNET50_rn50_b1_hd_n150" -> "resnet50_n150"
        parts = model_key.split('_')
        arch = parts[-1]  # n150 or n300

        # Both TT-Metal "ResNet" and Polaris "RESNET50" should normalize to "resnet50_{arch}"
        return f"resnet50_{arch}".lower()

    # Handle BERT variations
    if 'BERT' in model_key.upper():
        if 'bert_base' in model_key.lower():
            return f"bert_base_{model_key.split('_')[-1]}".lower()
        elif 'bert_large' in model_key.lower():
            return f"bert_large_{model_key.split('_')[-1]}".lower()

    # Handle Llama/LLM variations
    if 'llama' in model_key.lower() or 'basic_llm' in model_key.lower():
        parts = model_key.split('_')
        arch = parts[-1]  # n150 or n300

        # For Polaris basic_llm.gpt_mini -> llama_gpt_mini
        if 'basic_llm' in model_key.lower():
            return f"llama_gpt_mini_{arch}".lower()
        # For Polaris llama2 models -> llama2_{variant}
        elif 'llama2' in model_key.lower():
            # Extract variant from the instance name (e.g., llama2_small -> small)
            if 'small' in model_key.lower():
                return f"llama2_small_{arch}".lower()
            elif 'standard' in model_key.lower():
                return f"llama2_standard_{arch}".lower()
            else:
                return f"llama2_{arch}".lower()
        # For TT-Metal Llama models -> try to match by size
        elif 'llama' in model_key.lower():
            # Extract size info (7B, 8B, 11B, etc.)
            if '7b' in model_key.lower():
                return f"llama_7b_{arch}".lower()
            elif '8b' in model_key.lower():
                return f"llama_8b_{arch}".lower()
            elif '11b' in model_key.lower():
                return f"llama_11b_{arch}".lower()
            elif '70b' in model_key.lower():
                return f"llama_70b_{arch}".lower()
            elif '1b' in model_key.lower():
                return f"llama_1b_{arch}".lower()
            elif '3b' in model_key.lower():
                return f"llama_3b_{arch}".lower()
            else:
                return f"llama_{arch}".lower()

    # Handle YOLO variations - try to match version and size
    if 'YOLO' in model_key.upper():
        parts = model_key.split('_')
        arch = parts[-1]  # n150 or n300

        # For Polaris format: "YOLOv8_yolov8s_n300" -> extract "yolov8s"
        if len(parts) >= 3 and parts[0].startswith('YOLO') and parts[1].startswith('yolo'):
            yolo_variant = parts[1].lower()  # e.g., "yolov8s"
            return f"{yolo_variant}_{arch}".lower()
        # For TT-Metal format: "YOLOv8s_n150" -> extract "yolov8s"
        elif len(parts) >= 2 and parts[0].startswith('YOLO'):
            yolo_model = parts[0].lower()  # e.g., "yolov8s"
            return f"{yolo_model}_{arch}".lower()

    # Handle UNet variations - normalize to "unet_{arch}"
    if 'UNet' in model_key or 'unet' in model_key.lower():
        parts = model_key.split('_')
        arch = parts[-1]  # n150 or n300
        # Both TT-Metal "UNet" and Polaris "UNet_unet_b1" should normalize to "unet_{arch}"
        return f"unet_{arch}".lower()

    # Handle BEVDepth variations - normalize to "bevdepth_{arch}"
    if 'BEVDepth' in model_key or 'bevdepth' in model_key.lower():
        parts = model_key.split('_')
        arch = parts[-1]  # n150 or n300
        return f"bevdepth_{arch}".lower()

    # Handle SwinTransformer variations - normalize to "swintransformer_{arch}"
    if 'SwinTransformer' in model_key or 'swin' in model_key.lower():
        parts = model_key.split('_')
        arch = parts[-1]  # n150 or n300
        return f"swintransformer_{arch}".lower()

    # Default: return lowercase version
    return model_key.lower()


# File System Utilities
def ensure_directory(path: Path) -> None:
    """Ensure a directory exists, creating it if necessary."""
    path.mkdir(parents=True, exist_ok=True)

def safe_copy_file(src: Path, dst: Path) -> bool:
    """Safely copy a file with error handling."""
    try:
        shutil.copy2(src, dst)
        return True
    except Exception as e:
        logger.error(f"Failed to copy {src} to {dst}: {e}")
        return False

def archive_directory(src_dir: Path, dst_dir: Path, archive_name: str) -> bool:
    """Archive a directory to another location."""
    try:
        dst_path = dst_dir / archive_name
        if dst_path.exists():
            shutil.rmtree(dst_path)
        shutil.copytree(src_dir, dst_path, dirs_exist_ok=True)
        logger.info(f"Archived {src_dir} to {dst_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to archive {src_dir} to {dst_dir}: {e}")
        return False


# Data Processing Utilities
def select_best_entry(entries: List[Dict], key_func=None) -> Optional[Dict]:
    """
    Select the best entry from a list based on a key function.

    Args:
        entries: List of dictionary entries
        key_func: Function to extract comparison key (default: None)

    Returns:
        Best entry or None if list is empty
    """
    if not entries:
        return None
    if len(entries) == 1:
        return entries[0]

    # Default: return first entry if no key function provided
    if key_func is None:
        return entries[0]

    return max(entries, key=key_func)


def validate_model_data(data: Dict, required_fields: List[str]) -> bool:
    """
    Validate that model data contains all required fields.

    Args:
        data: Model data dictionary
        required_fields: List of required field names

    Returns:
        True if all fields are present and non-empty
    """
    for field in required_fields:
        if field not in data or data[field] is None or str(data[field]).strip() == '':
            logger.warning(f"Missing or empty required field: {field}")
            return False
    return True


def create_orderly_output_structure(base_dir: Path) -> Dict[str, Path]:
    """
    Create an orderly folder structure for HW vs Polaris comparison reports.

    Structure:
    HW_Polaris comparison reports/
    ├── Logs (HW & Polaris)/
    ├── Reports (Excel sheets)/
    └── Polaris runs/

    Args:
        base_dir: Base directory path

    Returns:
        Dictionary mapping structure names to Path objects
    """
    logger.info(f"Creating orderly output structure in: {base_dir}")

    # Define the structure
    structure = {
        'root': base_dir,
        'logs': base_dir / "Logs (HW & Polaris)",
        'reports': base_dir / "Reports (Excel sheets)",
        'polaris_runs': base_dir / "Polaris runs",
        # Legacy compatibility paths (tt_metal_hw_results removed since reports go to Reports folder)
        'polaris_simulation_results': base_dir / "Polaris runs" / "polaris_simulation_results",
        'comparison_reports': base_dir / "Reports (Excel sheets)",
        'raw_data': base_dir / "Polaris runs" / "raw_data"
    }

    # Create all directories
    for name, path in structure.items():
        if name != 'root':  # Don't create the root again
            ensure_directory(path)
            logger.debug(f"Created directory: {path}")

    logger.info(f"Orderly output structure ready: {base_dir}")
    return structure
