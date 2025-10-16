#!/usr/bin/env python3
"""
HW vs Polaris Performance Comparison Analyzer

Comprehensive tool for comparing TT-Metal hardware performance against Polaris simulation results,
generating detailed Excel reports that show simulation accuracy vs real silicon execution.

WORKFLOW:
1. Run comprehensive_n150_n300_parser.py to extract TT-Metal HW metrics from GitHub
2. Run polaris_workload_automation.py to generate simulation results
3. Generate unified comparison report showing HW vs Polaris metrics

USAGE MODES:
1. Complete HW vs Polaris Workflow (Recommended):
   python results_comparison_analyzer.py --run-hw-vs-polaris

2. Legacy Polaris-only Comparison:
   python results_comparison_analyzer.py --legacy-mode --optimized-dir DIR1 --full-dir DIR2

3. Dry Run: Test workflow without executing
   python results_comparison_analyzer.py --run-hw-vs-polaris --dry-run

FEATURES:
- TT-Metal HW metrics extraction from GitHub repositories
- Polaris simulation result generation and parsing
- Unified HW vs Simulation performance comparison
- Excel reports with HW/Polaris source identification and color coding
- Throughput accuracy analysis (Polaris estimates vs HW measurements)
- Model support matrix integration
- Comprehensive error handling and retry logic
- Validation of generated results

"""

import argparse
import json
import logging
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime
import re
import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

# Import common utilities
from common_utils import (
    normalize_model_name, create_excel_workbook, setup_excel_headers,
    apply_cell_formatting, adjust_column_widths, save_excel_workbook,
    ensure_directory, create_orderly_output_structure, EXCEL_COLORS
)


@dataclass
class RunResult:
    """Represents a single workload execution result"""
    model_name: str
    model_variant: str
    architecture: str
    run_type: str  # 'optimized' or 'full'
    success: bool
    execution_time: float
    metrics: Dict[str, Any]
    raw_data: Dict[str, Any]


@dataclass
class ComparisonResult:
    """Represents comparison between two runs for the same model"""
    model_name: str
    model_variant: str
    architecture: str
    optimized_result: Optional[RunResult]
    full_result: Optional[RunResult]
    metric_differences: Dict[str, Dict[str, Any]]


class ResultsComparisonAnalyzer:
    """Analyzes and compares profiling results from two Polaris runs"""

    def __init__(self, optimized_dir: Path, full_dir: Path, log_dir: Optional[Path] = None):
        self.optimized_dir = optimized_dir
        self.full_dir = full_dir
        self.optimized_results: List[RunResult] = []
        self.full_results: List[RunResult] = []
        self.comparisons: List[ComparisonResult] = []

        # Setup logging
        logger.remove()
        if log_dir:
            log_file = log_dir / f"results_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            logger.add(
                log_file,
                level="INFO",
                format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}"
            )
        logger.add(lambda msg: print(msg, end=""), level="INFO")

    def parse_run_results(self, run_dir: Path, run_type: str) -> List[RunResult]:
        """Parse all results from a single run directory, prioritizing SUMMARY JSON files"""
        results = []
        raw_results_dir = run_dir / "raw_polaris_results"

        if not raw_results_dir.exists():
            logger.warning(f"Raw results directory not found: {raw_results_dir}")
            return results

        logger.info(f"Parsing {run_type} run results from {raw_results_dir} (prioritizing SUMMARY data)")

        # For unified structure, results might be in raw_data/raw_polaris_results
        if not raw_results_dir.exists() and run_type == "polaris":
            # Try the unified structure path
            unified_raw_dir = run_dir.parent / "raw_data" / "raw_polaris_results"
            if unified_raw_dir.exists():
                raw_results_dir = unified_raw_dir
                logger.info(f"Using unified structure path: {raw_results_dir}")

        # PRIORITY 1: Find all study-summary.json files (highest quality data)
        summary_files_found = list(raw_results_dir.rglob("study-summary.json"))
        logger.info(f"Found {len(summary_files_found)} SUMMARY JSON files")

        for summary_file in summary_files_found:
            try:
                logger.debug(f"Processing SUMMARY file: {summary_file}")

                # Load JSON data from SUMMARY file (highest quality source)
                with open(summary_file, 'r') as f:
                    raw_data = json.load(f)

                # Extract model info from SUMMARY JSON data
                if 'summary' in raw_data and raw_data['summary']:
                    # Process each summary entry (Polaris can have multiple entries per file)
                    for summary_entry in raw_data['summary']:
                        model_name = summary_entry.get('wlname', '')
                        model_variant = summary_entry.get('wlinstance', '')
                        architecture = summary_entry.get('devname', '')

                        if not all([model_name, model_variant, architecture]):
                            logger.warning(f"Incomplete model info in SUMMARY {summary_file}: {model_name}, {model_variant}, {architecture}")
                            continue

                        # Extract execution time from run log if available
                        execution_time = self._extract_execution_time(run_dir, model_name, model_variant, architecture)
                        success = self._determine_success_from_summary(raw_data, summary_entry)

                        # Extract metrics directly from SUMMARY entry (highest quality)
                        metrics = self._extract_metrics_from_summary(summary_entry)

                        result = RunResult(
                            model_name=model_name,
                            model_variant=model_variant,
                            architecture=architecture,
                            run_type=run_type,
                            success=success,
                            execution_time=execution_time,
                            metrics=metrics,
                            raw_data=raw_data
                        )

                        results.append(result)
                        logger.debug(f"Parsed SUMMARY result: {model_name}.{model_variant} on {architecture} "
                                   f"(throughput: {metrics.get('ideal_throughput', 'N/A')})")

                else:
                    logger.warning(f"SUMMARY file {summary_file} has no valid summary data")

            except Exception as e:
                logger.error(f"Failed to parse SUMMARY file {summary_file}: {e}")
                continue

        logger.info(f"Successfully parsed {len(results)} results from SUMMARY files in {run_type} run")
        return results

    def _extract_execution_time(self, run_dir: Path, model_name: str, model_variant: str, architecture: str) -> float:
        """Extract execution time from run log file"""
        log_filename = f"{model_name}_{model_variant}_{architecture}_run_log.txt"
        log_file = run_dir / log_filename

        if not log_file.exists():
            return 0.0

        try:
            with open(log_file, 'r') as f:
                content = f.read()
                # Look for "Execution Time: X.XX seconds"
                match = re.search(r'Execution Time:\s*([\d.]+)\s*seconds', content)
                if match:
                    return float(match.group(1))
        except Exception as e:
            logger.warning(f"Could not extract execution time from {log_file}: {e}")

        return 0.0

    def _determine_success(self, raw_data: Dict[str, Any]) -> bool:
        """Determine if the run was successful based on raw data"""
        # Check if summary contains valid data
        summary_list = raw_data.get('summary', [])
        if not summary_list:
            return False

        summary = summary_list[0]
        # Basic success criteria: has cycles and time data
        return summary.get('tot_cycles', 0) > 0 and summary.get('tot_msecs', 0) > 0

    def _determine_success_from_summary(self, raw_data: Dict[str, Any], summary_entry: Dict[str, Any]) -> bool:
        """Determine if the run was successful based on SUMMARY entry data"""
        # Use SUMMARY data for success determination (highest quality)
        return (summary_entry.get('tot_cycles', 0) > 0 and
                summary_entry.get('tot_msecs', 0) > 0 and
                summary_entry.get('ideal_throughput', 0) > 0)

    def _extract_metrics(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract key metrics from raw Polaris data"""
        metrics = {}

        summary_list = raw_data.get('summary', [])
        if summary_list:
            summary = summary_list[0]
            metrics.update({
                'frequency_mhz': summary.get('freq_MHz', summary.get('frequency_MHz', summary.get('freq', 1000))),
                'total_cycles': summary.get('tot_cycles', summary.get('cycles', summary.get('total_cycles', 0))),
                'total_msecs': summary.get('tot_msecs', summary.get('msecs', summary.get('total_msecs', summary.get('time_ms', 0)))),
                'ideal_throughput': summary.get('ideal_throughput', 0),
                'memory_size_gb': summary.get('mem_size_GB', summary.get('memory_size', summary.get('mem_size', 0))),
                'device_memory_gb': summary.get('device_mem_size_GB', summary.get('device_memory', 0)),
                'fits_device': summary.get('fits_device', False),
                'performance_projection': summary.get('perf_projection', summary.get('performance', 0)),
                'resource_compute': summary.get('rsrc_comp', summary.get('compute_resource', 0)),
                'resource_memory': summary.get('rsrc_mem', summary.get('memory_resource', 0)),
                'batch_size': summary.get('batch', summary.get('batch_size', 1))
            })

        # Add operator count if available
        operator_stats = raw_data.get('operatorstats', [])
        metrics['operator_count'] = len(operator_stats)

        return metrics

    def _extract_metrics_from_summary(self, summary_entry: Dict[str, Any]) -> Dict[str, Any]:
        """Extract metrics directly from SUMMARY entry (highest quality data source)"""
        metrics = {}

        # Extract all available metrics from SUMMARY entry
        metrics.update({
            'frequency_mhz': summary_entry.get('freq_MHz', summary_entry.get('frequency_MHz', summary_entry.get('freq', 1000))),
            'total_cycles': summary_entry.get('tot_cycles', summary_entry.get('cycles', 0)),
            'total_msecs': summary_entry.get('tot_msecs', summary_entry.get('msecs', summary_entry.get('total_msecs', 0))),
            'ideal_throughput': summary_entry.get('ideal_throughput', 0),
            'memory_size_gb': summary_entry.get('mem_size_GB', summary_entry.get('memory_size', 0)),
            'device_memory_gb': summary_entry.get('device_memsize_GB', summary_entry.get('device_memory', 0)),
            'fits_device': summary_entry.get('fits_device', False),
            'performance_projection': summary_entry.get('perf_projection', summary_entry.get('performance', 0)),
            'resource_compute': summary_entry.get('rsrc_comp', summary_entry.get('compute_resource', 0)),
            'resource_memory': summary_entry.get('rsrc_mem', summary_entry.get('memory_resource', 0)),
            'batch_size': summary_entry.get('bs', summary_entry.get('batch', summary_entry.get('batch_size', 1))),
            'device_peak_bw_gbps': summary_entry.get('device_peak_bw_GBps', 0),
            'device_peak_fp8_tflops': summary_entry.get('device_peak_fp8_tflops', 0),
            'in_params': summary_entry.get('inParams', 0),
            'in_acts': summary_entry.get('inActs', 0),
            'out_acts': summary_entry.get('outActs', 0),
            'max_acts': summary_entry.get('maxActs', 0),
            'wlclass': summary_entry.get('wlcls', ''),
        })

        # Operator count from STATS file if available (fallback)
        stat_filename = summary_entry.get('stat_filename', '')
        if stat_filename:
            # Try to load operator count from the referenced STATS file
            try:
                # Construct path to STATS file relative to summary file
                stats_file_path = None
                # This is a simplified approach - in practice we'd need the full path
                metrics['operator_count'] = 0  # Will be populated if STATS file is accessible
            except:
                metrics['operator_count'] = 0
        else:
            metrics['operator_count'] = 0

        return metrics

    def perform_comparisons(self):
        """Compare results between optimized and full runs"""
        logger.info("Performing head-to-head comparisons...")

        # Create lookup dictionaries
        opt_lookup = {(r.model_name, r.model_variant, r.architecture): r for r in self.optimized_results}
        full_lookup = {(r.model_name, r.model_variant, r.architecture): r for r in self.full_results}

        # Get all unique model combinations
        all_keys = set(opt_lookup.keys()) | set(full_lookup.keys())

        for key in all_keys:
            model_name, model_variant, architecture = key
            opt_result = opt_lookup.get(key)
            full_result = full_lookup.get(key)

            # Calculate metric differences
            metric_differences = self._calculate_metric_differences(opt_result, full_result)

            comparison = ComparisonResult(
                model_name=model_name,
                model_variant=model_variant,
                architecture=architecture,
                optimized_result=opt_result,
                full_result=full_result,
                metric_differences=metric_differences
            )

            self.comparisons.append(comparison)

        logger.info(f"Generated {len(self.comparisons)} comparisons")

    def _calculate_metric_differences(self, opt_result: Optional[RunResult], full_result: Optional[RunResult]) -> Dict[str, Dict[str, Any]]:
        """Calculate differences between optimized and full run metrics"""
        differences = {}

        if not opt_result or not full_result:
            return differences

        # Compare execution times
        if opt_result.execution_time > 0 and full_result.execution_time > 0:
            time_diff = full_result.execution_time - opt_result.execution_time
            time_ratio = full_result.execution_time / opt_result.execution_time
            differences['execution_time'] = {
                'optimized': opt_result.execution_time,
                'full': full_result.execution_time,
                'difference': time_diff,
                'ratio': time_ratio,
                'percent_change': (time_ratio - 1) * 100
            }

        # Compare key metrics - separated by source
        hw_metrics = ['total_cycles', 'total_msecs', 'ideal_throughput']
        polaris_metrics = ['resource_compute', 'resource_memory', 'operator_count']

        for metric in hw_metrics + polaris_metrics:
            opt_val = opt_result.metrics.get(metric, 0)
            full_val = full_result.metrics.get(metric, 0)

            if opt_val > 0 and full_val > 0:
                diff = full_val - opt_val
                ratio = full_val / opt_val
                percent_change = (ratio - 1) * 100

                differences[metric] = {
                    'optimized': opt_val,
                    'full': full_val,
                    'difference': diff,
                    'ratio': ratio,
                    'percent_change': percent_change
                }

        return differences

    def generate_excel_report(self, output_path: Path) -> bool:
        """Generate comprehensive Excel comparison report"""
        logger.info(f"Generating Excel report: {output_path}")

        try:
            wb = create_excel_workbook()

            # Create worksheets
            self._create_legend_sheet(wb)
            self._create_summary_sheet(wb)
            self._create_head_to_head_sheet(wb)
            self._create_performance_analysis_sheet(wb)
            self._create_model_coverage_sheet(wb)

            # Remove default sheet if it exists
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Save workbook
            save_excel_workbook(wb, output_path)
            logger.info(f"Excel report saved successfully: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return False

    def _create_legend_sheet(self, wb: Workbook):
        """Create legend sheet explaining HW vs Polaris metrics"""
        ws = wb.create_sheet("Legend")

        ws['A1'] = "Metric Legend: HW vs Polaris"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # HW Metrics section
        ws['A3'] = "HARDWARE METRICS (HW)"
        ws['A3'].font = Font(size=14, bold=True)
        ws['A3'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        hw_metrics = [
            ["Exec Time (s)", "Actual execution time measured during Polaris run"],
            ["HW: Total Cycles", "Hardware cycle count from Tensix cores"],
            ["HW: Total Time (ms)", "Total time spent in hardware execution"],
            ["HW: Throughput", "Actual hardware throughput measurements"],
        ]

        row_idx = 4
        for metric, description in hw_metrics:
            ws.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=description)
            row_idx += 1

        # Polaris Metrics section
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="POLARIS SIMULATION METRICS").font = Font(size=14, bold=True)
        ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
        row_idx += 1

        polaris_metrics = [
            ["Polaris: Compute Resource", "Estimated compute resource utilization (simulation)"],
            ["Polaris: Memory Resource", "Estimated memory resource utilization (simulation)"],
            ["Polaris: Operators", "Count of neural network operators processed"],
            ["Batch Size", "Input batch size used in simulation"],
        ]

        for metric, description in polaris_metrics:
            ws.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=description)
            row_idx += 1

        # Key insights
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="KEY INSIGHTS:").font = Font(size=12, bold=True)
        row_idx += 1

        insights = [
            "• HW metrics come from actual Tensix core execution",
            "• Polaris metrics are simulation estimates and projections",
            "• HW metrics reflect real performance on silicon",
            "• Polaris metrics help with architectural planning",
            "• Compare HW metrics between runs for performance validation",
            "• Compare Polaris metrics between runs for design insights"
        ]

        for insight in insights:
            ws.cell(row=row_idx, column=1, value=insight)
            row_idx += 1

        self._adjust_column_widths(ws)

    def _create_summary_sheet(self, wb: Workbook):
        """Create summary overview sheet"""
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = "Polaris Profiling Results Comparison: HW vs Simulation Metrics"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        # Overview statistics
        ws['A3'] = "Overview Statistics"
        ws['A3'].font = Font(size=14, bold=True)

        note_cell = ws.cell(row=4, column=1, value="Note: HW metrics reflect actual Tensix core execution. Polaris metrics are simulation estimates.")
        note_cell.font = Font(italic=True, size=10)
        ws.merge_cells('A4:E4')

        data = [
            ["Metric", "Optimized Run", "Full Run", "Difference"],
            ["Total Models Tested", len(self.optimized_results), len(self.full_results),
             len(self.full_results) - len(self.optimized_results)],
            ["Successful Runs", sum(1 for r in self.optimized_results if r.success),
             sum(1 for r in self.full_results if r.success), None],
            ["Success Rate", f"{sum(1 for r in self.optimized_results if r.success)/len(self.optimized_results)*100:.1f}%" if self.optimized_results else "0%",
             f"{sum(1 for r in self.full_results if r.success)/len(self.full_results)*100:.1f}%" if self.full_results else "0%", None],
            ["Average Execution Time (s)", f"{sum(r.execution_time for r in self.optimized_results)/len(self.optimized_results):.2f}" if self.optimized_results else "0.00",
             f"{sum(r.execution_time for r in self.full_results)/len(self.full_results):.2f}" if self.full_results else "0.00", None],
        ]

        for row_idx, row in enumerate(data, 6):
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=row_idx, column=col_idx+1, value=value)
                if row_idx == 6:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

    def _create_head_to_head_sheet(self, wb: Workbook):
        """Create detailed head-to-head comparison sheet"""
        ws = wb.create_sheet("Head-to-Head Comparison")

        # Header
        headers = ["Model", "Variant", "Architecture", "Run Type", "Execution Source", "Success", "Exec Time (s)",
                  "HW: Total Cycles", "HW: Total Time (ms)", "HW: Throughput", "Polaris: Compute Resource", "Polaris: Memory Resource", "Polaris: Operators"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row_idx = 2

        # Filter to only include models that exist in both optimized and full runs
        filtered_comparisons = [
            comp for comp in self.comparisons
            if comp.optimized_result is not None and comp.full_result is not None
        ]

        logger.info(f"Filtered to {len(filtered_comparisons)} models that exist in both reports")

        # Sort comparisons by model name
        sorted_comparisons = sorted(filtered_comparisons, key=lambda x: (x.model_name, x.model_variant, x.architecture))

        for comparison in sorted_comparisons:
            # Optimized run row
            if comparison.optimized_result:
                self._add_result_row(ws, row_idx, comparison.optimized_result, "Optimized")
                row_idx += 1

            # Full run row
            if comparison.full_result:
                self._add_result_row(ws, row_idx, comparison.full_result, "Full")
                row_idx += 1

            # Difference row if both exist
            if comparison.optimized_result and comparison.full_result:
                self._add_difference_row(ws, row_idx, comparison)
                row_idx += 1

            # Separator
            row_idx += 1

        self._adjust_column_widths(ws)

    def _add_result_row(self, ws, row_idx: int, result: RunResult, run_label: str):
        """Add a single result row to the worksheet"""
        # Determine execution source based on HW metrics presence
        execution_source = "Polaris Simulation"  # Default
        if result.metrics.get('total_cycles', 0) > 0 or result.metrics.get('total_msecs', 0) > 0:
            execution_source = "HW Execution"

        data = [
            result.model_name,
            result.model_variant,
            result.architecture,
            run_label,
            execution_source,
            "Yes" if result.success else "No",
            f"{result.execution_time:.2f}",
            result.metrics.get('total_cycles', 0),
            f"{result.metrics.get('total_msecs', 0):.2f}",
            result.metrics.get('ideal_throughput', 0),
            f"{result.metrics.get('resource_compute', 0):.2f}",
            f"{result.metrics.get('resource_memory', 0):.2f}",
            result.metrics.get('operator_count', 0)
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Color coding based on run type
            if run_label == "Optimized":
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            elif run_label == "Full":
                cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")

    def _add_difference_row(self, ws, row_idx: int, comparison: ComparisonResult):
        """Add a difference row showing comparison between runs"""
        diff_data = ["", "", "", "Difference", "", "", "", "", "", "", "", "", ""]

        # Calculate key differences
        if 'execution_time' in comparison.metric_differences:
            diff_data[6] = f"{comparison.metric_differences['execution_time']['difference']:+.2f}"

        if 'total_cycles' in comparison.metric_differences:
            diff_data[7] = f"{comparison.metric_differences['total_cycles']['difference']:+.0f}"

        if 'total_msecs' in comparison.metric_differences:
            diff_data[8] = f"{comparison.metric_differences['total_msecs']['difference']:+.2f}"

        for col_idx, value in enumerate(diff_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.font = Font(italic=True)

    def _create_performance_analysis_sheet(self, wb: Workbook):
        """Create performance analysis sheet with insights"""
        ws = wb.create_sheet("Performance Analysis")

        ws['A1'] = "Performance Analysis & Insights"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Performance insights
        insights = [
            "Key Findings:",
            "",
            f"• Total models compared: {len([c for c in self.comparisons if c.optimized_result and c.full_result])}",
            f"• Models only in optimized run: {len([c for c in self.comparisons if c.optimized_result and not c.full_result])}",
            f"• Models only in full run: {len([c for c in self.comparisons if not c.optimized_result and c.full_result])}",
            "",
            "Performance Insights:",
        ]

        row_idx = 3
        for insight in insights:
            ws.cell(row=row_idx, column=1, value=insight)
            row_idx += 1

        # Calculate average differences
        exec_time_diffs = [c.metric_differences.get('execution_time', {}).get('ratio', 1)
                          for c in self.comparisons if 'execution_time' in c.metric_differences]

        if exec_time_diffs:
            avg_time_ratio = sum(exec_time_diffs) / len(exec_time_diffs)
            ws.cell(row=row_idx, column=1,
                   value=f"• Average execution time ratio (Full/Optimized): {avg_time_ratio:.2f}x")
            row_idx += 1

        self._adjust_column_widths(ws)

    def _create_model_coverage_sheet(self, wb: Workbook):
        """Create model coverage comparison sheet"""
        ws = wb.create_sheet("Model Coverage")

        ws['A1'] = "Model Coverage Comparison"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')

        headers = ["Model", "Variant", "Optimized Run", "Full Run", "Coverage"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row_idx = 3
        for comparison in sorted(self.comparisons, key=lambda x: (x.model_name, x.model_variant)):
            # Group by model/variant, show architectures
            opt_archs = []
            full_archs = []

            # Find all architectures for this model/variant
            for comp in self.comparisons:
                if comp.model_name == comparison.model_name and comp.model_variant == comparison.model_variant:
                    if comp.optimized_result:
                        opt_archs.append(comp.architecture)
                    if comp.full_result:
                        full_archs.append(comp.architecture)

            if opt_archs or full_archs:
                ws.cell(row=row_idx, column=1, value=comparison.model_name)
                ws.cell(row=row_idx, column=2, value=comparison.model_variant)
                ws.cell(row=row_idx, column=3, value=", ".join(sorted(set(opt_archs))))
                ws.cell(row=row_idx, column=4, value=", ".join(sorted(set(full_archs))))

                # Coverage indicator
                if set(opt_archs) == set(full_archs):
                    coverage = "Complete"
                    fill_color = "C8E6C9"  # Light green
                elif set(opt_archs).issubset(set(full_archs)):
                    coverage = "Subset"
                    fill_color = "FFF3E0"  # Light orange
                else:
                    coverage = "Partial"
                    fill_color = "FFEBEE"  # Light red

                cell = ws.cell(row=row_idx, column=5, value=coverage)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                row_idx += 1

        self._adjust_column_widths(ws)

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                # Skip merged cells to avoid errors
                if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                    continue

                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def run_workload_automation(self, automation_script: Path, output_dir: Path,
                               workload_configs: Optional[str] = None,
                               max_retries: int = 3, dry_run: bool = False,
                               run_type: str = "default") -> bool:
        """Run the workload automation script with proper error handling and retries

        Args:
            automation_script: Path to the polaris_workload_automation.py script
            output_dir: Directory to store results
            workload_configs: Comma-separated workload config files (optional)
            max_retries: Maximum number of retry attempts
            dry_run: If True, only show what would be done without executing
            run_type: Type of run ("optimized", "full", etc.) for unique output directories

        Returns:
            bool: True if successful, False otherwise
        """
        logger.info(f"Running workload automation: {automation_script}")
        logger.info(f"Output directory: {output_dir}")
        if workload_configs:
            logger.info(f"Workload configs: {workload_configs}")

        # Build command
        cmd = [sys.executable, str(automation_script), "--output-dir", str(output_dir), "--run-type", run_type]
        if workload_configs:
            cmd.extend(["--workload-configs", workload_configs])

        if dry_run:
            logger.info(f"DRY RUN: Would execute: {' '.join(cmd)}")
            logger.info("DRY RUN: Skipping actual execution")
            return True

        for attempt in range(max_retries):
            try:
                logger.info(f"Attempt {attempt + 1}/{max_retries}: Executing {' '.join(cmd)}")

                # Run the command with timeout
                result = subprocess.run(
                    cmd,
                    cwd=str(automation_script.parent),
                    capture_output=True,
                    text=True,
                    timeout=3600  # 1 hour timeout
                )

                if result.returncode == 0:
                    logger.info("Workload automation completed successfully")
                    logger.debug(f"STDOUT: {result.stdout}")
                    return True
                else:
                    logger.warning(f"Workload automation failed with return code {result.returncode}")
                    logger.warning(f"STDERR: {result.stderr}")
                    logger.warning(f"STDOUT: {result.stdout}")

                    # Check for specific error patterns
                    if "not a git repository" in result.stderr or "not a git repository" in result.stdout:
                        logger.error("Repository not properly initialized. Please run setup first.")
                        return False
                    elif "No such file or directory" in result.stderr:
                        logger.error("Missing required files. Check configuration.")
                        return False

            except subprocess.TimeoutExpired:
                logger.error(f"Workload automation timed out after 1 hour (attempt {attempt + 1})")
            except Exception as e:
                logger.error(f"Unexpected error running workload automation: {e}")

            # Wait before retry (except on last attempt)
            if attempt < max_retries - 1:
                wait_time = 10 * (attempt + 1)  # Progressive backoff
                logger.info(f"Waiting {wait_time} seconds before retry...")
                time.sleep(wait_time)

        logger.error(f"Workload automation failed after {max_retries} attempts")
        return False

    def run_hw_vs_polaris_comparison_workflow(self, tt_metal_parser_script: Path,
                                             polaris_script: Path,
                                             tt_metal_output_dir: Path,
                                             polaris_output_dir: Path,
                                             report_output_path: Path,
                                             force_regenerate: bool = False,
                                             dry_run: bool = False,
                                             max_retries: int = 3,
                                             unified_output_dir: Path = None,
                                             local_workload_config: Path = None) -> bool:
        """Run the complete HW vs Polaris comparison workflow with unified output structure

        Args:
            tt_metal_parser_script: Path to comprehensive_n150_n300_parser.py
            polaris_script: Path to polaris_workload_automation.py
            tt_metal_output_dir: Output directory for tt-metal HW metrics (within unified structure)
            polaris_output_dir: Output directory for Polaris simulation results (within unified structure)
            report_output_path: Path for the final comparison report (within unified structure)
            force_regenerate: If True, always regenerate results even if they exist
            dry_run: If True, only show what would be done without executing
            max_retries: Maximum retry attempts for failed runs
            unified_output_dir: Optional unified output directory to pass to the script
            local_workload_config: Optional path to local workload config file

        Returns:
            bool: True if the complete workflow succeeded
        """
        logger.info("=== Starting HW vs Polaris Comparison Workflow ===")
        logger.info(f"TT-Metal parser: {tt_metal_parser_script}")
        logger.info(f"Polaris script: {polaris_script}")
        logger.info(f"TT-Metal results: {tt_metal_output_dir}")
        logger.info(f"Polaris results: {polaris_output_dir}")
        logger.info(f"Comparison report: {report_output_path}")

        # Ensure unified output structure exists
        structure_paths = self._ensure_unified_output_structure(unified_output_dir, unified_output_dir)

        # Step 1: Generate tt-metal HW metrics (if needed)
        logger.info("=== Step 1: Generating TT-Metal HW Metrics ===")
        if not self.run_tt_metal_parser(
            tt_metal_parser_script,
            tt_metal_output_dir,
            force_regenerate,
            dry_run,
            max_retries,
            unified_output_dir  # Pass the unified output directory
        ):
            logger.error("Failed to generate TT-Metal HW metrics")
            return False

        # Step 2: Generate Polaris simulation results (if needed)
        logger.info("=== Step 2: Generating Polaris Simulation Results ===")
        if not self.run_polaris_workload_automation(
            polaris_script,
            polaris_output_dir,
            force_regenerate,
            dry_run,
            max_retries,
            unified_output_dir,  # Pass the unified output directory
            local_workload_config  # Pass the local workload config
        ):
            logger.error("Failed to generate Polaris simulation results")
            return False

        # Step 3: Load and compare results
        logger.info("=== Step 3: Loading and Comparing Results ===")

        # Load TT-Metal metrics
        tt_metal_metrics = self.load_tt_metal_metrics(tt_metal_output_dir)
        if not tt_metal_metrics:
            logger.error("Failed to load TT-Metal metrics")
            return False

        # Load Polaris metrics
        polaris_metrics = self.load_polaris_metrics(polaris_output_dir)
        if not polaris_metrics:
            logger.error("Failed to load Polaris metrics")
            return False

        # Create comparison
        self.create_hw_vs_polaris_comparison(tt_metal_metrics, polaris_metrics)

        # Generate unified report
        if self.generate_hw_vs_polaris_report(report_output_path):
            logger.info("=== HW vs Polaris Comparison Workflow Completed Successfully ===")
            logger.info(f"Report generated: {report_output_path}")
            return True
        else:
            logger.error("Failed to generate comparison report")
            return False

    def _ensure_unified_output_structure(self, base_dir: Path, polaris_base_dir: Path) -> Dict[str, Path]:
        """Ensure unified output directory structure exists

        Args:
            base_dir: Base directory for unified output structure
            polaris_base_dir: Base directory for Polaris results (may be different)

        Returns:
            Dictionary mapping structure names to Path objects
        """
        # Use the common orderly structure function
        return create_orderly_output_structure(base_dir)

    def run_tt_metal_parser(self, parser_script: Path, output_dir: Path,
                           force_regenerate: bool = False, dry_run: bool = False,
                           max_retries: int = 3, unified_output_dir: Path = None) -> bool:
        """Run the TT-Metal comprehensive parser to extract HW metrics

        Args:
            parser_script: Path to comprehensive_n150_n300_parser.py
            output_dir: Directory to store results
            force_regenerate: If True, always regenerate results
            dry_run: If True, only show what would be done without executing
            max_retries: Maximum retry attempts

        Returns:
            bool: True if successful
        """
        # Check if results already exist and are valid
        excel_files = [f for f in output_dir.glob("*n150_n300_report*.xlsx")
                      if not f.name.startswith('~$')]
        if not force_regenerate and excel_files:
            logger.info(f"TT-Metal metrics already exist: {excel_files[0]}")
            return True

        if dry_run:
            cmd_str = f"{sys.executable} {parser_script}"
            if unified_output_dir:
                cmd_str += f" --unified-output-dir {unified_output_dir}"
            logger.info(f"DRY RUN: Would execute: {cmd_str}")
            return True

        for attempt in range(max_retries):
            try:
                logger.info(f"Attempt {attempt + 1}/{max_retries}: Running TT-Metal parser")

                # Build command with unified output directory if provided
                cmd = [sys.executable, str(parser_script)]
                if unified_output_dir:
                    cmd.extend(["--unified-output-dir", str(unified_output_dir)])

                # Run the parser
                result = subprocess.run(
                    cmd,
                    cwd=str(parser_script.parent),
                    capture_output=True,
                    text=True,
                    timeout=1800  # 30 minutes timeout
                )

                if result.returncode == 0:
                    logger.info("TT-Metal parser completed successfully")
                    logger.debug(f"STDOUT: {result.stdout}")
                    return True
                else:
                    logger.warning(f"TT-Metal parser failed with return code {result.returncode}")
                    logger.warning(f"STDERR: {result.stderr}")

            except subprocess.TimeoutExpired:
                logger.error(f"TT-Metal parser timed out after 30 minutes (attempt {attempt + 1})")
            except Exception as e:
                logger.error(f"Unexpected error running TT-Metal parser: {e}")

            if attempt < max_retries - 1:
                wait_time = 10 * (attempt + 1)
                logger.info(f"Waiting {wait_time} seconds before retry...")
                time.sleep(wait_time)

        logger.error(f"TT-Metal parser failed after {max_retries} attempts")
        return False

    def run_polaris_workload_automation(self, polaris_script: Path, output_dir: Path,
                                       force_regenerate: bool = False, dry_run: bool = False,
                                       max_retries: int = 3, unified_output_dir: Path = None,
                                       local_workload_config: Path = None) -> bool:
        """Run the Polaris workload automation to generate simulation results

        Args:
            polaris_script: Path to polaris_workload_automation.py
            output_dir: Directory to store results
            force_regenerate: If True, always regenerate results
            dry_run: If True, only show what would be done without executing
            max_retries: Maximum retry attempts
            unified_output_dir: Optional unified output directory to pass to the script
            local_workload_config: Optional path to local workload config file

        Returns:
            bool: True if successful
        """
        # Check if results already exist and are valid
        if not force_regenerate and self._check_results_validity(output_dir):
            logger.info("Polaris simulation results already exist and are valid")
            return True

        if dry_run:
            cmd_str = f"{sys.executable} {polaris_script} --output-dir {output_dir}"
            if unified_output_dir:
                cmd_str += f" --unified-output-dir {unified_output_dir}"
            if local_workload_config:
                cmd_str += f" --local-workload-config {local_workload_config}"
            logger.info(f"DRY RUN: Would execute: {cmd_str}")
            return True

        for attempt in range(max_retries):
            try:
                logger.info(f"Attempt {attempt + 1}/{max_retries}: Running Polaris workload automation")

                # Build command with unified output directory and local workload config if provided
                cmd = [sys.executable, str(polaris_script), "--output-dir", str(output_dir)]
                if unified_output_dir:
                    cmd.extend(["--unified-output-dir", str(unified_output_dir)])
                if local_workload_config:
                    cmd.extend(["--local-workload-config", str(local_workload_config)])

                # Run the automation script
                result = subprocess.run(
                    cmd,
                    cwd=str(polaris_script.parent),
                    capture_output=True,
                    text=True,
                    timeout=3600  # 1 hour timeout
                )

                if result.returncode == 0:
                    logger.info("Polaris workload automation completed successfully")
                    logger.debug(f"STDOUT: {result.stdout}")
                    return True
                else:
                    logger.warning(f"Polaris automation failed with return code {result.returncode}")
                    logger.warning(f"STDERR: {result.stderr}")

            except subprocess.TimeoutExpired:
                logger.error(f"Polaris automation timed out after 1 hour (attempt {attempt + 1})")
            except Exception as e:
                logger.error(f"Unexpected error running Polaris automation: {e}")

            if attempt < max_retries - 1:
                wait_time = 10 * (attempt + 1)
                logger.info(f"Waiting {wait_time} seconds before retry...")
                time.sleep(wait_time)

        logger.error(f"Polaris workload automation failed after {max_retries} attempts")
        return False

    def load_tt_metal_metrics(self, output_dir: Path) -> Optional[Dict[str, Any]]:
        """Load TT-Metal metrics from the comprehensive parser output

        Args:
            output_dir: Directory containing the TT-Metal parser results

        Returns:
            Dict containing model support and performance metrics
        """
        try:
            # The parser saves reports directly in the Reports folder
            search_dirs = [output_dir, Path("./polaris_results")]

            latest_report = None
            for search_dir in search_dirs:
                if search_dir.exists():
                    # Filter out temporary Excel files (starting with ~$)
                    excel_files = [f for f in search_dir.glob("*n150_n300_report*.xlsx")
                                 if not f.name.startswith('~$')]
                    if excel_files:
                        # Find the latest in this directory
                        dir_latest = max(excel_files, key=lambda x: x.stat().st_mtime)
                        if latest_report is None or dir_latest.stat().st_mtime > latest_report.stat().st_mtime:
                            latest_report = dir_latest

            if not latest_report:
                logger.error("No TT-Metal report found in any search location")
                return None

            logger.info(f"Loading TT-Metal metrics from: {latest_report}")

            # Read the performance metrics sheet - it has a title before the actual header
            df_raw = pd.read_excel(latest_report, sheet_name="Performance Metrics", header=None)

            # Find the header row (where 'Model' is in the first column)
            header_row_idx = None
            for i, row in df_raw.iterrows():
                if str(row.iloc[0]).strip() == 'Model':
                    header_row_idx = i
                    break

            if header_row_idx is None:
                logger.error("Could not find header row in TT-Metal metrics sheet")
                return None

            # Extract data starting from header row
            df = df_raw.iloc[header_row_idx:]
            df.columns = df.iloc[0]  # Set column names from header row
            df = df.iloc[1:]  # Skip the header row itself
            df = df.reset_index(drop=True)  # Reset index

            logger.info(f"Found {len(df)} TT-Metal performance records")

            metrics = {}
            for _, row in df.iterrows():
                try:
                    model_key = f"{row['Model']}_{row['Variant']}_{row['Hardware']}"
                    metrics[model_key] = {
                        'model_name': str(row['Model']).strip(),
                        'model_variant': str(row['Variant']).strip() if pd.notna(row['Variant']) else '',
                        'hardware': str(row['Hardware']).strip(),
                        'metric_type': str(row['Metric Type']).strip() if pd.notna(row['Metric Type']) else '',
                        'value': float(row['Value']) if pd.notna(row['Value']) and str(row['Value']).strip() else 0.0,
                        'unit': str(row['Unit']).strip() if pd.notna(row['Unit']) else '',
                        'target': str(row['Target']).strip() if pd.notna(row['Target']) and str(row['Target']).strip() else None,
                        'batch_size': int(row['Batch Size']) if pd.notna(row['Batch Size']) and str(row['Batch Size']).strip() else None,
                        'source': 'TT-Metal HW'
                    }
                except (KeyError, ValueError) as e:
                    logger.warning(f"Skipping malformed row: {row}, error: {e}")
                    continue

            logger.info(f"Loaded {len(metrics)} TT-Metal performance metrics")
            return metrics

        except Exception as e:
            logger.error(f"Failed to load TT-Metal metrics: {e}")
            return None

    def load_polaris_metrics(self, output_dir: Path) -> Optional[Dict[str, Any]]:
        """Load Polaris simulation metrics from SUMMARY JSON files (highest quality data)

        Args:
            output_dir: Directory containing Polaris results

        Returns:
            Dict containing Polaris simulation metrics
        """
        try:
            # Parse the raw Polaris results using SUMMARY-focused method
            results = self.parse_run_results(output_dir, "polaris")

            if not results:
                logger.warning("No SUMMARY-based results found")
                return None

            # Group results by unique model key and select the best (highest throughput) SUMMARY result
            # This handles cases where multiple variants (HD/UHD) normalize to the same key
            metrics = {}
            grouped_results = {}

            for result in results:
                if result.success and result.metrics:
                    model_key = f"{result.model_name}_{result.model_variant}_{result.architecture}"
                    throughput = result.metrics.get('ideal_throughput', 0)

                    if model_key not in grouped_results:
                        grouped_results[model_key] = []
                    grouped_results[model_key].append((result, throughput))

            # For each model key, select the SUMMARY result with highest throughput
            for model_key, result_list in grouped_results.items():
                # Sort by throughput (descending) and pick the best SUMMARY result
                best_result, best_throughput = max(result_list, key=lambda x: x[1])

                # Extract comprehensive metrics from best SUMMARY data
                summary_metrics = best_result.metrics
                metrics[model_key] = {
                    'model_name': best_result.model_name,
                    'model_variant': best_result.model_variant,
                    'hardware': best_result.architecture,
                    'execution_time': best_result.execution_time,
                    'total_cycles': summary_metrics.get('total_cycles', 0),
                    'total_msecs': summary_metrics.get('total_msecs', 0),
                    'ideal_throughput': summary_metrics.get('ideal_throughput', 0),
                    'resource_compute': summary_metrics.get('resource_compute', 0),
                    'resource_memory': summary_metrics.get('resource_memory', 0),
                    'operator_count': summary_metrics.get('operator_count', 0),
                    'device_peak_bw_gbps': summary_metrics.get('device_peak_bw_gbps', 0),
                    'device_peak_fp8_tflops': summary_metrics.get('device_peak_fp8_tflops', 0),
                    'fits_device': summary_metrics.get('fits_device', False),
                    'performance_projection': summary_metrics.get('performance_projection', 0),
                    'batch_size': summary_metrics.get('batch_size', 1),
                    'source': 'Polaris Simulation (SUMMARY)'
                }

                logger.debug(f"Selected best SUMMARY metrics for {model_key}: "
                           f"throughput={best_throughput:.1f} (from {len(result_list)} variants)")

            logger.info(f"Loaded {len(metrics)} Polaris simulation metrics from SUMMARY JSON files "
                       f"(selected best results from {len(results)} total SUMMARY entries)")
            return metrics

        except Exception as e:
            logger.error(f"Failed to load Polaris metrics: {e}")
            return None


    def create_hw_vs_polaris_comparison(self, tt_metal_metrics: Dict[str, Any],
                                       polaris_metrics: Dict[str, Any]) -> None:
        """Create comparison data between TT-Metal HW and Polaris simulation results

        Args:
            tt_metal_metrics: Metrics from TT-Metal HW runs
            polaris_metrics: Metrics from Polaris simulations
        """
        logger.info("Creating HW vs Polaris comparison data...")

        # Create normalized name mappings - handle multiple TT-Metal entries per normalized key
        tt_metal_normalized = {}
        for k, v in tt_metal_metrics.items():
            normalized = normalize_model_name(k)
            if normalized not in tt_metal_normalized:
                tt_metal_normalized[normalized] = []
            tt_metal_normalized[normalized].append(v)

        # Select the best TT-Metal entry for each normalized key
        for normalized_key, entries in tt_metal_normalized.items():
            if len(entries) > 1:
                # Multiple entries - select the best one
                tt_metal_normalized[normalized_key] = self._select_best_tt_metal_entry(normalized_key, entries)
            else:
                tt_metal_normalized[normalized_key] = entries[0]

        polaris_normalized = {normalize_model_name(k): v for k, v in polaris_metrics.items()}

        # Find models that exist in both datasets using normalized names
        common_normalized = set(tt_metal_normalized.keys()) & set(polaris_normalized.keys())

        logger.info(f"Found {len(common_normalized)} matching models after normalization and selection")

        self.hw_vs_polaris_comparisons = []
        for normalized_key in common_normalized:
            tt_metal_data = tt_metal_normalized[normalized_key]
            polaris_data = polaris_normalized[normalized_key]

            comparison = {
                'model_name': tt_metal_data['model_name'],
                'model_variant': tt_metal_data['model_variant'],
                'hardware': tt_metal_data['hardware'],
                'normalized_key': normalized_key,
                'tt_metal_metric': tt_metal_data,
                'polaris_metric': polaris_data,
                'metric_differences': self._calculate_hw_vs_polaris_differences(tt_metal_data, polaris_data)
            }

            self.hw_vs_polaris_comparisons.append(comparison)

        logger.info(f"Created {len(self.hw_vs_polaris_comparisons)} HW vs Polaris comparisons")

    def _select_best_tt_metal_entry(self, normalized_key: str, entries: List[Dict]) -> Dict:
        """Select the best TT-Metal entry when multiple entries exist for the same normalized key

        Args:
            normalized_key: The normalized model key
            entries: List of TT-Metal entries for this key

        Returns:
            The best entry to use
        """
        if len(entries) == 1:
            return entries[0]

        # For YOLO models, prefer standard variants over specialized ones
        if 'yolo' in normalized_key.lower():
            # Prefer YOLOv8s over YOLOv8s_world
            for entry in entries:
                model_name = entry.get('model_name', '').lower()
                if 'yolov8s' in model_name and 'world' not in model_name:
                    logger.debug(f"Selected YOLOv8s over YOLOv8s_world for {normalized_key}")
                    return entry

        # For other models, just pick the first one (could be enhanced to pick by performance)
        logger.debug(f"Multiple entries for {normalized_key}, selecting first: {entries[0]['model_name']}")
        return entries[0]

    def _calculate_hw_vs_polaris_differences(self, tt_metal: Dict, polaris: Dict) -> Dict[str, Any]:
        """Calculate differences between TT-Metal HW and Polaris simulation metrics"""
        differences = {}

        # Compare execution time
        if 'execution_time' in polaris and polaris['execution_time'] > 0:
            differences['execution_time'] = {
                'tt_metal': None,  # TT-Metal doesn't have execution time
                'polaris': polaris['execution_time'],
                'difference': None,
                'ratio': None
            }

        # Compare throughput metrics where applicable
        tt_metal_throughput = tt_metal.get('value', 0) if tt_metal.get('metric_type') == 'images/sec' else 0
        polaris_throughput = polaris.get('ideal_throughput', 0)

        if tt_metal_throughput > 0 and polaris_throughput > 0:
            differences['throughput'] = {
                'tt_metal': tt_metal_throughput,
                'polaris': polaris_throughput,
                'difference': polaris_throughput - tt_metal_throughput,
                'ratio': polaris_throughput / tt_metal_throughput if tt_metal_throughput > 0 else None,
                'percent_diff': ((polaris_throughput - tt_metal_throughput) / tt_metal_throughput * 100) if tt_metal_throughput > 0 else None
            }

        return differences

    def generate_hw_vs_polaris_report(self, output_path: Path) -> bool:
        """Generate the unified HW vs Polaris comparison report

        Args:
            output_path: Path to save the Excel report

        Returns:
            bool: True if successful
        """
        try:
            logger.info(f"Generating HW vs Polaris comparison report: {output_path}")

            wb = create_excel_workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create overview sheet
            self._create_hw_vs_polaris_overview_sheet(wb)

            # Create detailed comparison sheet
            self._create_hw_vs_polaris_comparison_sheet(wb)

            # Create summary sheet
            self._create_hw_vs_polaris_summary_sheet(wb)

            save_excel_workbook(wb, output_path)
            logger.info(f"HW vs Polaris comparison report saved: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to generate HW vs Polaris report: {e}")
            return False

    def _create_hw_vs_polaris_overview_sheet(self, wb: Workbook) -> None:
        """Create overview sheet explaining the comparison"""
        ws = wb.create_sheet("Overview")

        ws.cell(row=1, column=1, value="HW vs Polaris Performance Comparison")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        ws.cell(row=3, column=1, value="This report compares:").font = Font(bold=True)
        ws.cell(row=4, column=1, value="1. TT-Metal HW Performance: Actual measurements from Tenstorrent hardware")
        ws.cell(row=5, column=1, value="2. Polaris Simulation: Performance estimates from simulation")

        ws.cell(row=7, column=1, value="Key Differences:").font = Font(bold=True)
        ws.cell(row=8, column=1, value="• HW metrics come from real silicon execution")
        ws.cell(row=9, column=1, value="• Polaris metrics are simulation-based estimates")
        ws.cell(row=10, column=1, value="• Comparison shows simulation accuracy vs real HW")

        self._adjust_column_widths(ws)

    def _create_hw_vs_polaris_comparison_sheet(self, wb: Workbook) -> None:
        """Create the detailed HW vs Polaris comparison sheet with populated fields"""
        ws = wb.create_sheet("HW vs Polaris Comparison")

        # Headers - one row per model with both HW and Polaris data
        headers = [
            "Model", "Variant", "Architecture", "HW Metric Type", "HW Value", "HW Unit",
            "Polaris Metric Type", "Polaris Value", "Polaris Unit"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Color code HW vs Polaris columns
            if "HW" in header and header != "HW Value":
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            elif "Polaris" in header and header != "Polaris Value":
                cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")

        row_idx = 2
        for comparison in self.hw_vs_polaris_comparisons:
            tt_metal = comparison['tt_metal_metric']
            polaris = comparison['polaris_metric']

            # Create one row per model with both HW and Polaris data
            ws.cell(row=row_idx, column=1, value=tt_metal['model_name'])
            ws.cell(row=row_idx, column=2, value=tt_metal.get('model_variant', polaris.get('model_variant', '')))
            ws.cell(row=row_idx, column=3, value=tt_metal['hardware'])

            # HW data
            ws.cell(row=row_idx, column=4, value=tt_metal['metric_type'])
            ws.cell(row=row_idx, column=5, value=tt_metal['value'])
            ws.cell(row=row_idx, column=6, value=tt_metal['unit'])
            ws.cell(row=row_idx, column=5).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            # Polaris data
            ws.cell(row=row_idx, column=7, value="Throughput (estimated)")
            ws.cell(row=row_idx, column=8, value=polaris.get('ideal_throughput', 'N/A'))
            ws.cell(row=row_idx, column=9, value="samples/sec")
            ws.cell(row=row_idx, column=8).fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")

            row_idx += 1

        self._adjust_column_widths(ws)

    def _create_hw_vs_polaris_summary_sheet(self, wb: Workbook) -> None:
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary Statistics")

        ws.cell(row=1, column=1, value="HW vs Polaris Summary Statistics")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Calculate summary stats
        if hasattr(self, 'hw_vs_polaris_comparisons') and self.hw_vs_polaris_comparisons:
            total_models = len(self.hw_vs_polaris_comparisons)
            throughput_comparisons = [c for c in self.hw_vs_polaris_comparisons
                                    if 'throughput' in c.get('metric_differences', {})]

            ws.cell(row=3, column=1, value="Total Models Compared:").font = Font(bold=True)
            ws.cell(row=3, column=2, value=total_models)

            ws.cell(row=4, column=1, value="Models with Throughput Data:").font = Font(bold=True)
            ws.cell(row=4, column=2, value=len(throughput_comparisons))

            if throughput_comparisons:
                ratios = [c['metric_differences']['throughput']['ratio']
                         for c in throughput_comparisons
                         if c['metric_differences']['throughput'].get('ratio')]

                if ratios:
                    avg_ratio = sum(ratios) / len(ratios)
                    ws.cell(row=5, column=1, value="Average Polaris/HW Throughput Ratio:").font = Font(bold=True)
                    ws.cell(row=5, column=2, value=f"{avg_ratio:.2f}x")

        self._adjust_column_widths(ws)

    def _check_results_validity(self, output_dir: Path) -> bool:
        """Check if results in output directory are valid and complete

        Args:
            output_dir: Directory containing results

        Returns:
            bool: True if results appear valid
        """
        try:
            raw_results_dir = output_dir / "raw_polaris_results"

            if not raw_results_dir.exists():
                logger.debug(f"Raw results directory missing: {raw_results_dir}")
                return False

            # Check for at least one study-summary.json file
            summary_files = list(raw_results_dir.rglob("study-summary.json"))
            if not summary_files:
                logger.debug("No study-summary.json files found")
                return False

            # Try to parse at least one file
            try:
                with open(summary_files[0], 'r') as f:
                    json.load(f)
            except Exception as e:
                logger.debug(f"Failed to parse study-summary.json: {e}")
                return False

            logger.debug(f"Results validation passed for {output_dir}")
            return True

        except Exception as e:
            logger.debug(f"Results validation error: {e}")
            return False


def main():
    parser = argparse.ArgumentParser(description="HW vs Polaris Performance Comparison Workflow")
    parser.add_argument("--run-hw-vs-polaris", action="store_true",
                       help="Run complete HW vs Polaris workflow: TT-Metal parser -> Polaris automation -> comparison")
    script_dir = Path(__file__).parent
    parser.add_argument("--tt-metal-parser", type=Path, default=script_dir / "comprehensive_n150_n300_parser.py",
                       help="Path to comprehensive_n150_n300_parser.py script")
    parser.add_argument("--polaris-script", type=Path, default=script_dir / "polaris_workload_automation.py",
                       help="Path to polaris_workload_automation.py script")
    parser.add_argument("--local-workload-config", type=Path, default=None,
                       help="Path to local workload configuration YAML file (default: wh_supported.yaml alongside script)")
    parser.add_argument("--unified-output-dir", type=Path, default=None,
                       help="Unified output directory for all profiling results (default: auto-generated with timestamp)")
    parser.add_argument("--tt-metal-dir", type=Path, default=None,
                       help="Directory for TT-Metal HW metrics (default: unified_output_dir/Reports (Excel sheets))")
    parser.add_argument("--polaris-dir", type=Path, default=None,
                       help="Directory for Polaris simulation results (default: unified_output_dir/polaris_simulation_results)")
    parser.add_argument("--output", type=Path, default=None,
                       help="Output Excel file path (default: unified_output_dir/comparison_reports/auto-generated)")
    parser.add_argument("--force-regenerate", action="store_true",
                       help="Force regeneration of results even if they exist")
    parser.add_argument("--max-retries", type=int, default=3,
                       help="Maximum retry attempts for failed runs (default: 3)")
    parser.add_argument("--debug", action="store_true",
                       help="Enable debug logging")
    parser.add_argument("--dry-run", action="store_true",
                       help="Dry run: show what would be done without executing")

    # Legacy mode arguments (for backward compatibility)
    parser.add_argument("--legacy-mode", action="store_true",
                       help="Use legacy mode: compare existing Polaris result directories")

    args = parser.parse_args()

    # Configure logging level
    if args.debug:
        logger.remove()
        logger.add(lambda msg: print(msg, end=""), level="DEBUG")
    else:
        logger.remove()
        logger.add(lambda msg: print(msg, end=""), level="INFO")

    # Set default local workload config if not provided
    if args.local_workload_config is None:
        script_dir = Path(__file__).parent
        # Look alongside this analyzer script (expected to be under polaris/tools/Polaris_vs_HW)
        default_config = (script_dir / "wh_supported.yaml")
        if default_config.exists():
            args.local_workload_config = default_config
            logger.info(f"Using default local workload config: {default_config}")
        else:
            logger.warning(f"Default local workload config not found: {default_config}")
    elif not args.local_workload_config.is_absolute():
        # Convert relative path to absolute path relative to script directory
        script_dir = Path(__file__).parent
        args.local_workload_config = script_dir / args.local_workload_config
        logger.info(f"Converted local workload config to absolute path: {args.local_workload_config}")

    # Create unified output directory structure
    if not hasattr(args, 'unified_output_dir') or args.unified_output_dir is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.unified_output_dir = Path(f"HW_Polaris_comparison_reports_{timestamp}")

    # Create orderly directory structure
    unified_base = args.unified_output_dir
    structure_paths = create_orderly_output_structure(unified_base)

    # Set up subdirectory structure within unified output using orderly paths
    if args.tt_metal_dir is None:
        args.tt_metal_dir = structure_paths['reports']
    if args.polaris_dir is None:
        args.polaris_dir = structure_paths['polaris_simulation_results']
    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = structure_paths['reports'] / f"hw_vs_polaris_comparison_{timestamp}.xlsx"


    # Determine mode of operation
    if args.run_hw_vs_polaris:
        # Run complete HW vs Polaris workflow with unified structure
        logger.info("Running complete HW vs Polaris workflow with unified output structure")
        logger.info(f"Unified output directory: {unified_base}")
        logger.info(f"TT-Metal parser: {args.tt_metal_parser}")
        logger.info(f"Polaris script: {args.polaris_script}")
        logger.info(f"TT-Metal results dir: {args.tt_metal_dir}")
        logger.info(f"Polaris results dir: {args.polaris_dir}")
        logger.info(f"Output report: {args.output}")

        # Validate scripts exist
        if not args.tt_metal_parser.exists():
            logger.error(f"TT-Metal parser script not found: {args.tt_metal_parser}")
            return 1

        if not args.polaris_script.exists():
            logger.error(f"Polaris script not found: {args.polaris_script}")
            return 1

        # Create analyzer
        analyzer = ResultsComparisonAnalyzer(Path("/tmp/placeholder"), Path("/tmp/placeholder"), structure_paths['logs'])

        # Run the complete HW vs Polaris workflow
        if analyzer.run_hw_vs_polaris_comparison_workflow(
            args.tt_metal_parser,
            args.polaris_script,
            args.tt_metal_dir,
            args.polaris_dir,
            args.output,
            args.force_regenerate,
            args.dry_run,
            args.max_retries,
            unified_base,  # Pass the unified output directory
            args.local_workload_config  # Pass the local workload config
        ):
            logger.info(f"✅ HW vs Polaris workflow completed successfully!")
            logger.info(f"📁 All results organized in: {unified_base}")
            return 0
        else:
            logger.error("❌ HW vs Polaris workflow failed")
            return 1

    else:
        # Legacy mode: use existing result directories for Polaris-only comparison
        logger.info("Running legacy Polaris comparison mode")
        logger.warning("⚠️  Legacy mode only compares Polaris results. Use --run-hw-vs-polaris for complete HW vs Polaris analysis.")

        # For legacy mode, we need the old arguments
        parser.add_argument("--optimized-dir", type=Path, default=Path("workflow_optimized_results"),
                           help="Directory for optimized run results")
        parser.add_argument("--full-dir", type=Path, default=Path("workflow_full_results"),
                           help="Directory for full run results")

        # Re-parse to get the legacy arguments
        args = parser.parse_args()

        # Set up output path for legacy mode
        if args.output is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Create a legacy output directory structure
            legacy_base = Path(f"polaris_legacy_comparison_{timestamp}")
            legacy_base.mkdir(parents=True, exist_ok=True)
            args.output = legacy_base / "comparison_reports" / f"polaris_comparison_{timestamp}.xlsx"

            # Create comparison_reports directory
            (legacy_base / "comparison_reports").mkdir(parents=True, exist_ok=True)

        logger.info(f"Optimized run: {args.optimized_dir}")
        logger.info(f"Full run: {args.full_dir}")
        logger.info(f"Output: {args.output}")

        # Validate input directories exist
        if not args.optimized_dir.exists():
            logger.error(f"Optimized run directory does not exist: {args.optimized_dir}")
            logger.error("Use --run-hw-vs-polaris to generate results automatically, or ensure directories exist")
            return 1

        if not args.full_dir.exists():
            logger.error(f"Full run directory does not exist: {args.full_dir}")
            logger.error("Use --run-hw-vs-polaris to generate results automatically, or ensure directories exist")
            return 1

        # Create analyzer and run comparison (legacy mode - no unified logs)
        analyzer = ResultsComparisonAnalyzer(args.optimized_dir, args.full_dir, None)

        logger.info("Starting results comparison analysis...")

        # Parse both runs
        analyzer.optimized_results = analyzer.parse_run_results(args.optimized_dir, "optimized")
        analyzer.full_results = analyzer.parse_run_results(args.full_dir, "full")

        # Perform comparisons
        analyzer.perform_comparisons()

        # Generate report
        if analyzer.generate_excel_report(args.output):
            logger.info("Comparison analysis completed successfully!")
            return 0
        else:
            logger.error("Failed to generate comparison report")
            return 1


if __name__ == "__main__":
    exit(main())
