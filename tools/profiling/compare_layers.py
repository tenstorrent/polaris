#!/usr/bin/env python
# SPDX-FileCopyrightText: (C) 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0

"""Compare Polaris and Profiler CSV layer sequences."""

import sys
import os
import argparse
import csv
from collections import defaultdict
from pathlib import Path
from typing import Any, DefaultDict, Dict, List, Optional, Tuple, Union
from dataclasses import dataclass

try:
    from op_canonical import normalize_polaris_optype, to_comparison_group  # type: ignore[import-not-found]
    from shape_canonical import (  # type: ignore[import-not-found]
        parse_shape_string as parse_shape,
        normalize_shape,
        compare_tensor_shapes,
        validate_binary_compatibility,
        validate_reshape_compatibility,
        compare_tensor_attributes,
    )
except ImportError:
    from .op_canonical import normalize_polaris_optype, to_comparison_group  # type: ignore
    from .shape_canonical import (  # type: ignore
        parse_shape_string as parse_shape,
        normalize_shape,
        compare_tensor_shapes,
        validate_binary_compatibility,
        validate_reshape_compatibility,
        compare_tensor_attributes,
    )

# Maximum distance to search forward for matching operations
DEFAULT_MAX_SEARCH_DISTANCE = 10

# Import layer extraction functions
try:
    from show_layers_polaris import layers_polaris  # type: ignore[import-not-found]
    from show_layers_profiler import layers_profiler  # type: ignore[import-not-found]
except ImportError:
    # Try relative imports
    from .show_layers_polaris import layers_polaris  # type: ignore
    from .show_layers_profiler import layers_profiler  # type: ignore


# Trace-replay dedup: HW captures running with METAL_TRACE_REPLAY emit one
# `(METAL TRACE REPLAY SESSION ID)` column whose value identifies the replay
# session for each row. The op sequence is repeated once per session, so the
# raw profiler CSV is N×ops_per_session rows long. compare_layers picks one
# representative session (median total duration) and discards the rest.
_REPLAY_SESSION_COL = 'METAL TRACE REPLAY SESSION ID'
_DURATION_NS_COL = 'DEVICE KERNEL DURATION [ns]'
_SEQNO_COL = 'GLOBAL CALL COUNT'


@dataclass
class TraceReplayInfo:
    """Structure detected in a profiler CSV."""
    has_trace_replay: bool
    n_sessions: int          # number of distinct replay session IDs
    ops_per_session: int     # ops in the largest session (expected uniform)
    setup_op_count: int      # rows with no session ID (warmup + trace-capture)
    selected_session: int    # session ID chosen for comparison (0 if N/A)
    selected_duration_ns: float  # total device-kernel duration of chosen session


def _detect_trace_replay(filepath: str) -> Tuple['TraceReplayInfo', Dict[int, List[Tuple[int, float]]]]:
    """Scan the raw profiler CSV for METAL TRACE REPLAY SESSION ID structure.

    Returns ``(info, sessions)`` where ``sessions`` maps session_id to a list
    of ``(row_index, duration_ns)`` tuples.  ``row_index`` is the 0-based
    position of the row in the CSV data (excluding header), which aligns
    directly with the index in the list returned by ``layers_profiler``.

    Using row index rather than GLOBAL CALL COUNT is necessary because the
    hardware trace replay re-uses the same GLOBAL CALL COUNT values for every
    replay session — seqnos are NOT unique across sessions.
    """
    sessions: DefaultDict[int, List[Tuple[int, float]]] = defaultdict(list)
    setup_row_count = 0
    has_replay_col = False

    with open(filepath, 'r') as fh:
        reader = csv.DictReader(fh)
        headers = reader.fieldnames or []
        has_replay_col = _REPLAY_SESSION_COL in headers
        for row_idx, row in enumerate(reader):
            dur_raw = (row.get(_DURATION_NS_COL) or '').strip()
            try:
                dur_ns = float(dur_raw) if dur_raw else 0.0
            except ValueError:
                dur_ns = 0.0
            sid_raw = (row.get(_REPLAY_SESSION_COL, '') or '').strip() if has_replay_col else ''
            if sid_raw:
                try:
                    sid = int(float(sid_raw))
                    sessions[sid].append((row_idx, dur_ns))
                except ValueError:
                    setup_row_count += 1
            else:
                setup_row_count += 1

    if not sessions:
        return TraceReplayInfo(
            has_trace_replay=False,
            n_sessions=0,
            ops_per_session=0,
            setup_op_count=setup_row_count,
            selected_session=0,
            selected_duration_ns=0.0,
        ), {}

    session_totals = {sid: sum(d for _, d in rows) for sid, rows in sessions.items()}
    sorted_sids = sorted(session_totals, key=lambda s: session_totals[s])
    selected_sid = sorted_sids[len(sorted_sids) // 2]
    ops_per_session = max(len(v) for v in sessions.values())

    return TraceReplayInfo(
        has_trace_replay=True,
        n_sessions=len(sessions),
        ops_per_session=ops_per_session,
        setup_op_count=setup_row_count,
        selected_session=selected_sid,
        selected_duration_ns=session_totals[selected_sid],
    ), dict(sessions)


def _maybe_dedup_profiler_layers(
    layers: List[Dict[str, Any]],
    filepath: str,
) -> List[Dict[str, Any]]:
    """Apply trace-replay deduplication to a profiler layer list if needed.

    Detects trace structure from the raw CSV, picks the median-total-duration
    replay session, and returns only those layers.  Setup-only ops (warmup,
    trace-capture) are excluded because they do not contribute to device-only FPS.

    Also performs self-checks:
    - Warns when dedup appears necessary but the session ID column is absent.
    - Warns when session sizes are uneven (unexpected replay structure).
    - Confirms when the file is a clean single-pass run requiring no dedup.
    """
    info, sessions = _detect_trace_replay(filepath)

    if not info.has_trace_replay:
        # Self-check: duplicate GLOBAL CALL COUNT (seqno) values are a reliable
        # indicator of trace replay without the session-id column — single-pass
        # profiler runs use unique seqnos; trace replay re-uses the same seqnos
        # across sessions. Repeated op types alone are NOT reliable (single-pass
        # runs naturally have many repeated optypes like conv2d).
        seqno_counts: DefaultDict[int, int] = defaultdict(int)
        for layer in layers:
            seqno_counts[layer['seqno']] += 1
        max_count = max(seqno_counts.values()) if seqno_counts else 0
        if max_count > 1:
            print(
                f"WARNING: '{_REPLAY_SESSION_COL}' column not found, but "
                f"'{_SEQNO_COL}' values repeat up to {max_count}x in the profiler "
                f"CSV. If this is a trace-replay run, add the column to enable auto-dedup.",
                file=sys.stderr,
            )
        else:
            print('Trace replay: not detected - using all profiler layers as-is (single-pass run)')
        return layers

    print(
        f'Trace replay detected: {info.n_sessions} session(s) x {info.ops_per_session} ops, '
        f'{info.setup_op_count} setup-only rows excluded'
    )

    session_sizes = {sid: len(rows) for sid, rows in sessions.items()}
    if len(set(session_sizes.values())) > 1:
        print(
            f'WARNING: replay sessions have uneven op counts: {session_sizes}  '
            f'- unexpected structure; using largest session size as reference.',
            file=sys.stderr,
        )

    print(
        f'Selected session {info.selected_session} '
        f'(median total duration {info.selected_duration_ns / 1e6:.3f} ms)'
    )

    selected_indices = {row_idx for row_idx, _ in sessions[info.selected_session]}
    deduped = [layer for i, layer in enumerate(layers) if i in selected_indices]

    expected = info.ops_per_session
    if len(deduped) != expected:
        print(
            f'WARNING: expected {expected} layers after dedup but got {len(deduped)} '
            f'- row-index alignment between raw CSV and layers_profiler may be off.',
            file=sys.stderr,
        )

    print(f'After dedup: {len(deduped)} profiler layers (was {len(layers)} total rows)')
    return deduped


def sanitize_file_path(filepath: str) -> Path:
    """
    Sanitize and validate a user-provided file path.

    Resolves the path to an absolute path and validates that it:
    - Exists
    - Is a regular file (not a directory or special file)
    - Has an allowed file extension (.csv, .txt, .log, .tsv)
    - Is within a safe base directory (POLARIS_BASE_DIR env var, defaults to HOME)

    Security Note: This is a development/profiling tool that reads CSV files.
    Access is restricted to files within the base directory (user's home by default).
    Set POLARIS_BASE_DIR environment variable to use a different base directory.

    Args:
        filepath: User-provided file path string

    Returns:
        Resolved absolute Path object

    Raises:
        ValueError: If the path is invalid, doesn't exist, or violates security constraints
    """
    # Allowed file extensions for CSV/text files
    ALLOWED_EXTENSIONS = {'.csv', '.txt', '.log', '.tsv'}

    # === STEP 1: Pre-validation of user input ===
    # Check for malicious patterns before any path operations
    if not filepath or not isinstance(filepath, str):
        raise ValueError("File path must be a non-empty string")

    # Check for null bytes (path traversal attack vector)
    if '\0' in filepath:
        raise ValueError("File path contains null bytes")

    # Check for excessively long paths (potential DoS)
    # Using hardcoded 4096 (typical POSIX PATH_MAX) rather than os.pathconf
    # for simplicity and portability. This is a pre-validation security check;
    # the actual OS limit will be enforced by Path.resolve() anyway.
    if len(filepath) > 4096:
        raise ValueError("File path exceeds maximum allowed length")

    # Input validation complete - filepath is now safe to process
    validated_input = filepath

    # === STEP 2: Establish safe base directory ===
    # Determine and validate base directory before processing user filepath
    base_dir_env = os.environ.get('POLARIS_BASE_DIR') or os.getenv('HOME')
    if not base_dir_env:
        raise ValueError("Cannot determine safe base directory (HOME not set)")

    try:
        base_dir = Path(base_dir_env).resolve(strict=True)
        if not base_dir.is_dir():
            raise ValueError(f"Base directory is not a directory: {base_dir}")
    except (FileNotFoundError, RuntimeError, OSError) as e:
        raise ValueError(f"Invalid base directory: {base_dir_env}") from e

    # === STEP 3: Resolve and validate file path ===
    try:
        # Convert validated input to Path and resolve to absolute path
        # This resolves symlinks and .. references
        resolved_path = Path(validated_input).resolve(strict=True)

        # Verify it's a regular file (not a directory or special file)
        if not resolved_path.is_file():
            raise ValueError(f"Path is not a regular file: {validated_input}")

        # Check file extension against allowlist
        if resolved_path.suffix.lower() not in ALLOWED_EXTENSIONS:
            raise ValueError(
                f"File must have one of these extensions: {', '.join(ALLOWED_EXTENSIONS)}. "
                f"Got: {resolved_path.suffix or '(no extension)'}"
            )

        # === STEP 4: Enforce base directory restriction ===
        # Verify resolved path is within the safe base directory
        try:
            resolved_path.relative_to(base_dir)
        except ValueError:
            raise ValueError(
                f"Access denied: File must be within {base_dir}. "
                f"Attempted to access: {resolved_path}"
            )

        # All security checks passed - return sanitized path
        return resolved_path

    except FileNotFoundError:
        raise ValueError(f"File not found: {validated_input}")
    except RuntimeError as e:
        # Can occur with symlink loops
        raise ValueError(f"Invalid file path (possible symlink loop): {validated_input}") from e
    except OSError as e:
        raise ValueError(f"Cannot access file: {validated_input}") from e


@dataclass
class ComparisonStats:
    """Statistics for layer comparison."""
    total_matches: int = 0
    name_mismatches: int = 0
    shape_mismatches: int = 0
    input_shape_mismatches: int = 0
    output_shape_mismatches: int = 0
    attr_mismatches: int = 0
    unmatched_polaris: int = 0
    unmatched_profiler: int = 0
    ambiguous: int = 0
    lut_key_mismatches: int = 0


def parse_args() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Compare Polaris and Profiler CSV layer sequences'
    )
    parser.add_argument('file1', type=str, help='First CSV file (polaris or profiler)')
    parser.add_argument('file2', type=str, nargs='?', default=None,
                        help='Second CSV file (optional; required for shape comparison)')
    parser.add_argument(
        '--perf',
        action='store_true',
        help='Enable performance matching: show network-total and layer-type-wise '
             'duration comparison (ms). With two files, shows gap w.r.t. profiler. '
             'With one file, shows standalone breakdown.'
    )
    parser.add_argument(
        '--max-search-distance',
        type=int,
        default=DEFAULT_MAX_SEARCH_DISTANCE,
        help=f'Maximum distance to search forward for matching operations (default: {DEFAULT_MAX_SEARCH_DISTANCE})'
    )
    parser.add_argument(
        '--strip-leading-ones',
        action=argparse.BooleanOptionalAction,
        default=True,
        help='Strip all leading 1s from shapes (default: enabled). '
             'Leading 1s are a batch-dimension convention difference between Polaris and HW. '
             'Use --no-strip-leading-ones for strict matching.'
    )
    # NOTE: --strip-singleton-dims is currently required for fused head ops
    # (CreateQKVHeads, ConcatHeads) because HW uses 4D shapes with a
    # seq_groups=1 singleton dim (e.g. [B, 1, S, H]) while Polaris emits 3D
    # shapes (e.g. [B, S, H]). Additionally, HW implicitly reinterprets the
    # output of ConcatHeads from [B, 1, S, H] to [1, B, S, H] for
    # downstream ops without an explicit reshape. Future work: update the
    # Polaris shim to emit 4D shapes and model this implicit view change,
    # which would allow removing this flag for those ops.
    parser.add_argument(
        '--strip-singleton-dims',
        action='store_true',
        help='Strip all singleton (=1) dimensions from shapes regardless of position '
             '(handles HW seq_groups=1 convention)'
    )
    parser.add_argument(
        '--filter-optype',
        type=str,
        default=None,
        help='Filter to only compare layers with this operation type (case-insensitive)'
    )
    parser.add_argument(
        '--ignore-attrs',
        action='store_true',
        help='Skip tensor attribute comparison (dtype, layout, memory); compare shapes only'
    )
    parser.add_argument(
        '--summarize-by-signature',
        action='store_true',
        help='Print a rollup table keyed by layer optype (CSV name) plus normalized '
             'input/output shape signature. Uses the same --strip-leading-ones / '
             '--strip-singleton-dims rules as shape comparison. With two CSVs, prints '
             'one table per file. With --perf, adds summed duration (and Polaris LUT '
             'hits when available), and the profiler-vs-Polaris performance comparison '
             'is grouped by type+signature instead of by optype alone.',
    )
    parser.add_argument(
        '--by-lut-key',
        action='store_true',
        help='Print a rollup table grouped by full LUT key (optype + per-input-slot '
             'padded shape, layout, dtype, memory + math_fidelity), with columns '
             'exploded.  Works for both profiler CSV (literal LUT entry key per row) '
             'and polaris CSV (uses the resolved key — the LUT entry the runtime '
             'lookup chain matched after any HEIGHT→BLOCK / L1→DRAM / arity-dup '
             'fallback — so polaris and profiler rows that share a LUT entry land '
             'in the same bucket).  With one file, prints a single-file count table. '
             'With two files, prints a side-by-side count table.  With --perf, adds '
             'summed duration columns and absolute/percent gap.',
    )
    parser.add_argument(
        '--xlsx',
        type=str,
        default=None,
        metavar='PATH',
        help='Also write an .xlsx report with four sheets: '
             '"Summary" (network-wide totals + shape/attr counts), '
             '"By Layer Type" (per canonical optype), '
             '"By Layer Signature" (per optype + normalized in/out shape signature), and '
             '"By LUT Key" (per full LUT key tuple with columns exploded). '
             'In two-file mode each sheet includes file1 vs file2 columns and gap; '
             'in single-file mode only that source\'s counts/ms are emitted. '
             'Requires openpyxl (already a polarisdev dep).',
    )
    parser.add_argument(
        '--label1',
        type=str,
        default=None,
        metavar='LABEL',
        help='Display label for the first file (default: auto-detected type, e.g. "Polaris" '
             'or "Profiler"; for same-type comparisons defaults to "Polaris 1" / "Profiler 1")',
    )
    parser.add_argument(
        '--label2',
        type=str,
        default=None,
        metavar='LABEL',
        help='Display label for the second file (default: auto-detected type, e.g. "Polaris" '
             'or "Profiler"; for same-type comparisons defaults to "Polaris 2" / "Profiler 2")',
    )
    return parser.parse_args()


def normalize_optype(optype: str) -> str:
    """Normalize operation type name to canonical form then coarsen for
    sequence matching (e.g. add/mul/sub → binary).

    Delegates to :mod:`op_canonical` for the canonical form and comparison
    group mapping.
    """
    return to_comparison_group(normalize_polaris_optype(optype))


def detect_file_type(filepath: str) -> Optional[str]:
    """
    Detect if file is polaris or profiler CSV.

    Args:
        filepath: User-provided file path (will be sanitized)

    Returns:
        'polaris' if archname column found
        'profiler' if OP CODE column found
        None if neither found
    """
    try:
        # Sanitize the file path to prevent path traversal attacks
        safe_path = sanitize_file_path(filepath)

        with open(safe_path, 'r') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            if 'archname' in headers:
                return 'polaris'
            elif 'OP CODE' in headers:
                return 'profiler'
            return None
    except ValueError as e:
        # Path sanitization failed
        print(f"Invalid file path {filepath}: {e}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"Error reading {filepath}: {e}", file=sys.stderr)
        return None




def find_next_match(
    layers: List[Dict[str, Any]],
    start_idx: int,
    target_optype: str,
    max_distance: Optional[int] = None
) -> Optional[int]:
    """
    Find the next occurrence of target_optype in layers starting from start_idx.

    Args:
        layers: List of layer dictionaries
        start_idx: Index to start searching from
        target_optype: Normalized operation type to search for
        max_distance: Maximum number of operations to search ahead (None = unlimited)

    Returns:
        Index of match, or None if not found within max_distance
    """
    end_idx = len(layers)
    if max_distance is not None:
        end_idx = min(end_idx, start_idx + max_distance)

    for i in range(start_idx, end_idx):
        if normalize_optype(layers[i]['optype']) == target_optype:
            return i
    return None


def format_shapes(shapes: List[str]) -> str:
    """Format list of shape strings for display."""
    if not shapes:
        return "[]"
    return "[" + ", ".join(shapes) + "]"


def compare_layers(
    layers1: List[Dict[str, Any]],
    layers2: List[Dict[str, Any]],
    max_search_distance: int = DEFAULT_MAX_SEARCH_DISTANCE,
    strip_leading_ones: bool = False,
    strip_singleton_dims: bool = False,
    ignore_attrs: bool = False,
    label1: str = 'File1',
    label2: str = 'File2',
) -> ComparisonStats:
    """
    Compare two layer sequences and print results.

    layers1 is the pivot (iterated in order); layers2 is searched for matches.
    label1/label2 are used in diagnostic output.

    Returns:
        ComparisonStats object with statistics
    """
    stats = ComparisonStats()
    ndx1 = 0
    ndx2 = 0

    while ndx1 < len(layers1) or ndx2 < len(layers2):
        # Case 3: One sequence exhausted
        if ndx1 >= len(layers1):
            layer = layers2[ndx2]
            print(f"⊘ [2:{layer['seqno']}] {layer['optype']} (not in {label1})")
            stats.unmatched_profiler += 1
            ndx2 += 1
            continue

        if ndx2 >= len(layers2):
            layer = layers1[ndx1]
            print(f"⊘ [1:{layer['seqno']}] {layer['optype']} (not in {label2})")
            stats.unmatched_polaris += 1
            ndx1 += 1
            continue

        # Get current layers
        l1 = layers1[ndx1]
        l2 = layers2[ndx2]

        # Normalize optypes for comparison
        l1_optype_norm = normalize_optype(l1['optype'])
        l2_optype_norm = normalize_optype(l2['optype'])

        # Case 1: optypes match (after normalization)
        if l1_optype_norm == l2_optype_norm:
            # Compare shapes
            input_match, input_details = compare_tensor_shapes(
                l1.get('input_tensors', []),
                l2.get('input_tensors', []),
                strip_leading_ones,
                l1['optype'],
                strip_singleton_dims=strip_singleton_dims,
            )
            output_match, output_details = compare_tensor_shapes(
                l1.get('output_tensors', []),
                l2.get('output_tensors', []),
                strip_leading_ones,
                l1['optype'],
                strip_singleton_dims=strip_singleton_dims,
            )

            # Special handling for binary ops (add/mul/sub) if input counts
            # or shapes don't match — one side may use scalar/untracked operands
            l1_canonical = normalize_polaris_optype(l1['optype'])
            if not input_match and to_comparison_group(l1_canonical) == 'binary':
                bin_valid, bin_details = validate_binary_compatibility(
                    l1.get('input_tensors', []),
                    l2.get('input_tensors', []),
                    strip_leading_ones,
                    strip_singleton_dims=strip_singleton_dims,
                )
                if bin_valid:
                    input_match = True
                    input_details = bin_details

            # Special handling for reshape if standard comparison fails
            if l1_canonical == 'reshape' and (not input_match or not output_match):
                reshape_valid, reshape_details = validate_reshape_compatibility(
                    l1.get('input_tensors', []),
                    l1.get('output_tensors', []),
                    l2.get('output_tensors', []),
                    strip_leading_ones,
                    strip_singleton_dims=strip_singleton_dims,
                )
                if reshape_valid:
                    # Accept reshape as valid - inputs may differ but outputs are compatible
                    input_match = True
                    output_match = True
                    output_details = reshape_details

            # Attribute comparison (dtype, layout, memory)
            attr_ok = True
            attr_details_parts = []
            if input_match and output_match and not ignore_attrs:
                in_attr_ok, in_attr_det = compare_tensor_attributes(l1, l2, 'input')
                out_attr_ok, out_attr_det = compare_tensor_attributes(l1, l2, 'output')
                if not in_attr_ok:
                    attr_ok = False
                    attr_details_parts.append(f"input attrs: {in_attr_det}")
                if not out_attr_ok:
                    attr_ok = False
                    attr_details_parts.append(f"output attrs: {out_attr_det}")

            if input_match and output_match and attr_ok:
                print(f"✓ [1:{l1['seqno']}] [2:{l2['seqno']}] {l1['optype']}  "
                      f"in: {format_shapes(l1.get('input_tensors', []))} | "
                      f"out: {format_shapes(l1.get('output_tensors', []))}")
                stats.total_matches += 1
            elif input_match and output_match and not attr_ok:
                print(f"✗ attr [1:{l1['seqno']}] [2:{l2['seqno']}] {l1['optype']}")
                for part in attr_details_parts:
                    print(f"  {part}")
                stats.attr_mismatches += 1
            else:
                print(f"✗ shape [1:{l1['seqno']}] [2:{l2['seqno']}] {l1['optype']}")
                if not input_match:
                    print(f"  input: {label1}={format_shapes(l1.get('input_tensors', []))} "
                          f"{label2}={format_shapes(l2.get('input_tensors', []))} ({input_details})")
                    stats.input_shape_mismatches += 1
                if not output_match:
                    print(f"  output: {label1}={format_shapes(l1.get('output_tensors', []))} "
                          f"{label2}={format_shapes(l2.get('output_tensors', []))} ({output_details})")
                    stats.output_shape_mismatches += 1
                stats.shape_mismatches += 1

            # LUT key comparison (for all matched pairs, both profiler outputs)
            lk1 = l1.get('lut_key_resolved') or l1.get('lut_key')
            lk2 = l2.get('lut_key_resolved') or l2.get('lut_key')
            if lk1 is not None and lk2 is not None and lk1 != lk2:
                print(f"  lut_key mismatch [1:{l1['seqno']}] [2:{l2['seqno']}] {l1['optype']}")
                print(f"    {label1}: {lk1}")
                print(f"    {label2}: {lk2}")
                stats.lut_key_mismatches += 1

            ndx1 += 1
            ndx2 += 1
            continue

        # Case 2: optypes don't match — search forward in layers2 (layers1 is pivot)
        match_idx2 = find_next_match(
            layers2, ndx2 + 1, l1_optype_norm, max_search_distance
        )

        if match_idx2 is not None:
            for i in range(ndx2, match_idx2):
                layer = layers2[i]
                print(f"⊘ [2:{layer['seqno']}] {layer['optype']} (not in {label1})")
                stats.unmatched_profiler += 1
            ndx2 = match_idx2
        else:
            print(f"✗ name [1:{l1['seqno']}] --- {l1['optype']} (not in {label2})")
            stats.name_mismatches += 1
            ndx1 += 1

    return stats


def _signature_string_for_layer(
    layer: Dict[str, Any],
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> str:
    """Normalized in/out shape string for grouping (matches compare_layers strip rules)."""

    def fmt_slot(shapes: List[str]) -> str:
        parts: List[str] = []
        for sh in shapes or []:
            dims = parse_shape(sh)
            nd = normalize_shape(dims, strip_leading_ones, strip_singleton_dims)
            parts.append("x".join(str(x) for x in nd) if nd else "")
        return ";".join(parts)

    ins = fmt_slot(layer.get("input_tensors") or [])
    outs = fmt_slot(layer.get("output_tensors") or [])
    return f"in[{ins}] out[{outs}]"


def _print_signature_summary(
    layers: List[Dict[str, Any]],
    label: str,
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
    *,
    include_perf: bool,
) -> None:
    """Rollup: count (and optionally ms / LUT) per (optype, shape signature)."""
    counts: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    ms_totals: DefaultDict[Tuple[str, str], float] = defaultdict(float)
    lut_totals: DefaultDict[Tuple[str, str], int] = defaultdict(int)

    for layer in layers:
        optype = str(layer.get("optype", ""))
        sig = _signature_string_for_layer(layer, strip_leading_ones, strip_singleton_dims)
        key = (optype, sig)
        counts[key] += 1
        if include_perf:
            d = layer.get("duration_ms")
            if d is not None:
                ms_totals[key] += float(d)
            if layer.get("uses_perf_lookup"):
                lut_totals[key] += 1

    keys_sorted = sorted(counts.keys(), key=lambda k: (-counts[k], k[0], k[1]))
    any_lut = include_perf and sum(lut_totals.values()) > 0

    print(f"\n{'=' * 72}")
    print(f"  Summary by layer type + signature ({label})")
    print(f"{'=' * 72}")

    hdr_count = "Count"
    hdr_type = "Layer type"
    hdr_sig = "Signature (normalized in / out)"
    col_c = max(5, len(hdr_count))
    col_t = max(12, max((len(k[0]) for k in keys_sorted), default=len(hdr_type)))
    col_m = 11
    col_l = 9

    if include_perf:
        hdr_ms = "Sum ms"
        hdr_lut = "LUT"
        print(
            f"  {hdr_count:>{col_c}}  {hdr_type:<{col_t}}  {hdr_ms:>{col_m}}  "
            f"{hdr_lut:>{col_l}}  {hdr_sig}"
        )
    else:
        print(f"  {hdr_count:>{col_c}}  {hdr_type:<{col_t}}  {hdr_sig}")
    print(f"  {'─' * 72}")

    total_n = 0
    total_ms = 0.0
    total_lut = 0
    for key in keys_sorted:
        n = counts[key]
        total_n += n
        op, sig = key
        if include_perf:
            ms = ms_totals.get(key, 0.0)
            lut = lut_totals.get(key, 0)
            total_ms += ms
            total_lut += lut
            lut_s = f"{lut}/{n}" if any_lut else "—"
            print(
                f"  {n:>{col_c}}  {op:<{col_t}}  {ms:>{col_m}.4f}  {lut_s:>{col_l}}  {sig}"
            )
        else:
            print(f"  {n:>{col_c}}  {op:<{col_t}}  {sig}")

    print(f"  {'─' * 72}")
    if include_perf:
        lut_footer = f"{total_lut}/{total_n}" if any_lut else "—"
        print(
            f"  {total_n:>{col_c}}  {'TOTAL':<{col_t}}  {total_ms:>{col_m}.4f}  {lut_footer:>{col_l}}"
        )
    else:
        print(f"  {total_n:>{col_c}}  {'TOTAL':<{col_t}}")
    print()


def print_summary(stats: ComparisonStats, label1: str = 'File1', label2: str = 'File2') -> None:
    """Print summary statistics."""
    print("\n=== Summary ===")
    print(f"Total matches: {stats.total_matches}")
    print(f"Name mismatches: {stats.name_mismatches}")
    print(f"Shape mismatches: {stats.shape_mismatches} "
          f"({stats.input_shape_mismatches} input, {stats.output_shape_mismatches} output)")
    print(f"Attribute mismatches: {stats.attr_mismatches}")
    print(f"LUT key mismatches: {stats.lut_key_mismatches}")
    print(f"Unmatched entries: {stats.unmatched_polaris + stats.unmatched_profiler} "
          f"({stats.unmatched_polaris} {label1}, {stats.unmatched_profiler} {label2})")
    print(f"Ambiguous: {stats.ambiguous}")


# ---------------------------------------------------------------------------
# Performance summary helpers
# ---------------------------------------------------------------------------

def _aggregate_duration_by_optype_signature(
    layers: List[Dict[str, Any]],
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> Dict[Tuple[str, str], Tuple[int, float, int]]:
    """Group layers by (optype, normalized shape signature); sum durations and LUT hits."""
    totals: DefaultDict[Tuple[str, str], float] = defaultdict(float)
    counts: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    lut_hits: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    for layer in layers:
        optype = str(layer.get("optype", ""))
        sig = _signature_string_for_layer(layer, strip_leading_ones, strip_singleton_dims)
        key = (optype, sig)
        counts[key] += 1
        dur = layer.get("duration_ms")
        if dur is not None:
            totals[key] += float(dur)
        if layer.get("uses_perf_lookup"):
            lut_hits[key] += 1
    all_keys = set(counts) | set(totals) | set(lut_hits)
    return {
        k: (counts[k], totals.get(k, 0.0), lut_hits.get(k, 0))
        for k in all_keys
    }


def _aggregate_duration_by_optype(
    layers: List[Dict[str, Any]],
) -> Dict[str, Tuple[int, float, int]]:
    """Group layers by fine-grained canonical optype and sum durations.

    Returns ``{optype: (count, total_ms, lut_hits)}`` sorted by descending
    total_ms.  Layers whose ``duration_ms`` is None are counted but
    contribute 0 ms.  ``lut_hits`` counts layers where
    ``uses_perf_lookup`` is True (Polaris-only; always 0 for profiler).
    """
    totals: Dict[str, float] = defaultdict(float)
    counts: Dict[str, int] = defaultdict(int)
    lut_hits: Dict[str, int] = defaultdict(int)
    for layer in layers:
        optype = layer['optype']
        counts[optype] += 1
        dur = layer.get('duration_ms')
        if dur is not None:
            totals[optype] += dur
        if layer.get('uses_perf_lookup'):
            lut_hits[optype] += 1
    all_optypes = set(counts) | set(totals)
    return {
        op: (counts[op], totals.get(op, 0.0), lut_hits.get(op, 0))
        for op in sorted(all_optypes, key=lambda o: totals.get(o, 0.0), reverse=True)
    }


def _pct_gap(reference: float, other: float) -> str:
    """Format percentage gap of *other* w.r.t. *reference*.

    Positive means *other* is larger (slower).
    """
    if reference == 0.0:
        return "N/A"
    gap = (other - reference) / reference * 100.0
    sign = "+" if gap >= 0 else ""
    return f"{sign}{gap:.2f}%"


def _print_perf_standalone_by_signature(
    layers: List[Dict[str, Any]],
    source_label: str,
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> None:
    """Standalone performance breakdown grouped by optype + signature."""
    by_key = _aggregate_duration_by_optype_signature(
        layers, strip_leading_ones, strip_singleton_dims
    )
    total_ms = sum(ms for _, ms, _ in by_key.values())
    total_count = sum(cnt for cnt, _, _ in by_key.values())
    total_lut = sum(lut for _, _, lut in by_key.values())
    has_lut = total_lut > 0

    print(f"\n{'=' * 60}")
    print(f"  Performance Summary by type + signature ({source_label})")
    print(f"{'=' * 60}")
    print(f"\n  Network total: {total_ms:.4f} ms")
    if has_lut:
        print(f"  LUT hits: {total_lut}/{total_count}")
    print()

    keys_sorted = sorted(
        by_key.keys(),
        key=lambda k: by_key.get(k, (0, 0.0, 0))[1],
        reverse=True,
    )

    hdr_type = "Layer type"
    hdr_sig = "Signature"
    col_w_t = max(10, max((len(k[0]) for k in keys_sorted), default=len(hdr_type)))
    col_w_s = max(24, min(72, max((len(k[1]) for k in keys_sorted), default=len(hdr_sig))))

    hdr_cnt = "Count"
    hdr_dur = "Duration (ms)"
    hdr_lut = "LUT"
    col_w_cnt = max(len(hdr_cnt), 6)
    col_w_dur = max(len(hdr_dur), 14)
    col_w_lut = max(len(hdr_lut), 8)

    header = (
        f"  {hdr_type:<{col_w_t}}  {hdr_sig:<{col_w_s}}  {hdr_cnt:>{col_w_cnt}}  "
        f"{hdr_dur:>{col_w_dur}}"
    )
    if has_lut:
        header += f"  {hdr_lut:>{col_w_lut}}"
    print(header)
    rule_len = col_w_t + col_w_s + col_w_cnt + col_w_dur + 8 + (col_w_lut + 2 if has_lut else 0)
    print(f"  {'─' * min(rule_len, 120)}")

    for key in keys_sorted:
        op, sig = key
        cnt, ms, lut = by_key[key]
        sig_disp = sig if len(sig) <= col_w_s else sig[: col_w_s - 3] + "..."
        line = (
            f"  {op:<{col_w_t}}  {sig_disp:<{col_w_s}}  {cnt:>{col_w_cnt}}  "
            f"{ms:>{col_w_dur}.4f}"
        )
        if has_lut:
            line += f"  {f'{lut}/{cnt}':>{col_w_lut}}"
        print(line)

    print(f"  {'─' * min(rule_len, 120)}")
    line = (
        f"  {'TOTAL':<{col_w_t}}  {'':<{col_w_s}}  {total_count:>{col_w_cnt}}  "
        f"{total_ms:>{col_w_dur}.4f}"
    )
    if has_lut:
        line += f"  {f'{total_lut}/{total_count}':>{col_w_lut}}"
    print(line)
    print()


def _print_perf_standalone(
    layers: List[Dict[str, Any]],
    source_label: str,
) -> None:
    """Print a standalone performance breakdown for a single source."""
    by_optype = _aggregate_duration_by_optype(layers)
    total_ms = sum(ms for _, ms, _ in by_optype.values())
    total_count = sum(cnt for cnt, _, _ in by_optype.values())
    total_lut = sum(lut for _, _, lut in by_optype.values())
    has_lut = total_lut > 0

    print(f"\n{'=' * 60}")
    print(f"  Performance Summary ({source_label})")
    print(f"{'=' * 60}")
    print(f"\n  Network total: {total_ms:.4f} ms")
    if has_lut:
        print(f"  LUT hits: {total_lut}/{total_count}")
    print()

    hdr_type = "Layer Type"
    hdr_cnt = "Count"
    hdr_dur = "Duration (ms)"
    hdr_lut = "LUT"
    col_w_type = max(len(hdr_type), max((len(op) for op in by_optype), default=10))
    col_w_cnt = max(len(hdr_cnt), 6)
    col_w_dur = max(len(hdr_dur), 14)
    col_w_lut = max(len(hdr_lut), 8)

    header = f"  {hdr_type:<{col_w_type}}  {hdr_cnt:>{col_w_cnt}}  {hdr_dur:>{col_w_dur}}"
    if has_lut:
        header += f"  {hdr_lut:>{col_w_lut}}"
    print(header)
    rule_len = col_w_type + col_w_cnt + col_w_dur + 4 + (col_w_lut + 2 if has_lut else 0)
    print(f"  {'─' * rule_len}")

    for op, (cnt, ms, lut) in by_optype.items():
        line = f"  {op:<{col_w_type}}  {cnt:>{col_w_cnt}}  {ms:>{col_w_dur}.4f}"
        if has_lut:
            line += f"  {f'{lut}/{cnt}':>{col_w_lut}}"
        print(line)

    print(f"  {'─' * rule_len}")
    line = f"  {'TOTAL':<{col_w_type}}  {total_count:>{col_w_cnt}}  {total_ms:>{col_w_dur}.4f}"
    if has_lut:
        line += f"  {f'{total_lut}/{total_count}':>{col_w_lut}}"
    print(line)
    print()


def _print_perf_comparison(
    layers1: List[Dict[str, Any]],
    layers2: List[Dict[str, Any]],
    *,
    by_signature: bool = False,
    strip_leading_ones: bool = False,
    strip_singleton_dims: bool = False,
    label1: str = 'File1',
    label2: str = 'File2',
) -> None:
    """Print side-by-side performance comparison; gap is (file2 − file1) / file1."""
    by1: Union[Dict[Tuple[str, str], Tuple[int, float, int]], Dict[str, Tuple[int, float, int]]]
    by2: Union[Dict[Tuple[str, str], Tuple[int, float, int]], Dict[str, Tuple[int, float, int]]]

    if by_signature:
        by1 = _aggregate_duration_by_optype_signature(layers1, strip_leading_ones, strip_singleton_dims)
        by2 = _aggregate_duration_by_optype_signature(layers2, strip_leading_ones, strip_singleton_dims)
        title = "Performance Summary (by layer type + signature)"
    else:
        by1 = _aggregate_duration_by_optype(layers1)
        by2 = _aggregate_duration_by_optype(layers2)
        title = "Performance Summary"

    total_ms1 = sum(ms for _, ms, _ in by1.values())
    total_ms2 = sum(ms for _, ms, _ in by2.values())
    total_cnt1 = sum(cnt for cnt, _, _ in by1.values())
    total_cnt2 = sum(cnt for cnt, _, _ in by2.values())
    total_lut1 = sum(lut for _, _, lut in by1.values())
    total_lut2 = sum(lut for _, _, lut in by2.values())

    print(f"\n{'=' * 82}")
    print(f"  {title}")
    print(f"{'=' * 82}")

    print("\n  Network total:")
    print(f"    {label1}:  {total_ms1:.4f} ms")
    print(f"    {label2}:  {total_ms2:.4f} ms")
    print(f"    Gap:  {_pct_gap(total_ms1, total_ms2)} (w.r.t. {label1})")
    if total_lut1 > 0:
        print(f"    {label1} LUT hits: {total_lut1}/{total_cnt1}")
    if total_lut2 > 0:
        print(f"    {label2} LUT hits: {total_lut2}/{total_cnt2}")
    print()

    all_keys: List[Any] = list(dict.fromkeys(list(by1.keys()) + list(by2.keys())))
    all_keys.sort(key=lambda k: by1.get(k, (0, 0.0, 0))[1], reverse=True)

    lbl1 = label1[:10]
    lbl2 = label2[:10]
    hdr_ms1 = f"{lbl1}(ms)"
    hdr_ms2 = f"{lbl2}(ms)"
    col_ms = max(13, len(hdr_ms1), len(hdr_ms2))
    has_lut1 = total_lut1 > 0
    has_lut2 = total_lut2 > 0

    if by_signature:
        col_w_op = max(10, max((len(k[0]) for k in all_keys), default=10))
        col_w_sig = max(28, min(56, max((len(k[1]) for k in all_keys), default=28)))
        hdr = (
            f"  {'Layer type':<{col_w_op}}"
            f"  {'Signature':<{col_w_sig}}"
            f"  {'#1':>6}  {hdr_ms1:>{col_ms}}"
            f"  {'#2':>6}  {hdr_ms2:>{col_ms}}"
        )
    else:
        col_w_op = max(10, max((len(str(op)) for op in all_keys), default=10))
        col_w_sig = 0
        hdr = (
            f"  {'Layer Type':<{col_w_op}}"
            f"  {'#1':>6}  {hdr_ms1:>{col_ms}}"
            f"  {'#2':>6}  {hdr_ms2:>{col_ms}}"
        )
    if has_lut1:
        hdr += f"  {'LUT1':>8}"
    if has_lut2:
        hdr += f"  {'LUT2':>8}"
    hdr += f"  {'Abs Gap(ms)':>12}  {'Gap%':>9}"
    print(hdr)
    rule_len = (col_w_op + 6 + col_ms + 6 + col_ms + 12 + 9 + 16
                + (col_w_sig + 2 if by_signature else 0)
                + (10 if has_lut1 else 0) + (10 if has_lut2 else 0))
    print(f"  {'─' * min(rule_len, 120)}")

    for key in all_keys:
        if by_signature:
            op, sig = key
            sig_disp = sig if len(sig) <= col_w_sig else sig[: col_w_sig - 3] + "..."
            cnt1, ms1, lut1 = by1.get(key, (0, 0.0, 0))
            cnt2, ms2, lut2 = by2.get(key, (0, 0.0, 0))
        else:
            op = key
            sig_disp = ""
            cnt1, ms1, lut1 = by1.get(op, (0, 0.0, 0))
            cnt2, ms2, lut2 = by2.get(op, (0, 0.0, 0))
        gap_pct = _pct_gap(ms1, ms2)
        abs_gap = ms2 - ms1
        abs_gap_s = f"{abs_gap:+.4f}" if (cnt1 and cnt2) else "—"
        cnt1_s = str(cnt1) if cnt1 else "—"
        cnt2_s = str(cnt2) if cnt2 else "—"
        ms1_s = f"{ms1:.4f}" if cnt1 else "—"
        ms2_s = f"{ms2:.4f}" if cnt2 else "—"
        lut1_s = f"{lut1}/{cnt1}" if has_lut1 and cnt1 else ("—" if has_lut1 else "")
        lut2_s = f"{lut2}/{cnt2}" if has_lut2 and cnt2 else ("—" if has_lut2 else "")
        if by_signature:
            line = (
                f"  {op:<{col_w_op}}"
                f"  {sig_disp:<{col_w_sig}}"
                f"  {cnt1_s:>6}  {ms1_s:>{col_ms}}"
                f"  {cnt2_s:>6}  {ms2_s:>{col_ms}}"
            )
        else:
            line = (
                f"  {op:<{col_w_op}}"
                f"  {cnt1_s:>6}  {ms1_s:>{col_ms}}"
                f"  {cnt2_s:>6}  {ms2_s:>{col_ms}}"
            )
        if has_lut1:
            line += f"  {lut1_s:>8}"
        if has_lut2:
            line += f"  {lut2_s:>8}"
        line += f"  {abs_gap_s:>12}  {gap_pct:>9}"
        print(line)

    abs_total = total_ms2 - total_ms1
    print(f"  {'─' * min(rule_len, 120)}")
    if by_signature:
        total_line = (
            f"  {'TOTAL':<{col_w_op}}"
            f"  {'':<{col_w_sig}}"
            f"  {total_cnt1:>6}  {total_ms1:>{col_ms}.4f}"
            f"  {total_cnt2:>6}  {total_ms2:>{col_ms}.4f}"
        )
    else:
        total_line = (
            f"  {'TOTAL':<{col_w_op}}"
            f"  {total_cnt1:>6}  {total_ms1:>{col_ms}.4f}"
            f"  {total_cnt2:>6}  {total_ms2:>{col_ms}.4f}"
        )
    if has_lut1:
        total_line += f"  {f'{total_lut1}/{total_cnt1}':>8}"
    if has_lut2:
        total_line += f"  {f'{total_lut2}/{total_cnt2}':>8}"
    total_line += f"  {abs_total:>+12.4f}  {_pct_gap(total_ms1, total_ms2):>9}"
    print(total_line)
    print()


# ---------------------------------------------------------------------------
# LUT key summary helpers
# ---------------------------------------------------------------------------

_LUT_SLOT_NAMES: Tuple[str, ...] = ('w', 'z', 'y', 'x', 'layout', 'dtype', 'memory')


def _lut_key_n_slots(key: Tuple[str, ...]) -> int:
    """Number of input slots encoded in a lut_key tuple."""
    if key and key[0] == 'halo':
        return 1  # halo keys have exactly one input slot regardless of length (15-tuple v3, 16-tuple v4)
    return max(0, (len(key) - 2) // 7)


def _lut_key_slot_field(key: Tuple[str, ...], slot: int, field_idx: int) -> str:
    """Return the string value for (slot, field_idx) in a lut_key, or '' if out of range."""
    # math_fidelity sits at index 8 between slot-0 and slot-1; skip it when accessing slot 1+.
    idx = 1 + slot * 7 + field_idx + (1 if slot > 0 else 0)
    return str(key[idx]) if idx < len(key) else ''


_HALO_V3_VARIANT_NAMES: Tuple[str, ...] = ('kernel_h', 'kernel_w', 'stride_h', 'stride_w', 'padding_h', 'padding_w')
_HALO_V4_VARIANT_NAMES: Tuple[str, ...] = _HALO_V3_VARIANT_NAMES + ('is_transpose',)
_ITS_VARIANT_NAMES: Tuple[str, ...] = ('output_0_memory',)


def _lut_key_variant_tail(key: Tuple[str, ...]) -> List[Tuple[str, str]]:
    """Return [(field_name, value), ...] for variant-specific fields at indices 9+ of a lut_key."""
    if not key:
        return []
    op = key[0]
    if op == 'halo':
        if len(key) == 16:
            names: Tuple[str, ...] = _HALO_V4_VARIANT_NAMES
        elif len(key) == 15:
            names = _HALO_V3_VARIANT_NAMES
        else:
            return []
        return [(n, str(key[9 + i])) for i, n in enumerate(names)]
    if op == 'interleavedtosharded' and len(key) == 10:
        return [('output_0_memory', str(key[9]))]
    return []


def _lut_key_variant_col_names(keys: List[Tuple[str, ...]]) -> List[str]:
    """Ordered union of all variant tail field names present across these keys."""
    seen: dict = {}
    for k in keys:
        for name, _ in _lut_key_variant_tail(k):
            seen[name] = None
    return list(seen.keys())


def _aggregate_by_lut_key(
    layers: List[Dict[str, Any]],
) -> Dict[Tuple[str, ...], Tuple[int, float, int]]:
    """Group layers by their *resolved* LUT key; return {key: (count, total_ms, lut_hits)}.

    Prefers ``lut_key_resolved`` (the tuple the polaris runtime LUT lookup actually
    matched, after any fallback substitution: HEIGHT→BLOCK shard, L1→DRAM
    interleaved, ROW_MAJOR→TILE layout, arity-1→arity-2 dup) over the literal
    ``lut_key``.  This makes polaris rows and profiler rows that semantically
    share a LUT entry land in the same bucket — without this, polaris's literal
    key (built from the workload op's tensor state) and the profiler's literal key
    (= the actual LUT entry) can differ in fallback-tolerable ways and over-segment.

    Profiler-side rows only carry ``lut_key`` (their literal key IS a LUT entry, no
    fallback needed); they fall through to the literal key automatically.  Older
    polaris CSVs predating the Tier-1 plumbing also lack ``lut_key_resolved``;
    those degrade to using the literal key — accurate for direct-hit rows,
    over-segmenting for fallback-resolved rows (same as before).
    """
    counts: DefaultDict[Tuple[str, ...], int] = defaultdict(int)
    ms_totals: DefaultDict[Tuple[str, ...], float] = defaultdict(float)
    lut_totals: DefaultDict[Tuple[str, ...], int] = defaultdict(int)
    for layer in layers:
        raw = layer.get('lut_key_resolved') or layer.get('lut_key')
        if raw is None:
            continue
        key: Tuple[str, ...] = tuple(str(f) for f in raw)
        counts[key] += 1
        d = layer.get('duration_ms')
        if d is not None:
            ms_totals[key] += float(d)
        if layer.get('uses_perf_lookup'):
            lut_totals[key] += 1
    return {k: (counts[k], ms_totals.get(k, 0.0), lut_totals.get(k, 0)) for k in counts}


def _lut_key_col_widths(
    keys: List[Tuple[str, ...]], max_slots: int, variant_col_names: Optional[List[str]] = None
) -> Tuple[int, int, List[List[int]], List[int]]:
    """Return (col_op, col_mf, slot_widths, variant_widths) computed from data."""
    col_op = max(10, max((len(k[0]) for k in keys), default=10))
    col_mf = max(4, max((len(str(k[8])) if len(k) > 8 else 0 for k in keys), default=4))
    slot_widths: List[List[int]] = []
    for s in range(max_slots):
        widths = []
        for fi, name in enumerate(_LUT_SLOT_NAMES):
            suffix = f'_{s}' if s > 0 else ''
            hdr_w = len(name + suffix)
            data_w = max((_lut_key_slot_field(k, s, fi).__len__() for k in keys), default=0)
            widths.append(max(hdr_w, data_w, 4))
        slot_widths.append(widths)
    variant_widths: List[int] = []
    for vname in (variant_col_names or []):
        data_w = max((len(dict(_lut_key_variant_tail(k)).get(vname, '')) for k in keys), default=0)
        variant_widths.append(max(len(vname), data_w, 4))
    return col_op, col_mf, slot_widths, variant_widths


def _lut_key_row(
    key: Tuple[str, ...],
    col_op: int,
    col_mf: int,
    slot_widths: List[List[int]],
    prefix: str = '  ',
    variant_col_names: Optional[List[str]] = None,
    variant_widths: Optional[List[int]] = None,
) -> str:
    """Format the optype + per-slot + mf + variant-tail portion of a LUT key row."""
    line = f'{prefix}{str(key[0]):<{col_op}}'
    for s, widths in enumerate(slot_widths):
        for fi in range(len(_LUT_SLOT_NAMES)):
            v = _lut_key_slot_field(key, s, fi)
            line += f'  {v:<{widths[fi]}}'
    line += f'  {str(key[8]) if len(key) > 8 else "":<{col_mf}}'
    if variant_col_names:
        tail_dict = dict(_lut_key_variant_tail(key))
        for vname, vw in zip(variant_col_names, variant_widths or []):
            line += f'  {tail_dict.get(vname, ""):<{vw}}'
    return line


def _lut_key_header(
    col_op: int,
    col_mf: int,
    slot_widths: List[List[int]],
    prefix: str = '  ',
    variant_col_names: Optional[List[str]] = None,
    variant_widths: Optional[List[int]] = None,
) -> str:
    """Format the column header for the LUT key columns."""
    hdr = f'{prefix}{"Layer type":<{col_op}}'
    for s, widths in enumerate(slot_widths):
        for fi, name in enumerate(_LUT_SLOT_NAMES):
            suffix = f'_{s}' if s > 0 else ''
            hdr += f'  {(name + suffix):<{widths[fi]}}'
    hdr += f'  {"MF":<{col_mf}}'
    if variant_col_names:
        for vname, vw in zip(variant_col_names, variant_widths or []):
            hdr += f'  {vname:<{vw}}'
    return hdr


def _print_lut_key_summary(
    layers: List[Dict[str, Any]],
    label: str,
    *,
    include_perf: bool,
) -> None:
    """Rollup: count (and optionally ms/LUT) per unique LUT key, columns exploded."""
    by_key = _aggregate_by_lut_key(layers)
    if not by_key:
        print(f'\n(No lut_key data for {label} — profiler CSV, or polaris CSV from a build that emits lut_key columns, required)')
        return

    keys_sorted: List[Tuple[str, ...]] = sorted(
        by_key.keys(),
        key=lambda k: (-by_key[k][1] if include_perf else -by_key[k][0], str(k)),
    )
    max_slots = max((_lut_key_n_slots(k) for k in keys_sorted), default=0)
    has_lut = include_perf and any(by_key[k][2] > 0 for k in keys_sorted)
    variant_col_names = _lut_key_variant_col_names(keys_sorted)
    col_op, col_mf, slot_widths, variant_widths = _lut_key_col_widths(keys_sorted, max_slots, variant_col_names)
    col_c = 6
    col_m = 12
    col_l = 9

    hdr_key = _lut_key_header(col_op, col_mf, slot_widths, variant_col_names=variant_col_names, variant_widths=variant_widths)
    hdr = f'  {"Count":>{col_c}}{hdr_key[1:]}'
    if include_perf:
        hdr += f'  {"Sum ms":>{col_m}}'
        if has_lut:
            hdr += f'  {"LUT":>{col_l}}'
    rule = '─' * max(len(hdr) - 2, 72)

    print(f'\n{"=" * max(len(hdr), 72)}')
    print(f'  Summary by LUT key ({label})')
    print(f'{"=" * max(len(hdr), 72)}')
    print(hdr)
    print(f'  {rule}')

    total_n = 0
    total_ms = 0.0
    total_lut = 0
    for key in keys_sorted:
        n, ms, lut = by_key[key]
        total_n += n
        total_ms += ms
        total_lut += lut
        line = f'  {n:>{col_c}}{_lut_key_row(key, col_op, col_mf, slot_widths, variant_col_names=variant_col_names, variant_widths=variant_widths)[1:]}'
        if include_perf:
            line += f'  {ms:>{col_m}.4f}'
            if has_lut:
                line += f'  {f"{lut}/{n}":>{col_l}}'
        print(line)

    print(f'  {rule}')
    total_line = f'  {total_n:>{col_c}}  {"TOTAL":<{col_op}}'
    for widths in slot_widths:
        for w in widths:
            total_line += f'  {"":>{w}}'
    total_line += f'  {"":>{col_mf}}'
    for vw in variant_widths:
        total_line += f'  {"":>{vw}}'
    if include_perf:
        total_line += f'  {total_ms:>{col_m}.4f}'
        if has_lut:
            total_line += f'  {f"{total_lut}/{total_n}":>{col_l}}'
    print(total_line)
    print()


def _print_lut_key_comparison(
    layers1: List[Dict[str, Any]],
    layers2: List[Dict[str, Any]],
    *,
    label1: str,
    label2: str,
    include_perf: bool,
) -> None:
    """Side-by-side LUT key comparison: count (and optionally ms/gap) per unique key."""
    by1 = _aggregate_by_lut_key(layers1)
    by2 = _aggregate_by_lut_key(layers2)

    if not by1 and not by2:
        print('\n(No lut_key data in either file — profiler CSV, or polaris CSV from a build that emits lut_key columns, required)')
        return
    if not by1:
        print(f'\n(No lut_key data for {label1})')
    if not by2:
        print(f'\n(No lut_key data for {label2})')

    all_keys: List[Tuple[str, ...]] = list(dict.fromkeys(list(by1.keys()) + list(by2.keys())))
    if include_perf:
        all_keys.sort(
            key=lambda k: max(by1.get(k, (0, 0.0, 0))[1], by2.get(k, (0, 0.0, 0))[1]),
            reverse=True,
        )
    else:
        all_keys.sort(
            key=lambda k: -(by1.get(k, (0, 0, 0))[0] + by2.get(k, (0, 0, 0))[0]),
        )

    max_slots = max((_lut_key_n_slots(k) for k in all_keys), default=0)
    variant_col_names = _lut_key_variant_col_names(all_keys)
    col_op, col_mf, slot_widths, variant_widths = _lut_key_col_widths(all_keys, max_slots, variant_col_names)
    total_ms1 = sum(ms for _, ms, _ in by1.values())
    total_ms2 = sum(ms for _, ms, _ in by2.values())
    total_cnt1 = sum(cnt for cnt, _, _ in by1.values())
    total_cnt2 = sum(cnt for cnt, _, _ in by2.values())
    total_lut1 = sum(lut for _, _, lut in by1.values())
    total_lut2 = sum(lut for _, _, lut in by2.values())
    has_lut = (total_lut1 + total_lut2) > 0

    lbl1 = label1[:10]
    lbl2 = label2[:10]
    col_cnt = 7
    col_ms = max(13, len(f'{lbl1}(ms)'), len(f'{lbl2}(ms)'))

    hdr_key = _lut_key_header(col_op, col_mf, slot_widths, variant_col_names=variant_col_names, variant_widths=variant_widths)
    hdr = hdr_key
    hdr += f'  {"#1":>{col_cnt}}'
    if include_perf:
        hdr += f'  {f"{lbl1}(ms)":>{col_ms}}'
    hdr += f'  {"#2":>{col_cnt}}'
    if include_perf:
        hdr += f'  {f"{lbl2}(ms)":>{col_ms}}  {"AbsGap(ms)":>12}  {"Gap%":>9}'
    rule = '─' * max(len(hdr) - 2, 82)

    title_w = max(len(hdr), 82)
    print(f'\n{"=" * title_w}')
    print(f'  LUT key comparison: {label1} vs {label2}')
    print(f'{"=" * title_w}')

    if include_perf:
        print('\n  Network total:')
        print(f'    {label1}:  {total_ms1:.4f} ms')
        print(f'    {label2}:  {total_ms2:.4f} ms')
        print(f'    Gap:  {_pct_gap(total_ms1, total_ms2)} (w.r.t. {label1})')
        if has_lut:
            print(f'    {label1} LUT hits: {total_lut1}/{total_cnt1}')
            print(f'    {label2} LUT hits: {total_lut2}/{total_cnt2}')
        print()

    print(hdr)
    print(f'  {rule}')

    for key in all_keys:
        cnt1, ms1, lut1 = by1.get(key, (0, 0.0, 0))
        cnt2, ms2, lut2 = by2.get(key, (0, 0.0, 0))
        line = _lut_key_row(key, col_op, col_mf, slot_widths, variant_col_names=variant_col_names, variant_widths=variant_widths)
        cnt1_s = str(cnt1) if cnt1 else '—'
        cnt2_s = str(cnt2) if cnt2 else '—'
        line += f'  {cnt1_s:>{col_cnt}}'
        if include_perf:
            ms1_s = f'{ms1:.4f}' if cnt1 else '—'
            line += f'  {ms1_s:>{col_ms}}'
        line += f'  {cnt2_s:>{col_cnt}}'
        if include_perf:
            ms2_s = f'{ms2:.4f}' if cnt2 else '—'
            abs_gap_s = f'{ms2 - ms1:+.4f}' if (cnt1 and cnt2) else '—'
            line += f'  {ms2_s:>{col_ms}}  {abs_gap_s:>12}  {_pct_gap(ms1, ms2):>9}'
        print(line)

    print(f'  {rule}')
    total_line = f'  {"TOTAL":<{col_op}}'
    for widths in slot_widths:
        for w in widths:
            total_line += f'  {"":>{w}}'
    total_line += f'  {"":>{col_mf}}'
    for vw in variant_widths:
        total_line += f'  {"":>{vw}}'
    total_line += f'  {total_cnt1:>{col_cnt}}'
    if include_perf:
        total_line += f'  {total_ms1:>{col_ms}.4f}'
    total_line += f'  {total_cnt2:>{col_cnt}}'
    if include_perf:
        abs_total = total_ms2 - total_ms1
        total_line += f'  {total_ms2:>{col_ms}.4f}  {abs_total:>+12.4f}  {_pct_gap(total_ms1, total_ms2):>9}'
    print(total_line)
    print()


# ---------------------------------------------------------------------------
# XLSX report (three sheets: Summary, By Layer Type, By Layer Signature)
# ---------------------------------------------------------------------------

def _write_xlsx_report(
    path: str,
    *,
    layers1: Optional[List[Dict[str, Any]]],
    layers2: Optional[List[Dict[str, Any]]],
    stats: Optional[ComparisonStats],
    label1: str,
    label2: str,
    source1: Optional[str],
    source2: Optional[str],
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> None:
    """Write a 4-sheet XLSX comparison report.

    Sheets:
      1. "Summary"          — model-wide totals, comparison, shape/attr stats.
      2. "By Layer Type"    — per canonical optype rollup with comparison.
      3. "By Layer Signature" — per (optype + normalized shape signature).
      4. "By LUT Key"       — per full LUT key tuple with columns exploded.

    In two-file mode (both ``layers1`` and ``layers2`` provided) each sheet
    includes File2 vs File1 columns and the absolute / percent gap.
    In single-file mode only the side that was loaded is emitted.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover - import guarded at call site
        raise RuntimeError(
            "openpyxl is required for --xlsx output. Install it via "
            "`pip install openpyxl` (already present in the polarisdev env)."
        ) from e

    have1 = layers1 is not None
    have2 = layers2 is not None
    two_sided = have1 and have2

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    total_font = Font(bold=True, italic=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _style_header(ws, row_idx: int, n_cols: int) -> None:
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    def _style_total(ws, row_idx: int, n_cols: int) -> None:
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = total_font
            cell.fill = total_fill

    def _autosize(ws, n_cols: int, max_width: int = 60) -> None:
        for c in range(1, n_cols + 1):
            letter = get_column_letter(c)
            best = 8
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
                v = row[0]
                if v is None:
                    continue
                if isinstance(v, float):
                    s = f"{v:.4f}"
                else:
                    s = str(v)
                if len(s) > best:
                    best = len(s)
            ws.column_dimensions[letter].width = min(best + 2, max_width)

    def _pct(reference: float, other: float) -> Optional[float]:
        if reference == 0.0:
            return None
        return (other - reference) / reference * 100.0

    wb = Workbook()

    # ------------------ Sheet 1: Summary ------------------
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Summary"

    ws1.cell(row=1, column=1, value="Compare Layers — XLSX Report").font = Font(bold=True, size=13)
    r = 2
    if source2:
        ws1.cell(row=r, column=1, value=f"{label2} CSV"); ws1.cell(row=r, column=2, value=source2); r += 1
    if source1:
        ws1.cell(row=r, column=1, value=f"{label1} CSV"); ws1.cell(row=r, column=2, value=source1); r += 1
    ws1.cell(row=r, column=1, value="strip_leading_ones"); ws1.cell(row=r, column=2, value=bool(strip_leading_ones)); r += 1
    ws1.cell(row=r, column=1, value="strip_singleton_dims"); ws1.cell(row=r, column=2, value=bool(strip_singleton_dims)); r += 1
    r += 1

    # Network totals
    ws1.cell(row=r, column=1, value="Network totals").font = Font(bold=True); r += 1
    if two_sided:
        headers = ["Metric", label1, label2, "Abs Gap", "Gap %"]
    elif have2:
        headers = ["Metric", label2]
    else:
        headers = ["Metric", label1]
    for c, h in enumerate(headers, start=1):
        ws1.cell(row=r, column=c, value=h)
    _style_header(ws1, r, len(headers))
    r += 1

    def _layer_totals(layers: List[Dict[str, Any]]) -> Tuple[int, float, int]:
        cnt = len(layers)
        ms = sum((float(l["duration_ms"]) for l in layers if l.get("duration_ms") is not None), 0.0)
        lut = sum(1 for l in layers if l.get("uses_perf_lookup"))
        return cnt, ms, lut

    cnt2 = ms2 = lut2 = 0
    cnt1 = ms1 = lut1 = 0
    if have2:
        cnt2, ms2, lut2 = _layer_totals(layers2)  # type: ignore[assignment,arg-type]
    if have1:
        cnt1, ms1, lut1 = _layer_totals(layers1)  # type: ignore[assignment,arg-type]

    def _write_metric(metric: str, v2, v1, *, fmt_ms: bool = False, percent: bool = False) -> None:
        nonlocal r
        # Column order matches the docstring convention: gap = (file2 − file1) / file1.
        # Header is [Metric, label1, label2, "Abs Gap", "Gap %"] so column 2 holds
        # label1's value (v1) and column 3 holds label2's value (v2).
        ws1.cell(row=r, column=1, value=metric)
        if two_sided:
            ws1.cell(row=r, column=2, value=v1)
            ws1.cell(row=r, column=3, value=v2)
            if isinstance(v2, (int, float)) and isinstance(v1, (int, float)):
                ws1.cell(row=r, column=4, value=v2 - v1)
                p = _pct(float(v1), float(v2))
                ws1.cell(row=r, column=5, value=(p if p is not None else "N/A"))
                if fmt_ms:
                    ws1.cell(row=r, column=2).number_format = "0.0000"
                    ws1.cell(row=r, column=3).number_format = "0.0000"
                    ws1.cell(row=r, column=4).number_format = "+0.0000;-0.0000"
                if p is not None:
                    ws1.cell(row=r, column=5).number_format = "+0.00\"%\";-0.00\"%\""
        elif have2:
            ws1.cell(row=r, column=2, value=v2)
            if fmt_ms and isinstance(v2, (int, float)):
                ws1.cell(row=r, column=2).number_format = "0.0000"
        else:
            ws1.cell(row=r, column=2, value=v1)
            if fmt_ms and isinstance(v1, (int, float)):
                ws1.cell(row=r, column=2).number_format = "0.0000"
        r += 1

    _write_metric("Total layers", cnt2, cnt1)
    _write_metric("Total duration (ms)", ms2, ms1, fmt_ms=True)
    if have1:
        # layers1-only metric — display only on layers1 column when single-sided.
        miss1 = cnt1 - lut1
        if two_sided:
            ws1.cell(row=r, column=1, value=f"{label1} LUT hits (count / total)")
            ws1.cell(row=r, column=2, value="—")
            ws1.cell(row=r, column=3, value=f"{lut1} / {cnt1}")
            r += 1
            ws1.cell(row=r, column=1, value=f"{label1} LUT misses (count / total)")
            ws1.cell(row=r, column=2, value="—")
            ws1.cell(row=r, column=3, value=f"{miss1} / {cnt1}")
            r += 1
        else:
            ws1.cell(row=r, column=1, value=f"{label1} LUT hits (count / total)")
            ws1.cell(row=r, column=2, value=f"{lut1} / {cnt1}")
            r += 1
            ws1.cell(row=r, column=1, value=f"{label1} LUT misses (count / total)")
            ws1.cell(row=r, column=2, value=f"{miss1} / {cnt1}")
            r += 1

    # Shape/attr stats (two-file only)
    if two_sided and stats is not None:
        r += 1
        ws1.cell(row=r, column=1, value="Shape / attribute comparison").font = Font(bold=True); r += 1
        for c, h in enumerate(["Metric", "Count"], start=1):
            ws1.cell(row=r, column=c, value=h)
        _style_header(ws1, r, 2); r += 1
        for label, val in [
            ("Total matches", stats.total_matches),
            ("Name mismatches", stats.name_mismatches),
            ("Shape mismatches", stats.shape_mismatches),
            ("  input shape mismatches", stats.input_shape_mismatches),
            ("  output shape mismatches", stats.output_shape_mismatches),
            ("Attribute mismatches", stats.attr_mismatches),
            ("LUT key mismatches", stats.lut_key_mismatches),
            (f"Unmatched ({label1})", stats.unmatched_polaris),
            (f"Unmatched ({label2})", stats.unmatched_profiler),
            ("Ambiguous", stats.ambiguous),
        ]:
            ws1.cell(row=r, column=1, value=label)
            ws1.cell(row=r, column=2, value=val)
            r += 1

    _autosize(ws1, 5 if two_sided else 2)

    # ---------- Sheet 2: By Layer Type ----------
    ws2 = wb.create_sheet("By Layer Type")
    by_op2: Dict[str, Tuple[int, float, int]] = (
        _aggregate_duration_by_optype(layers2) if have2 else {}  # type: ignore[arg-type]
    )
    by_op1: Dict[str, Tuple[int, float, int]] = (
        _aggregate_duration_by_optype(layers1) if have1 else {}  # type: ignore[arg-type]
    )
    keys_op = list(dict.fromkeys(list(by_op2.keys()) + list(by_op1.keys())))
    keys_op.sort(
        key=lambda k: max(by_op2.get(k, (0, 0.0, 0))[1], by_op1.get(k, (0, 0.0, 0))[1]),
        reverse=True,
    )
    if two_sided:
        # Column order: label1 first, label2 second (matches docstring "gap = (file2 − file1) / file1").
        op_headers = [
            "Layer Type",
            f"# {label1}", f"{label1} ms",
            f"# {label2}", f"{label2} ms",
            f"{label1} LUT hit", f"{label1} LUT miss", f"{label1} LUT total",
            f"Only in {label1}", f"Only in {label2}",
            "Abs Gap (ms)", "Gap %",
        ]
    elif have2:
        op_headers = ["Layer Type", f"# {label2}", f"{label2} ms"]
    else:
        op_headers = [
            "Layer Type", f"# {label1}", f"{label1} ms",
            f"{label1} LUT hit", f"{label1} LUT miss", f"{label1} LUT total",
        ]
    for c, h in enumerate(op_headers, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, 1, len(op_headers))

    rr = 2
    only1_op_total = 0
    only2_op_total = 0
    for k in keys_op:
        cnt2_row, ms2_row, _ = by_op2.get(k, (0, 0.0, 0))
        cnt1_row, ms1_row, lut1_row = by_op1.get(k, (0, 0.0, 0))
        miss1_row = cnt1_row - lut1_row
        only1 = max(0, cnt1_row - cnt2_row)
        only2 = max(0, cnt2_row - cnt1_row)
        only1_op_total += only1
        only2_op_total += only2
        if two_sided:
            ws2.cell(row=rr, column=1, value=k)
            ws2.cell(row=rr, column=2, value=cnt1_row)
            ws2.cell(row=rr, column=3, value=ms1_row).number_format = "0.0000"
            ws2.cell(row=rr, column=4, value=cnt2_row)
            ws2.cell(row=rr, column=5, value=ms2_row).number_format = "0.0000"
            ws2.cell(row=rr, column=6, value=lut1_row)
            ws2.cell(row=rr, column=7, value=miss1_row)
            ws2.cell(row=rr, column=8, value=cnt1_row)
            ws2.cell(row=rr, column=9, value=only1)
            ws2.cell(row=rr, column=10, value=only2)
            ws2.cell(row=rr, column=11, value=(ms2_row - ms1_row)).number_format = "+0.0000;-0.0000"
            p = _pct(ms1_row, ms2_row)
            cell = ws2.cell(row=rr, column=12, value=(p if p is not None else "N/A"))
            if p is not None:
                cell.number_format = "+0.00\"%\";-0.00\"%\""
        elif have2:
            ws2.cell(row=rr, column=1, value=k)
            ws2.cell(row=rr, column=2, value=cnt2_row)
            ws2.cell(row=rr, column=3, value=ms2_row).number_format = "0.0000"
        else:
            ws2.cell(row=rr, column=1, value=k)
            ws2.cell(row=rr, column=2, value=cnt1_row)
            ws2.cell(row=rr, column=3, value=ms1_row).number_format = "0.0000"
            ws2.cell(row=rr, column=4, value=lut1_row)
            ws2.cell(row=rr, column=5, value=miss1_row)
            ws2.cell(row=rr, column=6, value=cnt1_row)
        rr += 1

    # TOTAL row
    miss1_total = cnt1 - lut1
    ws2.cell(row=rr, column=1, value="TOTAL")
    if two_sided:
        ws2.cell(row=rr, column=2, value=cnt1)
        ws2.cell(row=rr, column=3, value=ms1).number_format = "0.0000"
        ws2.cell(row=rr, column=4, value=cnt2)
        ws2.cell(row=rr, column=5, value=ms2).number_format = "0.0000"
        ws2.cell(row=rr, column=6, value=lut1)
        ws2.cell(row=rr, column=7, value=miss1_total)
        ws2.cell(row=rr, column=8, value=cnt1)
        ws2.cell(row=rr, column=9, value=only1_op_total)
        ws2.cell(row=rr, column=10, value=only2_op_total)
        ws2.cell(row=rr, column=11, value=(ms2 - ms1)).number_format = "+0.0000;-0.0000"
        p = _pct(ms1, ms2)
        cell = ws2.cell(row=rr, column=12, value=(p if p is not None else "N/A"))
        if p is not None:
            cell.number_format = "+0.00\"%\";-0.00\"%\""
    elif have2:
        ws2.cell(row=rr, column=2, value=cnt2)
        ws2.cell(row=rr, column=3, value=ms2).number_format = "0.0000"
    else:
        ws2.cell(row=rr, column=2, value=cnt1)
        ws2.cell(row=rr, column=3, value=ms1).number_format = "0.0000"
        ws2.cell(row=rr, column=4, value=lut1)
        ws2.cell(row=rr, column=5, value=miss1_total)
        ws2.cell(row=rr, column=6, value=cnt1)
    _style_total(ws2, rr, len(op_headers))
    ws2.freeze_panes = "A2"
    _autosize(ws2, len(op_headers))

    # ---------- Sheet 3: By Layer Signature ----------
    ws3 = wb.create_sheet("By Layer Signature")
    by_sig2: Dict[Tuple[str, str], Tuple[int, float, int]] = (
        _aggregate_duration_by_optype_signature(
            layers2, strip_leading_ones, strip_singleton_dims  # type: ignore[arg-type]
        ) if have2 else {}
    )
    by_sig1: Dict[Tuple[str, str], Tuple[int, float, int]] = (
        _aggregate_duration_by_optype_signature(
            layers1, strip_leading_ones, strip_singleton_dims  # type: ignore[arg-type]
        ) if have1 else {}
    )
    keys_sig = list(dict.fromkeys(list(by_sig2.keys()) + list(by_sig1.keys())))
    keys_sig.sort(
        key=lambda k: max(
            by_sig2.get(k, (0, 0.0, 0))[1], by_sig1.get(k, (0, 0.0, 0))[1]
        ),
        reverse=True,
    )
    if two_sided:
        # Column order: label1 first, label2 second (matches docstring "gap = (file2 − file1) / file1").
        sig_headers = [
            "Layer Type", "Signature",
            f"# {label1}", f"{label1} ms",
            f"# {label2}", f"{label2} ms",
            f"{label1} LUT hit", f"{label1} LUT miss", f"{label1} LUT total",
            f"Only in {label1}", f"Only in {label2}",
            "Abs Gap (ms)", "Gap %",
        ]
    elif have2:
        sig_headers = ["Layer Type", "Signature", f"# {label2}", f"{label2} ms"]
    else:
        sig_headers = [
            "Layer Type", "Signature",
            f"# {label1}", f"{label1} ms",
            f"{label1} LUT hit", f"{label1} LUT miss", f"{label1} LUT total",
        ]
    for c, h in enumerate(sig_headers, start=1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, 1, len(sig_headers))

    rr = 2
    only1_sig_total = 0
    only2_sig_total = 0
    for sig_key in keys_sig:
        op, sig = sig_key
        cnt2_row, ms2_row, _ = by_sig2.get(sig_key, (0, 0.0, 0))
        cnt1_row, ms1_row, lut1_row = by_sig1.get(sig_key, (0, 0.0, 0))
        miss1_row = cnt1_row - lut1_row
        only1 = max(0, cnt1_row - cnt2_row)
        only2 = max(0, cnt2_row - cnt1_row)
        only1_sig_total += only1
        only2_sig_total += only2
        ws3.cell(row=rr, column=1, value=op)
        ws3.cell(row=rr, column=2, value=sig).alignment = left
        if two_sided:
            ws3.cell(row=rr, column=3, value=cnt1_row)
            ws3.cell(row=rr, column=4, value=ms1_row).number_format = "0.0000"
            ws3.cell(row=rr, column=5, value=cnt2_row)
            ws3.cell(row=rr, column=6, value=ms2_row).number_format = "0.0000"
            ws3.cell(row=rr, column=7, value=lut1_row)
            ws3.cell(row=rr, column=8, value=miss1_row)
            ws3.cell(row=rr, column=9, value=cnt1_row)
            ws3.cell(row=rr, column=10, value=only1)
            ws3.cell(row=rr, column=11, value=only2)
            ws3.cell(row=rr, column=12, value=(ms2_row - ms1_row)).number_format = "+0.0000;-0.0000"
            p = _pct(ms1_row, ms2_row)
            cell = ws3.cell(row=rr, column=13, value=(p if p is not None else "N/A"))
            if p is not None:
                cell.number_format = "+0.00\"%\";-0.00\"%\""
        elif have2:
            ws3.cell(row=rr, column=3, value=cnt2_row)
            ws3.cell(row=rr, column=4, value=ms2_row).number_format = "0.0000"
        else:
            ws3.cell(row=rr, column=3, value=cnt1_row)
            ws3.cell(row=rr, column=4, value=ms1_row).number_format = "0.0000"
            ws3.cell(row=rr, column=5, value=lut1_row)
            ws3.cell(row=rr, column=6, value=miss1_row)
            ws3.cell(row=rr, column=7, value=cnt1_row)
        rr += 1

    miss1_total = cnt1 - lut1
    ws3.cell(row=rr, column=1, value="TOTAL")
    ws3.cell(row=rr, column=2, value="")
    if two_sided:
        ws3.cell(row=rr, column=3, value=cnt1)
        ws3.cell(row=rr, column=4, value=ms1).number_format = "0.0000"
        ws3.cell(row=rr, column=5, value=cnt2)
        ws3.cell(row=rr, column=6, value=ms2).number_format = "0.0000"
        ws3.cell(row=rr, column=7, value=lut1)
        ws3.cell(row=rr, column=8, value=miss1_total)
        ws3.cell(row=rr, column=9, value=cnt1)
        ws3.cell(row=rr, column=10, value=only1_sig_total)
        ws3.cell(row=rr, column=11, value=only2_sig_total)
        ws3.cell(row=rr, column=12, value=(ms2 - ms1)).number_format = "+0.0000;-0.0000"
        p = _pct(ms1, ms2)
        cell = ws3.cell(row=rr, column=13, value=(p if p is not None else "N/A"))
        if p is not None:
            cell.number_format = "+0.00\"%\";-0.00\"%\""
    elif have2:
        ws3.cell(row=rr, column=3, value=cnt2)
        ws3.cell(row=rr, column=4, value=ms2).number_format = "0.0000"
    else:
        ws3.cell(row=rr, column=3, value=cnt1)
        ws3.cell(row=rr, column=4, value=ms1).number_format = "0.0000"
        ws3.cell(row=rr, column=5, value=lut1)
        ws3.cell(row=rr, column=6, value=miss1_total)
        ws3.cell(row=rr, column=7, value=cnt1)
    _style_total(ws3, rr, len(sig_headers))
    ws3.freeze_panes = "C2"
    _autosize(ws3, len(sig_headers))

    # ---------- Sheet 4: By LUT Key ----------
    ws4 = wb.create_sheet("By LUT Key")
    by_lk1: Dict[Tuple[str, ...], Tuple[int, float, int]] = (
        _aggregate_by_lut_key(layers1) if have1 else {}  # type: ignore[arg-type]
    )
    by_lk2: Dict[Tuple[str, ...], Tuple[int, float, int]] = (
        _aggregate_by_lut_key(layers2) if have2 else {}  # type: ignore[arg-type]
    )
    all_lk: List[Tuple[str, ...]] = list(dict.fromkeys(list(by_lk2.keys()) + list(by_lk1.keys())))
    all_lk.sort(
        key=lambda k: max(by_lk2.get(k, (0, 0.0, 0))[1], by_lk1.get(k, (0, 0.0, 0))[1]),
        reverse=True,
    )

    if not all_lk:
        ws4.cell(row=1, column=1, value="(No lut_key data in either file — polaris CSV with lut_key columns required)")
    else:
        max_slots_lk = max((_lut_key_n_slots(k) for k in all_lk), default=0)
        variant_col_names_lk = _lut_key_variant_col_names(all_lk)
        mf_col = 2 + max_slots_lk * len(_LUT_SLOT_NAMES)  # op_type + slots + mf
        n_key_cols = mf_col + len(variant_col_names_lk)

        lk_headers: List[str] = ['op_type']
        for s in range(max_slots_lk):
            for fi, fname in enumerate(_LUT_SLOT_NAMES):
                lk_headers.append(fname + (f'_{s}' if s > 0 else ''))
        lk_headers.append('mf')
        lk_headers.extend(variant_col_names_lk)
        if two_sided:
            lk_headers += [
                f'# {label1}', f'{label1} ms',
                f'# {label2}', f'{label2} ms',
                f'{label1} LUT hit', f'{label1} LUT miss', f'{label1} LUT total',
                'Abs Gap (ms)', 'Gap %',
            ]
        elif have1:
            lk_headers += [
                f'# {label1}', f'{label1} ms',
                f'{label1} LUT hit', f'{label1} LUT miss', f'{label1} LUT total',
            ]
        else:
            lk_headers += [f'# {label2}', f'{label2} ms']

        for c, h in enumerate(lk_headers, start=1):
            ws4.cell(row=1, column=c, value=h)
        _style_header(ws4, 1, len(lk_headers))

        rr = 2
        lk_cnt1_total = lk_ms1_total = lk_cnt2_total = lk_ms2_total = 0.0
        lk_lut1_total = 0
        for lk in all_lk:
            ws4.cell(row=rr, column=1, value=lk[0])
            for s in range(max_slots_lk):
                for fi in range(len(_LUT_SLOT_NAMES)):
                    ws4.cell(row=rr, column=2 + s * len(_LUT_SLOT_NAMES) + fi,
                             value=_lut_key_slot_field(lk, s, fi))
            ws4.cell(row=rr, column=mf_col, value=lk[8] if len(lk) > 8 else '')
            lk_tail = dict(_lut_key_variant_tail(lk))
            for vi, vname in enumerate(variant_col_names_lk):
                ws4.cell(row=rr, column=mf_col + 1 + vi, value=lk_tail.get(vname, ''))

            cnt2_row, ms2_row, _ = by_lk2.get(lk, (0, 0.0, 0))
            cnt1_row, ms1_row, lut1_row = by_lk1.get(lk, (0, 0.0, 0))
            miss1_row = cnt1_row - lut1_row
            lk_cnt1_total += cnt1_row; lk_ms1_total += ms1_row
            lk_cnt2_total += cnt2_row; lk_ms2_total += ms2_row
            lk_lut1_total += lut1_row

            col = n_key_cols + 1
            if two_sided:
                ws4.cell(row=rr, column=col, value=cnt1_row); col += 1
                ws4.cell(row=rr, column=col, value=ms1_row).number_format = "0.0000"; col += 1
                ws4.cell(row=rr, column=col, value=cnt2_row); col += 1
                ws4.cell(row=rr, column=col, value=ms2_row).number_format = "0.0000"; col += 1
                ws4.cell(row=rr, column=col, value=lut1_row); col += 1
                ws4.cell(row=rr, column=col, value=miss1_row); col += 1
                ws4.cell(row=rr, column=col, value=cnt1_row); col += 1
                ws4.cell(row=rr, column=col, value=(ms2_row - ms1_row)).number_format = "+0.0000;-0.0000"; col += 1
                p = _pct(ms1_row, ms2_row)
                cell = ws4.cell(row=rr, column=col, value=(p if p is not None else "N/A"))
                if p is not None:
                    cell.number_format = "+0.00\"%\";-0.00\"%\""
            elif have1:
                ws4.cell(row=rr, column=col, value=cnt1_row); col += 1
                ws4.cell(row=rr, column=col, value=ms1_row).number_format = "0.0000"; col += 1
                ws4.cell(row=rr, column=col, value=lut1_row); col += 1
                ws4.cell(row=rr, column=col, value=miss1_row); col += 1
                ws4.cell(row=rr, column=col, value=cnt1_row)
            else:
                ws4.cell(row=rr, column=col, value=cnt2_row); col += 1
                ws4.cell(row=rr, column=col, value=ms2_row).number_format = "0.0000"
            rr += 1

        # TOTAL row
        lk_miss1_total = int(lk_cnt1_total) - lk_lut1_total
        ws4.cell(row=rr, column=1, value="TOTAL")
        col = n_key_cols + 1
        if two_sided:
            ws4.cell(row=rr, column=col, value=int(lk_cnt1_total)); col += 1
            ws4.cell(row=rr, column=col, value=lk_ms1_total).number_format = "0.0000"; col += 1
            ws4.cell(row=rr, column=col, value=int(lk_cnt2_total)); col += 1
            ws4.cell(row=rr, column=col, value=lk_ms2_total).number_format = "0.0000"; col += 1
            ws4.cell(row=rr, column=col, value=lk_lut1_total); col += 1
            ws4.cell(row=rr, column=col, value=lk_miss1_total); col += 1
            ws4.cell(row=rr, column=col, value=int(lk_cnt1_total)); col += 1
            ws4.cell(row=rr, column=col, value=(lk_ms2_total - lk_ms1_total)).number_format = "+0.0000;-0.0000"; col += 1
            p = _pct(lk_ms1_total, lk_ms2_total)
            cell = ws4.cell(row=rr, column=col, value=(p if p is not None else "N/A"))
            if p is not None:
                cell.number_format = "+0.00\"%\";-0.00\"%\""
        elif have1:
            ws4.cell(row=rr, column=col, value=int(lk_cnt1_total)); col += 1
            ws4.cell(row=rr, column=col, value=lk_ms1_total).number_format = "0.0000"; col += 1
            ws4.cell(row=rr, column=col, value=lk_lut1_total); col += 1
            ws4.cell(row=rr, column=col, value=lk_miss1_total); col += 1
            ws4.cell(row=rr, column=col, value=int(lk_cnt1_total))
        else:
            ws4.cell(row=rr, column=col, value=int(lk_cnt2_total)); col += 1
            ws4.cell(row=rr, column=col, value=lk_ms2_total).number_format = "0.0000"
        _style_total(ws4, rr, len(lk_headers))
        ws4.freeze_panes = f"{get_column_letter(n_key_cols + 1)}2"
        _autosize(ws4, len(lk_headers))

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    """Main entry point."""
    args = parse_args()

    # Sanitize and validate file paths to prevent path traversal attacks
    try:
        file1_path = sanitize_file_path(args.file1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    file2_path = None
    if args.file2 is not None:
        try:
            file2_path = sanitize_file_path(args.file2)
        except ValueError as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1

    # --- Standalone mode (single file + --perf / --summarize-by-signature / --by-lut-key) ---
    if file2_path is None:
        if not args.perf and not args.summarize_by_signature and not args.by_lut_key:
            print("Error: Two files are required for shape comparison. "
                  "Use --perf, --summarize-by-signature, or --by-lut-key with a single file.",
                  file=sys.stderr)
            return 1

        ftype = detect_file_type(str(file1_path))
        if ftype is None:
            print("Error: Could not determine file type. Expected CSV with "
                  "'archname' (polaris) or 'OP CODE' (profiler) columns.",
                  file=sys.stderr)
            return 1

        try:
            if ftype == 'polaris':
                layers = layers_polaris(str(file1_path))
                label = "Polaris"
            else:
                layers = layers_profiler(str(file1_path))
                label = "Profiler"
        except Exception as e:
            print(f"Error extracting layers: {e}", file=sys.stderr)
            return 1

        print(f"{label} CSV: {file1_path}")
        print(f"Loaded {len(layers)} {label.lower()} layers")

        if ftype == 'profiler':
            layers = _maybe_dedup_profiler_layers(layers, str(file1_path))

        if args.filter_optype:
            filter_norm = normalize_optype(args.filter_optype)
            layers = [layer for layer in layers if normalize_optype(layer['optype']) == filter_norm]
            print(f"Filtered to {len(layers)} layers with optype='{args.filter_optype}'")

        if args.summarize_by_signature:
            _print_signature_summary(
                layers,
                label,
                args.strip_leading_ones,
                args.strip_singleton_dims,
                include_perf=args.perf,
            )
        if args.by_lut_key:
            _print_lut_key_summary(layers, label, include_perf=args.perf)
        if args.perf:
            if args.summarize_by_signature:
                _print_perf_standalone_by_signature(
                    layers,
                    label,
                    args.strip_leading_ones,
                    args.strip_singleton_dims,
                )
            elif not args.by_lut_key:
                _print_perf_standalone(layers, label)

        if args.xlsx:
            try:
                _write_xlsx_report(
                    args.xlsx,
                    layers1=layers if ftype == 'polaris' else None,
                    layers2=layers if ftype == 'profiler' else None,
                    stats=None,
                    label1="Polaris",
                    label2="Profiler",
                    source1=args.file1 if ftype == 'polaris' else None,
                    source2=args.file1 if ftype == 'profiler' else None,
                    strip_leading_ones=args.strip_leading_ones,
                    strip_singleton_dims=args.strip_singleton_dims,
                )
                print(f"\nWrote XLSX report: {args.xlsx}")
            except Exception as e:
                print(f"Error writing XLSX report: {e}", file=sys.stderr)
                return 1
        return 0

    # --- Two-file mode ---
    type1 = detect_file_type(str(file1_path))
    type2 = detect_file_type(str(file2_path))

    if type1 is None or type2 is None:
        print("Error: Could not determine file types. Expected CSV files with "
              "'archname' (polaris) or 'OP CODE' (profiler) columns.", file=sys.stderr)
        return 1

    # Determine loader for each file based on its detected type
    loader1 = layers_polaris if type1 == 'polaris' else layers_profiler
    loader2 = layers_polaris if type2 == 'polaris' else layers_profiler

    # Set default display labels; user overrides via --label1/--label2
    if args.label1:
        label1 = args.label1
    else:
        label1 = f"{type1.capitalize()} 1" if type1 == type2 else type1.capitalize()
    if args.label2:
        label2 = args.label2
    else:
        label2 = f"{type2.capitalize()} 2" if type1 == type2 else type2.capitalize()

    print(f"File 1 ({label1}): {file1_path}")
    print(f"File 2 ({label2}): {file2_path}")
    print()

    # Extract layers
    try:
        layers1 = loader1(str(file1_path))
        layers2 = loader2(str(file2_path))
    except Exception as e:
        print(f"Error extracting layers: {e}", file=sys.stderr)
        return 1

    print(f"Loaded {len(layers1)} {label1} layers, {len(layers2)} {label2} layers")

    if type1 == 'profiler':
        print(f'\n[{label1}] ', end='')
        layers1 = _maybe_dedup_profiler_layers(layers1, str(file1_path))
    if type2 == 'profiler':
        print(f'\n[{label2}] ', end='')
        layers2 = _maybe_dedup_profiler_layers(layers2, str(file2_path))

    # Filter by optype if requested
    if args.filter_optype:
        filter_optype_norm = normalize_optype(args.filter_optype)

        layers1 = [
            layer for layer in layers1
            if normalize_optype(layer['optype']) == filter_optype_norm
        ]
        layers2 = [
            layer for layer in layers2
            if normalize_optype(layer['optype']) == filter_optype_norm
        ]

        print(f"Filtered to {len(layers1)} {label1} layers, {len(layers2)} {label2} layers with optype='{args.filter_optype}'")

        if len(layers1) == 0 and len(layers2) == 0:
            print(f"Warning: No layers found with optype='{args.filter_optype}'")
            return 0

    print()

    # Compare layers (shape / attribute matching)
    stats = compare_layers(layers1, layers2, args.max_search_distance,
                           args.strip_leading_ones, args.strip_singleton_dims,
                           ignore_attrs=args.ignore_attrs,
                           label1=label1, label2=label2)

    # Print shape-comparison summary
    print_summary(stats, label1=label1, label2=label2)

    if args.summarize_by_signature:
        _print_signature_summary(
            layers1,
            label1,
            args.strip_leading_ones,
            args.strip_singleton_dims,
            include_perf=args.perf,
        )
        _print_signature_summary(
            layers2,
            label2,
            args.strip_leading_ones,
            args.strip_singleton_dims,
            include_perf=args.perf,
        )

    if args.by_lut_key:
        _print_lut_key_comparison(
            layers1,
            layers2,
            label1=label1,
            label2=label2,
            include_perf=args.perf,
        )

    # Performance comparison (when --perf is enabled)
    if args.perf:
        _print_perf_comparison(
            layers1,
            layers2,
            by_signature=args.summarize_by_signature,
            strip_leading_ones=args.strip_leading_ones,
            strip_singleton_dims=args.strip_singleton_dims,
            label1=label1,
            label2=label2,
        )

    if args.xlsx:
        try:
            _write_xlsx_report(
                args.xlsx,
                layers1=layers1,
                layers2=layers2,
                stats=stats,
                label1=label1,
                label2=label2,
                source1=str(file1_path),
                source2=str(file2_path),
                strip_leading_ones=args.strip_leading_ones,
                strip_singleton_dims=args.strip_singleton_dims,
            )
            print(f"\nWrote XLSX report: {args.xlsx}")
        except Exception as e:
            print(f"Error writing XLSX report: {e}", file=sys.stderr)
            return 1

    return 0


if __name__ == '__main__':
    sys.exit(main())
